from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import io
from django.views.decorators.http import require_GET
@require_GET
def export_merit_list_pdf(request):
    """
    Export the new CBC merit list as a simple PDF (total points, average level per student).
    """
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()
    if is_teacher and not request.user.teacher.is_class_teacher:
        return HttpResponseForbidden('Merit lists are available to class teachers only.')
    class_id = request.GET.get('class_id')
    exam_id = request.GET.get('exam_id')
    stream_id = request.GET.get('stream_id')
    term = (request.GET.get('term') or '').strip()

    from .utils.grading import get_level_and_points_for_score
    from .models import Student, Subject, MarkSheet, StudentMark, ClassRoom, Exam
    from collections import Counter
    from django.shortcuts import get_object_or_404

    classroom = get_object_or_404(ClassRoom, id=class_id, school=school)
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
            teacher, school, cast(Any, classroom).id
        )
        if not allowed:
            return HttpResponseForbidden('Not allowed for this class.')
        if stream_id and not all_streams:
            try:
                if int(stream_id) not in allowed_stream_ids:
                    return HttpResponseForbidden('Not allowed for selected stream.')
            except (TypeError, ValueError):
                return HttpResponse('Invalid stream_id', status=400)
    exam = get_object_or_404(Exam, id=exam_id, school=school)
    students = Student.objects.filter(classroom=classroom, school=school)
    if stream_id:
        students = students.filter(stream_id=stream_id)
    elif is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        _allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
            teacher, school, cast(Any, classroom).id
        )
        if not all_streams and allowed_stream_ids:
            students = students.filter(stream_id__in=allowed_stream_ids)
    subjects = Subject.objects.filter(subjectallocation__class_id=class_id, school=school).distinct()
    mark_sheets = MarkSheet.objects.filter(school_class=classroom, exam_id=exam_id, status='published')
    if term:
        mark_sheets = mark_sheets.filter(term=term)
    mark_sheets = mark_sheets.select_related('subject')
    marks_map = {}
    for mark in StudentMark.objects.filter(marksheet__in=mark_sheets, score__isnull=False):
        marks_map[(mark.student_id, mark.marksheet.subject_id)] = (mark.score, mark.marksheet.out_of)
    merit_list = []
    for student in students:
        total_points = 0
        level_list = []
        for subject in subjects:
            key = (student.id, subject.id)
            if key not in marks_map:
                continue
            score, out_of = marks_map[key]
            level, points = get_level_and_points_for_score(
                student=student,
                subject=subject,
                score=score,
                out_of=out_of,
                term=term
            )
            if points:
                total_points += points
            if level:
                level_list.append(level)
        average_level = Counter(level_list).most_common(1)[0][0] if level_list else "-"
        merit_list.append({
            "adm": student.admission_number,
            "name": student.user.get_full_name() if hasattr(student, 'user') else str(student),
            "stream": student.stream.name if student.stream else "",
            "total_points": total_points,
            "average_level": average_level
        })
    merit_list.sort(key=lambda x: x["total_points"], reverse=True)

    # PDF generation
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 40
    p.setFont("Helvetica-Bold", 16)
    p.drawString(40, y, f"Merit List: {classroom.name} - {exam.title}")
    y -= 30
    p.setFont("Helvetica", 10)
    p.drawString(40, y, f"Generated: {term or ''}")
    y -= 30
    p.setFont("Helvetica-Bold", 11)
    p.drawString(40, y, "Pos")
    p.drawString(70, y, "Adm")
    p.drawString(120, y, "Name")
    p.drawString(300, y, "Stream")
    p.drawString(370, y, "Total Points")
    p.drawString(460, y, "Avg Level")
    y -= 18
    p.setFont("Helvetica", 10)
    for i, row in enumerate(merit_list, 1):
        if y < 60:
            p.showPage()
            y = height - 40
        p.drawString(40, y, str(i))
        p.drawString(70, y, str(row["adm"]))
        p.drawString(120, y, str(row["name"]))
        p.drawString(300, y, str(row["stream"]))
        p.drawString(370, y, str(row["total_points"]))
        p.drawString(460, y, str(row["average_level"]))
        y -= 16
    p.showPage()
    p.save()
    buffer.seek(0)
    from django.http import HttpResponse
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="merit_list_{classroom.name}_{exam.title}.pdf"'
    return response
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from schools.models import ClassRoom, Subject, TeacherAssignment, GradeScale
import datetime

@login_required
def load_subjects_for_class(request):
    user = request.user
    school = get_user_school(user)
    class_id = request.GET.get('class_id')
    if not class_id or not school:
        return JsonResponse({'subjects': []})

    is_headteacher = hasattr(user, 'headteacher')
    is_superuser = user.is_superuser

    if is_headteacher or is_superuser:
        # All subjects for this class
        subject_ids = TeacherAssignment.objects.filter(classroom_id=class_id, classroom__school=school).values_list('subject_id', flat=True).distinct()
        subjects = Subject.objects.filter(id__in=subject_ids, school=school)
    elif hasattr(user, 'teacher'):
        teacher = user.teacher
        # Only subjects assigned to this teacher for this class
        subject_ids = TeacherAssignment.objects.filter(teacher=teacher, classroom_id=class_id, classroom__school=school).values_list('subject_id', flat=True).distinct()
        subjects = Subject.objects.filter(id__in=subject_ids, school=school)
    else:
        subjects = Subject.objects.none()

    subject_payload = []
    for s in subjects:
        s_any = cast(Any, s)
        subject_payload.append({
            'id': s_any.id,
            'name': s_any.name,
            'pathway_code': s_any.pathway.code if s_any.pathway else None,
            'level': s_any.education_level.name if s_any.education_level else None,
        })
    return JsonResponse({'subjects': subject_payload})
from schools.forms import SchoolRegistrationForm
from schools.forms import SchoolRegistrationForm, SubjectForm
from schools.models import Student, Teacher
from finance.models import FeePayment as Payment, FeeStructure
from payroll.models import Staff
from schools.models import (
    School, ClassRoom, StreamClassTeacher, TeacherAssignment, Exam, TermDate, MarkSheet, Announcement, Stream, Subject, PromotionLog, EducationLevel, SubjectAllocation, HeadTeacher, SchoolUserAccess
)
from schools.models import StudentMark
from schools.forms import StudentForm, AnnouncementForm, ClassRoomForm
from django.http import HttpResponse
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.pdfgen import canvas
from django.contrib.auth import get_user_model
import io
from schools.forms import TeacherForm
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login
from django.utils import timezone
from django.urls import reverse
from .access import (
    get_user_role as resolve_user_role,
    get_user_school as resolve_user_school,
    has_full_headteacher_access,
    user_has_permission,
    user_has_any_permission,
)

def post_login_redirect(request):
    """Role-aware post-login redirect."""
    if not request.user.is_authenticated:
        return redirect('login')
    if hasattr(request.user, 'headteacher'):
        return redirect('headteacher_dashboard')
    if hasattr(request.user, 'teacher'):
        return redirect('teacher_dashboard')
    if Student.objects.filter(parent_user=request.user).exists():
        return redirect('parent_dashboard')
    # Backward-compat fallback for legacy accounts tied to school email.
    school = School.objects.filter(email__iexact=request.user.username).first()
    if school:
        HeadTeacher.objects.get_or_create(
            user=request.user,
            defaults={
                'school': school,
                'full_name': request.user.get_full_name(),
            },
        )
        return redirect('headteacher_dashboard')
    if request.user.is_superuser:
        return redirect('headteacher_dashboard')
    school = resolve_user_school(request.user)
    role = resolve_user_role(request.user, school)
    if role == SchoolUserAccess.ROLE_ACCOUNTS:
        return redirect('bursar_dashboard')
    if role in (SchoolUserAccess.ROLE_DEAN, SchoolUserAccess.ROLE_SECRETARY, SchoolUserAccess.ROLE_DEPUTY):
        return redirect('headteacher_dashboard')
    if request.user.is_staff:
        return redirect('admin:index')
    return redirect('landing')


def _build_login_form(request, data=None):
    form = AuthenticationForm(request, data=data)
    form.fields['username'].label = 'Email / Phone'
    form.fields['username'].widget.attrs.update({
        'placeholder': 'Email or Phone',
        'autocomplete': 'email',
        'inputmode': 'text',
    })
    form.fields['password'].widget.attrs.update({
        'placeholder': 'Password',
        'autocomplete': 'current-password',
    })
    return form


def signup_modal_redirect(request):
    return redirect(f"{reverse('landing')}?auth=signup")


def _require_school_permission(request, *permissions):
    school = resolve_user_school(request.user)
    if not school:
        return None, HttpResponseForbidden('No school is linked to this account.')
    if request.user.is_superuser:
        return school, None
    if permissions and not user_has_any_permission(request.user, school, permissions):
        return None, HttpResponseForbidden('Access denied for your role.')
    return school, None
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.utils import timezone
from django.db import transaction, IntegrityError
from django.db.models import Sum, Avg, Count, Value, IntegerField, Max, Q
from django.http import HttpResponseForbidden, JsonResponse
import json
from collections import defaultdict
from typing import Any, IO, cast
import re
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect

from django.views.decorators.http import require_POST
from django.views.decorators.http import require_GET

from .cbe import (
    get_primary_level_order,
    ensure_primary_learning_areas,
    get_primary_level,
    get_primary_points,
    is_primary_subject_name,
    recommend_primary_interest,
    ensure_junior_learning_areas,
    ensure_cbe_learning_areas,
    get_junior_level,
    get_junior_points,
    get_junior_level_from_points,
    is_junior_subject_name,
    JUNIOR_LEVEL_BANDS,
    JUNIOR_LEVEL_ORDER,
    recommend_junior_pathway,
    get_performance_level,
    get_comment_variants,
    get_random_comment,
)

from .cbc_utils import get_primary_level_and_points, LOWER_PRIMARY_LEVEL_BANDS, UPPER_PRIMARY_LEVEL_BANDS


def grade_from_mark(mark):
    if mark is None:
        return ""
    if mark >= 80:
        return "A"
    if mark >= 70:
        return "B"
    if mark >= 60:
        return "C"
    if mark >= 50:
        return "D"
    return "E"


LOWER_PRIMARY_PERFORMANCE_LEVELS = ["EE", "ME", "AE", "BE"]
UPPER_PRIMARY_PERFORMANCE_LEVELS = ["EE1", "EE2", "ME1", "ME2", "AE1", "AE2", "BE1", "BE2"]

PRIMARY_GENERAL_TEMPLATES = {
    "Lower Primary": {
        "EE": [
            "Consistently exceeds expectations and demonstrates strong understanding.",
            "Shows outstanding mastery and applies skills confidently.",
            "Excellent performance with clear understanding and creativity.",
            "Demonstrates exceptional competence and independence.",
            "Highly skilled and consistently delivers quality work.",
            "Outstanding progress and surpasses expected standards.",
            "Shows excellent grasp and applies learning in new contexts.",
        ],
        "ME": [
            "Meets expectations and shows steady understanding.",
            "Demonstrates good competence with minor support.",
            "Shows consistent progress and meets required standards.",
            "Competent performance with occasional guidance.",
            "Applies skills appropriately and meets expectations.",
            "Good effort and understanding across tasks.",
            "Shows solid performance at the expected level.",
        ],
        "AE": [
            "Approaching expectations; needs more practice.",
            "Shows developing understanding with support needed.",
            "Progressing but requires reinforcement of key skills.",
            "Needs additional guidance to meet expectations.",
            "Demonstrates partial understanding; more practice required.",
            "Improving steadily; continue targeted support.",
            "Shows potential but needs more consistency.",
        ],
        "BE": [
            "Below expectations; requires significant support.",
            "Needs focused assistance to grasp key skills.",
            "Requires more time and practice to improve.",
            "Limited understanding; intensive support recommended.",
            "Struggles with core skills; needs close guidance.",
            "Below expected level; extra practice is essential.",
            "Requires consistent support to build basic competence.",
        ],
    },
    "Upper Primary": {
        "EE1": [
            "Exceptional performance; consistently exceeds expectations.",
            "Outstanding mastery and application of skills.",
            "Excellent understanding with confident application.",
            "Shows exceptional competence and independence.",
            "Highly proficient with exemplary work quality.",
            "Consistently outstanding across tasks.",
            "Demonstrates superior understanding and creativity.",
        ],
        "EE2": [
            "Very strong performance; exceeds expectations.",
            "Shows high competence and reliable application.",
            "Excellent effort with strong understanding.",
            "Performs above expected standard consistently.",
            "Demonstrates advanced understanding.",
            "Shows strong mastery with minor refinement needed.",
            "Very good performance and application.",
        ],
        "ME1": [
            "Meets expectations with good competence.",
            "Shows reliable understanding and application.",
            "Good performance with occasional support.",
            "Consistently meets the expected standard.",
            "Competent and progressing well.",
            "Applies skills appropriately and effectively.",
            "Good effort and understanding across tasks.",
        ],
        "ME2": [
            "Generally meets expectations; needs minor support.",
            "Shows adequate understanding with some guidance.",
            "Competent but requires consolidation.",
            "Progressing steadily toward expectations.",
            "Meets many expectations with support.",
            "Shows developing competence.",
            "Fair performance with room for improvement.",
        ],
        "AE1": [
            "Approaching expectations; needs more practice.",
            "Shows developing understanding.",
            "Requires guidance to meet expectations.",
            "Progressing but needs reinforcement.",
            "Improving but not yet consistent.",
            "Partial understanding; more practice required.",
            "Needs additional support to improve.",
        ],
        "AE2": [
            "Below expectations; requires more support.",
            "Limited understanding; needs reinforcement.",
            "Requires focused practice to improve.",
            "Needs consistent guidance.",
            "Shows basic understanding only.",
            "Progress is slow; needs support.",
            "Requires targeted intervention.",
        ],
        "BE1": [
            "Significantly below expectations.",
            "Struggles with key skills; needs support.",
            "Requires intensive guidance to improve.",
            "Needs substantial practice and support.",
            "Limited achievement; requires close monitoring.",
            "Below expected level; extra help needed.",
            "Needs consistent intervention to progress.",
        ],
        "BE2": [
            "Very low performance; urgent support needed.",
            "Requires immediate intervention and guidance.",
            "Severely limited understanding.",
            "Needs intensive support to build basics.",
            "Struggles significantly with core skills.",
            "Urgent remediation required.",
            "Requires close supervision and support.",
        ],
    },
}

PRIMARY_SUBJECT_TEMPLATES = {
        "creative_arts": [
            "Shows creativity in arts and crafts.",
            "Participates actively in creative arts activities.",
            "Needs more practice with creative arts skills.",
            "Demonstrates progress in artistic expression.",
            "Requires support to improve creative arts knowledge.",
            "Shows originality in creative arts work.",
            "Applies creative arts skills in various contexts.",
        ],
        "foreign_languages": [
            "Shows good understanding of foreign language concepts.",
            "Participates actively in foreign language lessons.",
            "Needs more practice with foreign language skills.",
            "Demonstrates progress in foreign language comprehension.",
            "Requires support to improve foreign language knowledge.",
            "Shows confidence in foreign language communication.",
            "Applies foreign language in daily life.",
        ],
    "math": [
        "Demonstrates strong number sense and accurate calculations.",
        "Applies mathematical concepts correctly in problem-solving.",
        "Shows good understanding of operations and patterns.",
        "Needs more practice with mathematical procedures.",
        "Solves problems effectively and checks work.",
        "Shows developing reasoning in mathematics.",
        "Requires support to improve accuracy and speed.",
    ],
    "mathematical_activities": [
        "Shows good understanding of mathematical activities.",
        "Participates actively in math-related tasks.",
        "Needs more practice with basic math concepts.",
        "Applies math skills in daily life.",
        "Demonstrates progress in mathematical thinking.",
        "Requires support to improve math skills.",
        "Shows creativity in solving math problems.",
    ],
    "english": [
        "Reads fluently and expresses ideas clearly.",
        "Uses appropriate vocabulary and grammar in writing.",
        "Shows strong comprehension of texts.",
        "Needs more practice with spelling and sentence structure.",
        "Communicates ideas confidently in speaking.",
        "Shows improvement in reading and writing skills.",
        "Requires support to improve comprehension.",
    ],
    "kiswahili": [
        "Reads Kiswahili texts fluently.",
        "Expresses ideas clearly in Kiswahili.",
        "Shows good comprehension of Kiswahili.",
        "Needs more practice with Kiswahili grammar.",
        "Communicates confidently in Kiswahili.",
        "Shows improvement in Kiswahili skills.",
        "Requires support to improve Kiswahili comprehension.",
    ],
    "ksl": [
        "Demonstrates understanding of Kenyan Sign Language.",
        "Communicates effectively using KSL.",
        "Shows progress in KSL vocabulary.",
        "Needs more practice with KSL grammar.",
        "Participates actively in KSL activities.",
        "Requires support to improve KSL skills.",
        "Shows confidence in KSL communication.",
    ],
    "health_education": [
        "Shows good understanding of health concepts.",
        "Participates actively in health education activities.",
        "Needs more practice with healthy habits.",
        "Demonstrates progress in health awareness.",
        "Requires support to improve health knowledge.",
        "Shows responsibility in health-related tasks.",
        "Applies health concepts in daily life.",
    ],
    "religious_education": [
        "Shows good understanding of religious concepts.",
        "Participates actively in religious education.",
        "Needs more practice with religious studies.",
        "Demonstrates respect for religious diversity.",
        "Requires support to improve religious knowledge.",
        "Shows progress in religious values.",
        "Applies religious teachings in daily life.",
    ],
    "creative_activities": [
        "Shows creativity in artistic activities.",
        "Participates actively in creative tasks.",
        "Needs more practice with creative skills.",
        "Demonstrates progress in creative expression.",
        "Requires support to improve creativity.",
        "Shows originality in creative work.",
        "Applies creative skills in various contexts.",
    ],
    "physical_education": [
        "Shows good understanding of physical education concepts.",
        "Participates actively in PE activities.",
        "Needs more practice with physical skills.",
        "Demonstrates progress in fitness and health.",
        "Requires support to improve PE skills.",
        "Shows teamwork in physical activities.",
        "Applies PE concepts in daily life.",
    ],
    "indigenous_language": [
        "Shows good understanding of indigenous language.",
        "Participates actively in indigenous language activities.",
        "Needs more practice with indigenous language skills.",
        "Demonstrates progress in indigenous language comprehension.",
        "Requires support to improve indigenous language.",
        "Shows confidence in indigenous language communication.",
        "Applies indigenous language in daily life.",
    ],
    "mathematics": [
        "Demonstrates strong number sense and accurate calculations.",
        "Applies mathematical concepts correctly in problem-solving.",
        "Shows good understanding of operations and patterns.",
        "Needs more practice with mathematical procedures.",
        "Solves problems effectively and checks work.",
        "Shows developing reasoning in mathematics.",
        "Requires support to improve accuracy and speed.",
    ],
    "science_technology": [
        "Demonstrates good understanding of science and technology concepts.",
        "Applies inquiry skills and observes accurately.",
        "Shows curiosity and engages well in science and technology.",
        "Needs more practice with scientific explanations.",
        "Understands cause and effect in science and technology topics.",
        "Shows progress in scientific reasoning.",
        "Requires support to grasp key concepts.",
    ],
    "social_studies": [
        "Shows good understanding of society and environment.",
        "Explains social concepts clearly and accurately.",
        "Demonstrates awareness of community roles.",
        "Needs more practice with social studies content.",
        "Participates well in discussions about society.",
        "Shows developing understanding of civic issues.",
        "Requires support to improve content knowledge.",
    ],
    "home_science": [
        "Shows good understanding of home science concepts.",
        "Participates actively in home science activities.",
        "Needs more practice with home science skills.",
        "Demonstrates progress in home science.",
        "Requires support to improve home science knowledge.",
        "Shows responsibility in home science tasks.",
        "Applies home science concepts in daily life.",
    ],
    "agriculture": [
        "Shows good understanding of agriculture concepts.",
        "Participates actively in agriculture activities.",
        "Needs more practice with agriculture skills.",
        "Demonstrates progress in agriculture.",
        "Requires support to improve agriculture knowledge.",
        "Shows responsibility in agriculture tasks.",
        "Applies agriculture concepts in daily life.",
    ],
}

PRIMARY_SUBJECT_NAME_MAP = {
        "creative_arts": ["creative arts", "creative activities", "arts", "art"],
        "foreign_languages": ["foreign languages", "french", "german", "arabic", "chinese", "language"],
    "math": ["mathematics", "mathematical activities", "math"],
    "mathematical_activities": ["mathematical activities"],
    "english": ["english"],
    "kiswahili": ["kiswahili", "kiswahili/ksl"],
    "ksl": ["ksl", "kiswahili/ksl"],
    "health_education": ["health education"],
    "religious_education": ["religious education"],
    "creative_activities": ["creative activities"],
    "physical_education": ["physical education"],
    "indigenous_language": ["indigenous language"],
    "mathematics": ["mathematics"],
    "science_technology": ["science & technology", "science", "technology", "integrated science"],
    "social_studies": ["social studies"],
    "home_science": ["home science"],
    "agriculture": ["agriculture"],
}


def _seed_missing_primary_comments(school) -> dict[str, int]:
    from .models import CompetencyComment, Subject

    added = 0
    levels = {
        "Lower Primary": LOWER_PRIMARY_PERFORMANCE_LEVELS,
        "Upper Primary": UPPER_PRIMARY_PERFORMANCE_LEVELS,
    }

    def ensure_variants(level_name: str, performance_level: str, subject_id, comments: list[str]):
        nonlocal added
        existing = CompetencyComment.objects.filter(
            education_level=level_name,
            performance_level=performance_level,
            subject_id=subject_id,
        ).count()
        missing = max(0, 7 - existing)
        for idx in range(missing):
            CompetencyComment.objects.create(
                education_level=level_name,
                performance_level=performance_level,
                subject_id=subject_id,
                comment_text=comments[idx % len(comments)],
            )
            added += 1

    for level_name, perf_levels in levels.items():
        for perf in perf_levels:
            ensure_variants(level_name, perf, None, PRIMARY_GENERAL_TEMPLATES[level_name][perf])

    subjects = list(
        Subject.objects.filter(school=school, education_level__name__in=["Lower Primary", "Upper Primary"]).select_related('education_level')
    )
    for subject in subjects:
        level_name = subject.education_level.name if subject.education_level else None
        if level_name not in levels:
            continue
        name = (subject.name or "").lower()
        subject_key = None
        for key, patterns in PRIMARY_SUBJECT_NAME_MAP.items():
            if any(p in name for p in patterns):
                subject_key = key
                break
        if not subject_key:
            continue
        for perf in levels[level_name]:
            ensure_variants(level_name, perf, cast(Any, subject).id, PRIMARY_SUBJECT_TEMPLATES[subject_key])

    return {"added": added}


def _build_primary_comment_report(school):
    from .models import CompetencyComment, Subject

    levels = {
        "Lower Primary": LOWER_PRIMARY_PERFORMANCE_LEVELS,
        "Upper Primary": UPPER_PRIMARY_PERFORMANCE_LEVELS,
    }
    summary = []
    subject_missing = []

    for level_name, perf_levels in levels.items():
        for perf in perf_levels:
            general_count = CompetencyComment.objects.filter(
                education_level=level_name,
                performance_level=perf,
                subject__isnull=True,
            ).count()
            summary.append({
                "level": level_name,
                "performance_level": perf,
                "general_count": general_count,
                "general_missing": max(0, 7 - general_count),
            })

    subjects = list(
        Subject.objects.filter(school=school, education_level__name__in=["Lower Primary", "Upper Primary"]).select_related('education_level')
    )
    for subject in subjects:
        level_name = subject.education_level.name if subject.education_level else None
        if level_name not in levels:
            continue
        for perf in levels[level_name]:
            count = CompetencyComment.objects.filter(
                education_level=level_name,
                performance_level=perf,
                subject=subject,
            ).count()
            if count < 7:
                subject_missing.append({
                    "level": level_name,
                    "performance_level": perf,
                    "subject": subject.name,
                    "missing": max(0, 7 - count),
                })

    return {
        "summary": summary,
        "subject_missing": subject_missing,
    }


GRADE_POINTS = {
    'A': 12,
    'A-': 11,
    'B+': 10,
    'B': 9,
    'B-': 8,
    'C+': 7,
    'C': 6,
    'C-': 5,
    'D+': 4,
    'D': 3,
    'D-': 2,
    'E': 1,
}


def calculate_points(grade):
    return GRADE_POINTS.get(grade, 0)


CAMBRIDGE_9_1_BANDS = [
    (90, 100, '9', 9),
    (80, 89.99, '8', 8),
    (70, 79.99, '7', 7),
    (60, 69.99, '6', 6),
    (50, 59.99, '5', 5),
    (40, 49.99, '4', 4),
    (30, 39.99, '3', 3),
    (20, 29.99, '2', 2),
    (0, 19.99, '1', 1),
]

CAMBRIDGE_A_G_BANDS = [
    (80, 100, 'A', 7),
    (70, 79.99, 'B', 6),
    (60, 69.99, 'C', 5),
    (50, 59.99, 'D', 4),
    (40, 49.99, 'E', 3),
    (30, 39.99, 'F', 2),
    (0, 29.99, 'G', 1),
]

CAMBRIDGE_COMMENT_BANK = {
    '9': 'Outstanding mastery of the subject. Keep stretching with advanced tasks.',
    '8': 'Excellent performance with strong understanding and consistent effort.',
    '7': 'Very good performance. Continue refining depth and accuracy.',
    '6': 'Good performance with clear progress. Build consistency for higher bands.',
    '5': 'Satisfactory performance. More targeted practice will raise attainment.',
    '4': 'Basic pass achieved. Focus on core gaps to strengthen understanding.',
    '3': 'Below expected level. Regular guided practice is needed.',
    '2': 'Significant support needed. Focus on fundamentals and stepwise improvement.',
    '1': 'Very low attainment currently. Intensive support and practice required.',
    'A': 'Excellent standard attained. Maintain this high-quality performance.',
    'B': 'Strong performance with good command of key concepts.',
    'C': 'Credit-level performance. Continue focused practice to improve.',
    'D': 'Developing pass. Work on weaker areas to secure stronger outcomes.',
    'E': 'Borderline pass. Consistent revision and support are required.',
    'F': 'Limited attainment. Urgent reinforcement and guided practice needed.',
    'G': 'Very weak attainment. Immediate intervention is required.',
}


def _default_cambridge_bands(school) -> list[tuple[float, float, str, int]]:
    scheme = getattr(school, 'cambridge_grading_system', 'CAMB_9_1')
    if scheme == 'CAMB_A_G':
        return CAMBRIDGE_A_G_BANDS
    return CAMBRIDGE_9_1_BANDS


def _normalize_pct(score: Any, out_of: Any) -> float:
    try:
        score_val = float(score or 0)
        out_val = float(out_of or 0)
        if out_val <= 0:
            return 0.0
        return round((score_val / out_val) * 100, 4)
    except Exception:
        return 0.0


def _parse_exam_weights_param(raw_weights: str | None, exam_ids: list[int]) -> dict[int, float]:
    weights: dict[int, float] = {int(eid): 1.0 for eid in exam_ids}
    if not raw_weights:
        return weights
    try:
        payload = json.loads(raw_weights)
    except Exception:
        return weights
    if not isinstance(payload, dict):
        return weights
    for key, value in payload.items():
        try:
            exam_id = int(key)
            if exam_id not in weights:
                continue
            weight = float(value)
            if weight > 0:
                weights[exam_id] = weight
        except Exception:
            continue
    return weights


def _weighted_mean(pairs: list[tuple[float, float]]) -> float | None:
    if not pairs:
        return None
    total_weight = sum(float(w or 0) for _v, w in pairs if float(w or 0) > 0)
    if total_weight <= 0:
        return None
    weighted_total = sum(float(v or 0) * float(w or 0) for v, w in pairs if float(w or 0) > 0)
    return round(weighted_total / total_weight, 2)


def _cambridge_comment_for_level(level: str) -> str:
    return CAMBRIDGE_COMMENT_BANK.get(str(level or '').upper(), 'Steady progress noted. Keep practicing consistently.')


def resolve_cbe_level(school, class_level_name):
    if not school:
        return class_level_name
    if hasattr(school, 'resolve_cbe_level'):
        return school.resolve_cbe_level(class_level_name)
    return class_level_name


def _resolve_primary_band_from_class_name(class_name: str) -> str | None:
    if not class_name:
        return None
    match = re.search(r"(\d+)", class_name)
    if not match:
        return None
    try:
        grade = int(match.group(1))
    except ValueError:
        return None
    if 1 <= grade <= 3:
        return "Lower Primary"
    if 4 <= grade <= 6:
        return "Upper Primary"
    return None


def school_allows_pathways(school) -> bool:
    if not school:
        return False
    if hasattr(school, 'allows_pathways'):
        return school.allows_pathways()
    return school.system_type == 'CBE'


def filter_subjects_for_school(school, queryset):
    if not school or school.system_type != 'CBE':
        return queryset
    category = getattr(school, 'school_category', None)
    if category == 'PRIMARY':
        return queryset.filter(education_level__name__in=['Lower Primary', 'Upper Primary'])

    if category == 'COMPREHENSIVE':
        return queryset.filter(education_level__name__in=['Lower Primary', 'Upper Primary', 'Junior'])

    if category == 'JUNIOR':
        return queryset.filter(education_level__name='Junior')
    if category == 'SENIOR':
        return queryset.filter(education_level__name='Senior')
    return queryset


@login_required
def teacher_dashboard(request):
    """Teacher dashboard showing classes, subjects, and marks entry status"""
    if not hasattr(request.user, 'teacher'):
        return HttpResponseForbidden()
    
    teacher = request.user.teacher
    school = teacher.school
    
    # Get all classes where this teacher is a class teacher (general level)
    class_teacher_classrooms = ClassRoom.objects.filter(
        class_teacher=teacher,
        school=school
    ).select_related('school')
    
    # Get all stream-specific class teacher assignments
    stream_class_teacher = StreamClassTeacher.objects.filter(
        teacher=teacher,
        classroom__school=school
    ).select_related('classroom', 'stream', 'classroom__school')
    
    stream_class_info = []
    for sct in stream_class_teacher:
        stream_class_info.append({
            'classroom': sct.classroom,
            'stream': sct.stream,
        })
    
    # Get all subject assignments for this teacher (scoped to school)
    subject_allocations = TeacherAssignment.objects.filter(
        teacher=teacher,
        classroom__school=school,
        classroom__isnull=False
    ).select_related('classroom', 'classroom__level', 'subject', 'subject__pathway', 'stream')
    
    # Group allocations by classroom for display
    subject_by_classroom = {}
    assigned_classrooms_set = set()
    for alloc in subject_allocations:
        classroom = alloc.classroom
        if not classroom:
            continue
        classroom_any = cast(Any, classroom)
        assigned_classrooms_set.add(classroom_any.id)
        key = classroom_any.id
        if key not in subject_by_classroom:
            subject_by_classroom[key] = []
        subject_by_classroom[key].append({
            'subject': alloc.subject,
            'stream': alloc.stream,
        })
    
    # Get recent exams for the school
    recent_exams = Exam.objects.filter(school=school).order_by('-year', '-id')[:10]
    
    # Get marking status for each exam - only for teacher's assigned classes/subjects
    exam_status = []
    for exam in recent_exams:
        # Get marksheets for this teacher's assignments
        mark_sheets = MarkSheet.objects.filter(
            exam=exam,
            school_class__in=assigned_classrooms_set,
            created_by=request.user,
            status__in=['draft', 'published']
        )
        
        if mark_sheets.exists():
            total = mark_sheets.count()
            published = mark_sheets.filter(status='published').count()
            draft = mark_sheets.filter(status='draft').count()
            
            exam_status.append({
                'exam': exam,
                'total_marksheets': total,
                'published': published,
                'draft': draft,
                'pending': total - published - draft,
                'completion_percentage': round((published / total * 100)) if total > 0 else 0,
            })
    
    # Get current term
    current_term = TermDate.objects.filter(
        school=school,
        start_date__lte=timezone.now().date(),
        end_date__gte=timezone.now().date()
    ).first()
    
    # Get all classrooms this teacher is involved with
    all_classrooms = set()
    for classroom in class_teacher_classrooms:
        all_classrooms.add(classroom)
    for sct in stream_class_teacher:
        all_classrooms.add(sct.classroom)
    for alloc in subject_allocations:
        all_classrooms.add(alloc.classroom)
    
    all_classrooms = sorted(list(all_classrooms), key=lambda x: x.name)
    
    # Get student count per classroom and add to classroom objects
    for classroom in all_classrooms:
        classroom.student_count = Student.objects.filter(
            classroom=classroom,
            school=school
        ).count()
    
    # Calculate summary metrics
    total_classes = len(all_classrooms)
    total_subjects = len({cast(Any, alloc).subject_id for alloc in subject_allocations})
    total_students = sum(c.student_count for c in all_classrooms)
    overall_completion = 0
    if exam_status:
        overall_completion = int(sum(es['completion_percentage'] for es in exam_status) / len(exam_status))
    
    quick_entry_exam = Exam.objects.filter(school=school).order_by('-year', '-id').first()

    context = {
        'teacher': teacher,
        'school': school,
        'allow_pathways': school_allows_pathways(school),
        'class_teacher_classrooms': class_teacher_classrooms,
        'stream_class_info': stream_class_info,
        'subject_allocations': subject_allocations,
        'subject_by_classroom': subject_by_classroom,
        'exam_status': exam_status,
        'current_term': current_term,
        'all_classrooms': all_classrooms,
        'total_classes': total_classes,
        'total_subjects': total_subjects,
        'total_students': total_students,
        'overall_completion': overall_completion,
        'quick_entry_exam': quick_entry_exam,
    }
    
    return render(request, 'schools/teacher_dashboard.html', context)


def _teacher_assigned_class_ids(teacher, school):
    return set(
        TeacherAssignment.objects.filter(
            teacher=teacher,
            classroom__school=school,
            classroom__isnull=False,
        ).values_list('classroom_id', flat=True)
    )


def _teacher_class_teacher_class_ids(teacher, school):
    class_ids = set(
        ClassRoom.objects.filter(class_teacher=teacher, school=school).values_list('id', flat=True)
    )
    class_ids.update(
        StreamClassTeacher.objects.filter(
            teacher=teacher,
            classroom__school=school,
        ).values_list('classroom_id', flat=True)
    )
    return class_ids


def _teacher_allowed_subject_ids_for_class(teacher, school, class_id):
    return set(
        TeacherAssignment.objects.filter(
            teacher=teacher,
            classroom__school=school,
            classroom_id=class_id,
        ).values_list('subject_id', flat=True)
    )


def _teacher_class_teacher_scope_for_class(teacher, school, class_id):
    is_class_teacher_for_whole_class = ClassRoom.objects.filter(
        id=class_id,
        school=school,
        class_teacher=teacher,
    ).exists()
    if is_class_teacher_for_whole_class:
        return True, True, set()

    stream_ids = set(
        StreamClassTeacher.objects.filter(
            teacher=teacher,
            classroom__school=school,
            classroom_id=class_id,
        ).values_list('stream_id', flat=True)
    )
    if stream_ids:
        return True, False, stream_ids
    return False, False, set()


@login_required
def exams_management(request):
    # only headteachers may manage exams
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)

    # POST - create exam (AJAX or normal form)
    if request.method == 'POST':
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        try:
            data = json.loads(request.body) if is_ajax else request.POST

            title = (data.get('title') or '').strip()
            year = int(data.get('year') or 0)
            term = data.get('term')
            start_str = data.get('start_date') or data.get('start')
            end_str = data.get('end_date') or data.get('end')

            # server-side validation
            if not title:
                return JsonResponse({'success': False, 'error': 'Title is required'})
            if year < 1900 or year > 2100:
                return JsonResponse({'success': False, 'error': 'Year must be between 1900 and 2100'})
            if term not in dict(Exam.TERM_CHOICES):
                return JsonResponse({'success': False, 'error': 'Invalid term'})

            start = datetime.datetime.strptime(start_str, "%Y-%m-%d").date()
            end = datetime.datetime.strptime(end_str, "%Y-%m-%d").date()
            if end < start:
                return JsonResponse({'success': False, 'error': 'End date cannot be before start date'})

            exam = Exam.objects.create(
                title=title,
                year=year,
                term=term,
                start_date=start,
                end_date=end,
                school=school,
            )

            # If AJAX, return created exam data so client can update without reload
            if is_ajax:
                exam_any = cast(Any, exam)
                exam_data = {
                    'id': exam_any.id,
                    'title': exam_any.title,
                    'year': exam_any.year,
                    'term': exam_any.term,
                    'start_date': exam_any.start_date.isoformat(),
                    'end_date': exam_any.end_date.isoformat(),
                }
                return JsonResponse({'success': True, 'exam': exam_data})
            # Non-AJAX: add a persistent flash message and redirect back to exams page
            messages.success(request, 'Exam created successfully.')
            return redirect('exams_management')
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'Error creating exam: {e}')
            return redirect('exams_management')

    # GET - render page (only exams for this school)
    exams = Exam.objects.filter(school=school).order_by('-year', 'term')
    return render(request, 'schools/exams.html', {'exams': exams})


@login_required
def term_dates(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)

    # render page with existing term dates
    dates = TermDate.objects.filter(school=school).order_by('-year', 'term')
    return render(request, 'schools/term_dates.html', {'dates': dates})


@login_required
@require_POST
def create_term_date(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    # Defensive fallback: try to resolve headteacher.school directly if get_user_school for some reason returned None
    if not school:
        try:
            from .models import HeadTeacher
            ht = HeadTeacher.objects.filter(user=request.user).first()
            if ht:
                school = ht.school
        except Exception:
            pass
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    try:
        try:
            data = json.loads(request.body) if request.body else request.POST
        except Exception:
            data = request.POST
        year = int(data.get('year') or 0)
        term = data.get('term')
        start = datetime.datetime.strptime(data.get('start_date') or data.get('start'), "%Y-%m-%d").date()
        end = datetime.datetime.strptime(data.get('end_date') or data.get('end'), "%Y-%m-%d").date()

        if year < 1900 or year > 2100:
            return JsonResponse({'success': False, 'error': 'Invalid year'})
        if term not in dict(Exam.TERM_CHOICES):
            return JsonResponse({'success': False, 'error': 'Invalid term'})
        if end < start:
            return JsonResponse({'success': False, 'error': 'End date cannot be before start date'})

        # Ensure we create/update a TermDate instance tied to this school explicitly
        td = TermDate.objects.filter(school=school, year=year, term=term).first()
        if td:
            td.start_date = start
            td.end_date = end
            td.save()
            created = False
        else:
            if not school:
                # defensive error for debugging: return informative message
                print('create_term_date: no school for user', getattr(request.user, 'username', None))
                return JsonResponse({'success': False, 'error': 'No school found for current user'})
            td = TermDate.objects.create(school=school, year=year, term=term, start_date=start, end_date=end)
            created = True

        td_any = cast(Any, td)
        return JsonResponse({'success': True, 'term_date': {'id': td_any.id, 'year': td_any.year, 'term': td_any.term, 'start_date': td_any.start_date.isoformat(), 'end_date': td_any.end_date.isoformat()}})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def edit_term_date(request, pk):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    try:
        td = get_object_or_404(TermDate, pk=pk, school=school)
        try:
            data = json.loads(request.body) if request.body else request.POST
        except Exception:
            data = request.POST
        year = int(data.get('year') or td.year)
        term = data.get('term') or td.term
        start = datetime.datetime.strptime(data.get('start_date') or data.get('start'), "%Y-%m-%d").date()
        end = datetime.datetime.strptime(data.get('end_date') or data.get('end'), "%Y-%m-%d").date()
        if term not in dict(Exam.TERM_CHOICES):
            return JsonResponse({'success': False, 'error': 'Invalid term'})
        if end < start:
            return JsonResponse({'success': False, 'error': 'End date cannot be before start date'})
        td.year = year
        td.term = term
        td.start_date = start
        td.end_date = end
        td.save()
        td_any = cast(Any, td)
        return JsonResponse({'success': True, 'term_date': {'id': td_any.id, 'year': td_any.year, 'term': td_any.term, 'start_date': td_any.start_date.isoformat(), 'end_date': td_any.end_date.isoformat()}})
    except TermDate.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Term date not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def delete_term_date(request, pk):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    try:
        td = get_object_or_404(TermDate, pk=pk, school=school)
        td.delete()
        return JsonResponse({'success': True})
    except TermDate.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Term date not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def edit_exam(request, pk):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    try:
        school = get_user_school(request.user)
        exam = get_object_or_404(Exam, pk=pk, school=school)
        data = json.loads(request.body)

        title = (data.get('title') or '').strip()
        year = int(data.get('year') or 0)
        term = data.get('term')
        start = datetime.datetime.strptime(data['start_date'], "%Y-%m-%d").date()
        end = datetime.datetime.strptime(data['end_date'], "%Y-%m-%d").date()

        if not title:
            return JsonResponse({'success': False, 'error': 'Title required'})
        if year < 1900 or year > 2100:
            return JsonResponse({'success': False, 'error': 'Invalid year'})
        if term not in dict(Exam.TERM_CHOICES):
            return JsonResponse({'success': False, 'error': 'Invalid term'})
        if end < start:
            return JsonResponse({'success': False, 'error': 'End date cannot be before start date'})

        exam.title = title
        exam.year = year
        exam.term = term
        exam.start_date = start
        exam.end_date = end
        exam.save()

        exam_any = cast(Any, exam)
        return JsonResponse({'success': True, 'exam': {'id': exam_any.id, 'title': exam_any.title, 'year': exam_any.year, 'term': exam_any.term, 'start_date': exam_any.start_date.isoformat(), 'end_date': exam_any.end_date.isoformat()}})
    except Exam.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Exam not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def delete_exam(request, pk):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    try:
        school = get_user_school(request.user)
        exam = get_object_or_404(Exam, pk=pk, school=school)
        exam.delete()
        return JsonResponse({'success': True})
    except Exam.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Exam not found'})
    except Exam.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Exam not found'})

def get_user_school(user):
    return resolve_user_school(user)


def get_students_with_balances(school):
    """Return a list of dicts {'student': Student, 'balance': Decimal/float} for students owing money."""
    debtors = []
    students = Student.objects.filter(school=school)
    current_year = timezone.now().year
    terms = ['Term 1', 'Term 2', 'Term 3']

    for student in students:
        expected = sum(student.total_fees_due(term, current_year) for term in terms)
        total_paid = Payment.objects.filter(student=student, year=current_year).aggregate(total=Sum('amount_paid'))['total'] or 0
        try:
            balance = float(expected) - float(total_paid)
        except Exception:
            balance = 0
        if balance > 0:
            debtors.append({'student': student, 'balance': balance})
    return debtors

@login_required
def headteacher_dashboard(request):
    school, denied = _require_school_permission(request, 'students', 'teachers', 'academics')
    if denied:
        return denied
    debtors = get_students_with_balances(school)

    students_count = Student.objects.filter(school=school).count()
    teachers_count = Teacher.objects.filter(school=school).count()
    classes_count = ClassRoom.objects.filter(school=school).count()

    total_paid = Payment.objects.filter(student__school=school).aggregate(
        total=Sum('amount_paid')
    )['total'] or 0

    recent_announcements = Announcement.objects.filter(
        school=school
    ).order_by('-created_at')[:5]

    # 📊 Performance Data using marks
    subject_performance = (
        StudentMark.objects
        .filter(student__school=school)
        .values('marksheet__subject__name')
        .annotate(avg_score=Avg('score'))
        .order_by('marksheet__subject__name')
    )

    context = {
        'school': school,
        'students_count': students_count,
        'teachers_count': teachers_count,
        'classes_count': classes_count,
        'total_paid': total_paid,
        'recent_announcements': recent_announcements,
        'debtors': debtors,
        'subject_performance': subject_performance,
    }

    return render(request, 'schools/headteacher_dashboard.html', context)


@login_required
def bursar_dashboard(request):
    school, denied = _require_school_permission(request, 'finance')
    if denied:
        return denied
    return render(request, 'schools/bursar_dashboard.html', {'school': school})



@login_required
def manage_teachers(request):
    school, denied = _require_school_permission(request, 'teachers')
    if denied:
        return denied
    teachers = Teacher.objects.filter(school=school)

    # support AJAX creation of teacher user + Teacher record
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == "POST":
        if is_ajax:
            # handle new teacher creation
            data = json.loads(request.body)
            password = data.get('password') or ''
            first_name = (data.get('first_name') or '').strip()
            last_name = (data.get('last_name') or '').strip()
            email_input = (data.get('email') or '').strip().lower()
            is_class_teacher = bool(data.get('is_class_teacher'))
            class_teacher_for = data.get('class_teacher_for')
            class_teacher_stream = data.get('class_teacher_stream')

            # basic validation
            if not first_name or not last_name:
                return JsonResponse({'success': False, 'error': 'First and last name are required'}, status=400)
            if not password:
                return JsonResponse({'success': False, 'error': 'Password is required'}, status=400)
            if len(password) < 6:
                return JsonResponse({'success': False, 'error': 'Password must be at least 6 characters'}, status=400)
            
            # validate class assignment if is_class_teacher is True
            if is_class_teacher and class_teacher_for:
                try:
                    classroom = ClassRoom.objects.get(id=class_teacher_for, school=school)
                    # validate stream if provided
                    if class_teacher_stream:
                        Stream.objects.get(id=class_teacher_stream, classroom=classroom)
                except ClassRoom.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Invalid class selection'}, status=400)
                except Stream.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Invalid stream for this class'}, status=400)

            # generate username/email from name and school (or use provided email)
            def slugify_name(s):
                return ''.join(ch for ch in s.lower() if ch.isalnum())
            base_user = f"{slugify_name(first_name)}.{slugify_name(last_name)}"
            school_slug = slugify_name(school.name or 'school')
            domain = f"{school_slug}.skulplus.com"
            User = get_user_model()
            if email_input:
                if User.objects.filter(username=email_input).exists() or User.objects.filter(email=email_input).exists():
                    return JsonResponse({'success': False, 'error': 'A user with that email already exists.'}, status=400)
                username = email_input
                email = email_input
            else:
                # Generate email first, then use it as username for login consistency
                username = base_user
                counter = 1
                while True:
                    candidate_email = f"{username}@{domain}"
                    if not User.objects.filter(username=candidate_email).exists() and not User.objects.filter(email=candidate_email).exists():
                        email = candidate_email
                        username = candidate_email
                        break
                    counter += 1
                    username = f"{base_user}{counter}"

            try:
                user = User.objects.create_user(username=username, email=email, password=password, first_name=first_name, last_name=last_name)
                teacher = Teacher.objects.create(user=user, school=school, is_class_teacher=is_class_teacher)
                
                # Set class teacher assignment if provided
                if is_class_teacher and class_teacher_for:
                    classroom = ClassRoom.objects.get(id=class_teacher_for, school=school)
                    if class_teacher_stream:
                        # Assign to specific stream
                        try:
                            stream = Stream.objects.get(id=class_teacher_stream, classroom=classroom)
                            stream_class_teacher, created = StreamClassTeacher.objects.get_or_create(
                                classroom=classroom,
                                stream=stream
                            )
                            stream_class_teacher.teacher = teacher
                            stream_class_teacher.save()
                        except Stream.DoesNotExist:
                            pass
                    else:
                        # No stream specified - assign to classroom level
                        classroom.class_teacher = teacher
                        classroom.save()
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Error creating teacher: {str(e)}'}, status=500)

            teacher_id = cast(Any, teacher).id
            teacher_data = {
                'id': teacher_id,
                'name': user.get_full_name() or user.username,
                'email': user.email,
                'is_class_teacher': teacher.is_class_teacher,
            }
            return JsonResponse({'success': True, 'teacher': teacher_data})
        else:
            form = TeacherForm(request.POST)
            if form.is_valid():
                teacher = form.save(commit=False)
                teacher.school = school
                teacher.save()
                return redirect('manage_teachers')

    # GET: render page
    form = TeacherForm()
    allocation_form = None
    classes = ClassRoom.objects.filter(school=school).order_by('name')
    
    # Get class teacher assignments for each teacher
    teachers_with_classes = []
    for teacher in teachers:
        teacher_any = cast(Any, teacher)
        assigned_classes = []
        
        # Get classroom-level assignments (no specific stream)
        for classroom in teacher_any.assigned_classes.all():
            class_info = {
                'id': cast(Any, classroom).id,
                'name': classroom.name,
                'section': classroom.section,
                'stream_name': None,
                'stream_id': None,
            }
            assigned_classes.append(class_info)
        
        # Get stream-specific class teacher assignments
        stream_assignments = StreamClassTeacher.objects.filter(teacher=teacher).select_related('classroom', 'stream')
        for stream_ct in stream_assignments:
            classroom_any = cast(Any, stream_ct.classroom)
            stream_any = cast(Any, stream_ct.stream)
            class_info = {
                'id': classroom_any.id,
                'name': classroom_any.name,
                'section': classroom_any.section,
                'stream_name': stream_any.name,
                'stream_id': stream_any.id,
            }
            assigned_classes.append(class_info)
        
        teachers_with_classes.append({
            'teacher': teacher,
            'class_teacher_for': assigned_classes
        })
    
    return render(request, 'schools/manage_teachers.html', {
        'teachers': teachers,
        'teachers_with_classes': teachers_with_classes,
        'form': form,
        'allocation_form': allocation_form,
        'classes': classes,
        'school': school,
    })


@login_required
@require_POST
def edit_teacher(request, pk):
    school, denied = _require_school_permission(request, 'teachers')
    if denied:
        return denied
    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    try:
        try:
            data = json.loads(request.body) if request.body else request.POST
        except Exception:
            data = request.POST

        first_name = (data.get('first_name') or teacher.user.first_name).strip()
        last_name = (data.get('last_name') or teacher.user.last_name).strip()
        is_class_teacher = bool(data.get('is_class_teacher'))
        password = data.get('password', '').strip()
        email = data.get('email', '').strip().lower()
        class_teacher_for = data.get('class_teacher_for')
        class_teacher_stream = data.get('class_teacher_stream')

        # update fields
        user = teacher.user
        user.first_name = first_name
        user.last_name = last_name

        # update password if provided
        if password and len(password) >= 6:
            user.set_password(password)
        
        # update email/username if provided, otherwise keep existing
        if email:
            User = get_user_model()
            user_id = cast(Any, user).id
            email_exists = User.objects.filter(email=email).exclude(id=user_id).exists()
            username_exists = User.objects.filter(username=email).exclude(id=user_id).exists()
            if email_exists or username_exists:
                return JsonResponse({'success': False, 'error': 'A user with that email already exists.'}, status=400)
            user.email = email
            user.username = email
        
        teacher.is_class_teacher = is_class_teacher
        user.save()
        teacher.save()
        
        # Handle class teacher assignment
        # First, remove teacher from any StreamClassTeacher assignments
        StreamClassTeacher.objects.filter(teacher=teacher).delete()
        # Also remove from classroom-level assignment
        ClassRoom.objects.filter(class_teacher=teacher).update(class_teacher=None)
        
        # If class teacher is enabled and a class is selected, assign to that class
        if is_class_teacher and class_teacher_for:
            try:
                classroom = ClassRoom.objects.get(id=class_teacher_for, school=school)
                if class_teacher_stream:
                    # Assign to specific stream
                    try:
                        stream = Stream.objects.get(id=class_teacher_stream, classroom=classroom)
                        stream_class_teacher, created = StreamClassTeacher.objects.get_or_create(
                            classroom=classroom,
                            stream=stream
                        )
                        stream_class_teacher.teacher = teacher
                        stream_class_teacher.save()
                    except Stream.DoesNotExist:
                        pass
                else:
                    # No stream specified - assign to classroom level
                    classroom.class_teacher = teacher
                    classroom.save()
            except ClassRoom.DoesNotExist:
                pass

        teacher_id = cast(Any, teacher).id
        return JsonResponse({'success': True, 'teacher': {'id': teacher_id, 'name': user.get_full_name(), 'email': user.email, 'is_class_teacher': teacher.is_class_teacher}})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def delete_teacher(request, pk):
    school, denied = _require_school_permission(request, 'teachers')
    if denied:
        return denied
    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    try:
        user = teacher.user
        user.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def allocate_teacher(request, teacher_id):
    school, denied = _require_school_permission(request, 'teachers')
    if denied:
        return denied
    try:
        teacher = Teacher.objects.get(id=teacher_id, school=school)
    except Teacher.DoesNotExist:
        return HttpResponseForbidden()

    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if request.method == 'POST' and is_ajax:
        data = json.loads(request.body)
        
        # Support both single allocation (new style) and multiple (legacy)
        subject_id = data.get('subject_id')
        class_id = data.get('class_id')
        stream_id = data.get('stream_id')
        
        if subject_id and class_id:
            # New single allocation style
            # Validate subject and class belong to school
            try:
                subject = Subject.objects.get(id=subject_id, school=school)
                classroom = ClassRoom.objects.get(id=class_id, school=school)
            except (Subject.DoesNotExist, ClassRoom.DoesNotExist):
                return JsonResponse({'success': False, 'error': 'Invalid subject or class'}, status=400)
            
            # Validate stream if provided
            stream = None
            if stream_id:
                try:
                    stream = Stream.objects.get(id=stream_id, classroom=classroom)
                except Stream.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Invalid stream for this class'}, status=400)

            # Enforce allocation source: subject must already have students allocated in selected class/stream.
            allocation_qs = SubjectAllocation.objects.filter(
                subject=subject,
                student__school=school,
                student__classroom=classroom,
            )
            if stream is not None:
                allocation_qs = allocation_qs.filter(student__stream=stream)
            if not allocation_qs.exists():
                return JsonResponse(
                    {'success': False, 'error': 'Selected subject has no allocated students in this class/stream.'},
                    status=400
                )
            
            # Create the assignment (check for duplicates)
            assignment, created = TeacherAssignment.objects.get_or_create(
                teacher=teacher,
                subject=subject,
                classroom=classroom,
                stream=stream
            )
            
            assignment_id = cast(Any, assignment).id
            return JsonResponse({'success': True, 'assignment_id': assignment_id})
        
        else:
            # Legacy multiple allocation style
            subject_ids = data.get('subjects', [])
            class_ids = data.get('classes', [])
            stream_ids = data.get('streams', [])
            
            # validate subjects/classes/streams belong to this school
            valid_subjects = set(Subject.objects.filter(id__in=subject_ids, school=school).values_list('id', flat=True))
            if len(valid_subjects) != len(subject_ids):
                return JsonResponse({'success': False, 'error': 'One or more subjects are invalid'}, status=400)

            valid_classes = set(ClassRoom.objects.filter(id__in=class_ids, school=school).values_list('id', flat=True))
            if len(valid_classes) != len(class_ids):
                return JsonResponse({'success': False, 'error': 'One or more classes are invalid'}, status=400)

            valid_streams = set(Stream.objects.filter(id__in=stream_ids).values_list('id', flat=True))
            if len(valid_streams) != len(stream_ids):
                return JsonResponse({'success': False, 'error': 'One or more streams are invalid'}, status=400)

            # remove existing assignments for this teacher for this school
            TeacherAssignment.objects.filter(teacher=teacher).delete()

            # create assignments: 
            # if both classes and streams provided, create per-class-stream assignment
            # if only classes provided, create per-class assignment (no stream restriction)
            # if neither provided, create global assignment
            if class_ids:
                if stream_ids:
                    # Class + Stream specific
                    for sid in subject_ids:
                        for cid in class_ids:
                            for stm_id in stream_ids:
                                TeacherAssignment.objects.create(
                                    teacher=teacher,
                                    subject_id=sid,
                                    classroom_id=cid,
                                    stream_id=stm_id
                                )
                else:
                    # Class only (all streams in class)
                    for sid in subject_ids:
                        for cid in class_ids:
                            TeacherAssignment.objects.create(teacher=teacher, subject_id=sid, classroom_id=cid)
            else:
                # Global (all classes and streams)
                for sid in subject_ids:
                    TeacherAssignment.objects.create(teacher=teacher, subject_id=sid)

            return JsonResponse({'success': True})

    # For GET (AJAX) return current allocations
    if is_ajax and request.method == 'GET':
        assignments = TeacherAssignment.objects.filter(teacher=teacher).select_related(
            'subject', 'subject__pathway', 'classroom', 'classroom__level', 'stream'
        )
        assignments_data = []
        for a in assignments:
            a_any = cast(Any, a)
            subject = a_any.subject
            classroom = a_any.classroom
            stream = a_any.stream
            assignments_data.append({
                'id': a_any.id,
                'subject_name': subject.name,
                'subject_id': subject.id,
                'subject_pathway_code': subject.pathway.code if subject.pathway else None,
                'class_name': f"{classroom.name} {classroom.section}" if classroom else None,
                'class_id': classroom.id if classroom else None,
                'class_level': classroom.level.name if classroom and classroom.level else None,
                'stream_name': stream.name if stream else None,
                'stream_id': stream.id if stream else None,
            })
        
        # Also return simple lists for backwards compatibility
        subjects = [cast(Any, a).subject.id for a in assignments]
        classes = [cast(Any, a).classroom.id for a in assignments if cast(Any, a).classroom]
        streams = [cast(Any, a).stream.id for a in assignments if cast(Any, a).stream]
        
        return JsonResponse({
            'assignments': assignments_data,
            'subjects': subjects,
            'classes': classes,
            'streams': streams
        })

    return HttpResponseForbidden()


@login_required
def delete_teacher_assignment(request, assignment_id):
    school, denied = _require_school_permission(request, 'teachers')
    if denied:
        return denied
    
    try:
        assignment = TeacherAssignment.objects.select_related('teacher').get(id=assignment_id, teacher__school=school)
    except TeacherAssignment.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Assignment not found'}, status=404)

    if request.method == 'POST':
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        if is_ajax:
            assignment.delete()
            return JsonResponse({'success': True})
        
    return HttpResponseForbidden()


@login_required
def load_school_subjects_classes(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    class_id = request.GET.get('class_id')
    stream_id = request.GET.get('stream_id')

    classes = ClassRoom.objects.filter(school=school).values('id', 'name', 'section')
    subjects = Subject.objects.none()

    if class_id:
        students_qs = Student.objects.filter(
            school=school,
            classroom_id=class_id,
        )
        if stream_id:
            students_qs = students_qs.filter(stream_id=stream_id)

        allocated_subject_ids = SubjectAllocation.objects.filter(
            student__in=students_qs,
            subject__school=school,
        ).values_list('subject_id', flat=True).distinct()

        subjects = filter_subjects_for_school(
            school,
            Subject.objects.filter(
                school=school,
                id__in=allocated_subject_ids,
            ).select_related('pathway', 'education_level')
        ).order_by('name')

    subject_payload = []
    for s in subjects:
        s_any = cast(Any, s)
        subject_payload.append({
            'id': s_any.id,
            'name': s_any.name,
            'pathway_code': s_any.pathway.code if s_any.pathway else None,
            'level': s_any.education_level.name if s_any.education_level else None,
        })
    return JsonResponse({'subjects': subject_payload, 'classes': list(classes)})


@login_required
@require_GET
def load_streams_for_class(request):
    """Returns list of streams for a given classroom."""
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()

    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    scope = (request.GET.get('scope') or '').strip().lower()

    if not class_id:
        return JsonResponse({'streams': []})

    # Verify classroom belongs to this school
    classroom = get_object_or_404(ClassRoom, id=class_id, school=school)
    streams_qs = Stream.objects.filter(classroom=classroom)

    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        if scope == 'class_teacher':
            allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
                teacher, school, cast(Any, classroom).id
            )
            if not allowed:
                return JsonResponse({'streams': []})
            if not all_streams:
                streams_qs = streams_qs.filter(id__in=allowed_stream_ids)
            streams = streams_qs.values('id', 'name', 'code').order_by('name')
            return JsonResponse({'streams': list(streams)})

        assignments = TeacherAssignment.objects.filter(
            teacher=teacher,
            classroom=classroom,
            classroom__school=school,
        )
        if subject_id:
            assignments = assignments.filter(subject_id=subject_id)

        if not assignments.exists():
            return JsonResponse({'streams': []})

        # stream=NULL assignment means teacher is assigned all streams for this class(/subject)
        has_all_streams = assignments.filter(stream__isnull=True).exists()
        if not has_all_streams:
            assigned_stream_ids = assignments.exclude(stream__isnull=True).values_list('stream_id', flat=True).distinct()
            streams_qs = streams_qs.filter(id__in=assigned_stream_ids)

    streams = streams_qs.values('id', 'name', 'code').order_by('name')
    return JsonResponse({'streams': list(streams)})


@login_required
def admit_student(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    # Determine if request is AJAX early (used for both GET and POST handling)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == "POST":
        # include files (photo) when binding form
        form = StudentForm(request.POST, request.FILES, school=school)
        if form.is_valid():
            student = form.save(commit=False)
            student.school = school
            student.save()
            parent_username, parent_password, parent_error = _sync_parent_account_for_student(student)
            if parent_error:
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': {'parent_phone': [parent_error]}}, status=400)
                form.add_error('parent_phone', parent_error)
                return render(request, 'schools/admit_student.html', {'form': form, 'students': Student.objects.filter(school=school).order_by('last_name', 'first_name'), 'classes': ClassRoom.objects.filter(school=school)})

            # If AJAX, return JSON with new student info for client-side insertion
            if is_ajax:
                # build minimal student payload for the table row
                photo_url = ''
                try:
                    if student.photo:
                        photo_url = student.photo.url
                except Exception:
                    photo_url = ''

                student_data = {
                    'id': student.id,
                    'admission_number': student.admission_number,
                    'first_name': student.first_name,
                    'last_name': student.last_name,
                    'classroom': str(student.classroom) if student.classroom else '',
                    'stream': str(student.stream) if student.stream else '',
                    'parent_name': getattr(student, 'parent_name', '') or '',
                    'parent_phone': getattr(student, 'parent_phone', '') or '',
                    'parent_username': parent_username,
                    'parent_password': parent_password,
                    'photo_url': photo_url,
                }
                return JsonResponse({'success': True, 'student': student_data})

            # non-AJAX POST: redirect back to admit page so the list refreshes
            if parent_username and parent_password:
                messages.success(request, f'Parent account created. Username: {parent_username} | Password: {parent_password}')
            return redirect('admit_student')
        else:
            # form invalid
            if is_ajax:
                # return form errors as JSON so client can show them in the modal
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    else:
        form = StudentForm(school=school)

    # If this is an AJAX GET request, return just the form partial (for modal)
    if is_ajax and request.method == 'GET':
        return render(request, 'schools/_admit_student_form.html', {'form': form})

    # Non-AJAX: render full page with students and classes scoped to the headteacher's school
    students = Student.objects.filter(school=school).order_by('last_name', 'first_name')
    classes = ClassRoom.objects.filter(school=school)

    return render(request, 'schools/admit_student.html', {
        'form': form,
        'students': students,
        'classes': classes,
    })


@login_required
def admit_student_new(request):
    """Full page admit form (navigates like classes)."""
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, school=school)

        if form.is_valid():
            student = form.save(commit=False)
            student.school = school
            student.save()
            parent_username, parent_password, parent_error = _sync_parent_account_for_student(student)
            if parent_error:
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': {'parent_phone': [parent_error]}}, status=400)
                form.add_error('parent_phone', parent_error)
                return render(request, 'schools/admit_student_form.html', {'form': form})
            if is_ajax:
                photo_url = ''
                try:
                    if student.photo:
                        photo_url = student.photo.url
                except Exception:
                    photo_url = ''
                student_data = {
                    'id': student.id,
                    'admission_number': student.admission_number,
                    'first_name': student.first_name,
                    'last_name': student.last_name,
                    'classroom': str(student.classroom) if student.classroom else '',
                    'stream': str(student.stream) if student.stream else '',
                    'parent_name': getattr(student, 'parent_name', '') or '',
                    'parent_phone': getattr(student, 'parent_phone', '') or '',
                    'parent_username': parent_username,
                    'parent_password': parent_password,
                    'photo_url': photo_url,
                }
                return JsonResponse({'success': True, 'student': student_data})

            if parent_username and parent_password:
                messages.success(request, f'Parent account created. Username: {parent_username} | Password: {parent_password}')
            return redirect('admit_student')
        else:
            if is_ajax:
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = StudentForm(school=school)

    return render(request, 'schools/admit_student_form.html', {'form': form})


@login_required
def post_announcement(request):
    school, denied = _require_school_permission(request, 'academics')
    if denied:
        return denied

    if request.method == "POST":
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.school = school
            announcement.posted_by = request.user
            announcement.save()
            return redirect('headteacher_dashboard')
    else:
        form = AnnouncementForm()

    return render(request, 'schools/post_announcement.html', {'form': form})


def landing(request):
    """Public landing page — always render homepage first.

    Note: role-based redirects happen after login via `post_login_redirect`.
    """
    context = {'user': request.user} if request.user.is_authenticated else {}
    if not request.user.is_authenticated:
        open_auth_modal = (request.GET.get('auth') or '').strip().lower()
        if open_auth_modal not in ('login', 'signup'):
            open_auth_modal = ''
        context.update({
            'form': _build_login_form(request),
            'signup_form': SchoolRegistrationForm(),
            'open_auth_modal': open_auth_modal,
        })
    return render(request, 'landing.html', context)


@ensure_csrf_cookie
@csrf_protect
def login_view(request):
    """Combined login + signup view handling both actions in one template."""
    login_form = _build_login_form(request)
    signup_form = SchoolRegistrationForm()

    if request.method == 'GET':
        return redirect(f"{reverse('landing')}?auth=login")

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'login':
            login_form = _build_login_form(request, data=request.POST)
            if login_form.is_valid():
                user = login_form.get_user()
                login(request, user)
                return redirect('post_login')

        elif form_type == 'signup':
            signup_form = SchoolRegistrationForm(request.POST)
            if signup_form.is_valid():
                User = get_user_model()
                head_email = signup_form.cleaned_data['head_email'].strip().lower()
                school_email = signup_form.cleaned_data['school_email'].strip().lower()
                head_full_name = signup_form.cleaned_data['head_full_name']
                head_password = signup_form.cleaned_data['head_password']
                first_name = ''
                last_name = ''
                if ' ' in head_full_name:
                    parts = head_full_name.split(' ', 1)
                    first_name, last_name = parts[0], parts[1]
                else:
                    first_name = head_full_name

                try:
                    with transaction.atomic():
                        school = School.objects.create(
                            name=signup_form.cleaned_data['school_name'],
                            school_type=signup_form.cleaned_data['school_type'],
                            system_type=('CBE' if signup_form.cleaned_data['school_type'] == 'CBE' else '844'),
                            school_category=signup_form.cleaned_data.get('school_category', 'PRIMARY'),
                            cambridge_grading_system=signup_form.cleaned_data.get('cambridge_grading_system') or 'CAMB_9_1',
                            address=signup_form.cleaned_data.get('address', ''),
                            phone=signup_form.cleaned_data.get('phone', ''),
                            email=school_email,
                        )

                        # Headteacher account logs in with school email.
                        headteacher_user = User.objects.create_user(
                            username=school_email,
                            email=school_email,
                            password=head_password,
                            first_name=first_name,
                            last_name=last_name,
                        )
                        HeadTeacher.objects.create(
                            user=headteacher_user,
                            school=school,
                            full_name=head_full_name,
                            phone=signup_form.cleaned_data.get('head_phone', ''),
                        )

                        # Teacher account logs in with provided headteacher personal email.
                        teacher_user = User.objects.create_user(
                            username=head_email,
                            email=head_email,
                            password=head_password,
                            first_name=first_name,
                            last_name=last_name,
                        )
                        Teacher.objects.create(
                            user=teacher_user,
                            school=school,
                            is_class_teacher=False,
                        )
                except IntegrityError:
                    signup_form.add_error(None, 'Unable to create account due to a duplicate record. Use different school/headteacher emails.')
                else:
                    login(request, headteacher_user)
                    return redirect('headteacher_dashboard')

    return render(request, 'landing.html', {
        'form': login_form,
        'signup_form': signup_form,
        'open_auth_modal': 'signup' if request.POST.get('form_type') == 'signup' else 'login',
    })


def signup(request):
    # Provide a combined school + headteacher registration form
    if request.method == 'POST':
        form = SchoolRegistrationForm(request.POST)
        if form.is_valid():
            User = get_user_model()
            head_email = form.cleaned_data['head_email'].strip().lower()
            school_email = form.cleaned_data['school_email'].strip().lower()
            head_full_name = form.cleaned_data['head_full_name']
            head_password = form.cleaned_data['head_password']
            first_name = ''
            last_name = ''
            if ' ' in head_full_name:
                parts = head_full_name.split(' ', 1)
                first_name, last_name = parts[0], parts[1]
            else:
                first_name = head_full_name

            try:
                with transaction.atomic():
                    school = School.objects.create(
                        name=form.cleaned_data['school_name'],
                        school_type=form.cleaned_data['school_type'],
                        system_type=('CBE' if form.cleaned_data['school_type'] == 'CBE' else '844'),
                        school_category=form.cleaned_data.get('school_category', 'PRIMARY'),
                        cambridge_grading_system=form.cleaned_data.get('cambridge_grading_system') or 'CAMB_9_1',
                        address=form.cleaned_data.get('address', ''),
                        phone=form.cleaned_data.get('phone', ''),
                        email=school_email,
                    )
                    headteacher_user = User.objects.create_user(
                        username=school_email,
                        email=school_email,
                        password=head_password,
                        first_name=first_name,
                        last_name=last_name,
                    )
                    HeadTeacher.objects.create(
                        user=headteacher_user,
                        school=school,
                        full_name=head_full_name,
                        phone=form.cleaned_data.get('head_phone', ''),
                    )
                    teacher_user = User.objects.create_user(
                        username=head_email,
                        email=head_email,
                        password=head_password,
                        first_name=first_name,
                        last_name=last_name,
                    )
                    Teacher.objects.create(
                        user=teacher_user,
                        school=school,
                        is_class_teacher=False,
                    )
            except IntegrityError:
                form.add_error(None, 'Unable to create account due to a duplicate record. Use different school/headteacher emails.')
            else:
                login(request, headteacher_user)
                messages.success(request, 'School and Headteacher account created. Welcome!')
                return redirect('headteacher_dashboard')
    else:
        form = SchoolRegistrationForm()
    return render(request, 'registration/signup.html', {'form': form})


# ------------------ Placeholder pages for sidebar ------------------
@login_required
def edit_students(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    students = Student.objects.filter(school=school).select_related(
        'classroom__level', 'stream'
    ).select_related('studentpathway__pathway').order_by('last_name', 'first_name')

    # ensure template-safe `photo` attribute exists (some Student records may not have it)
    for s in students:
        if not hasattr(s, 'photo'):
            setattr(s, 'photo', None)

    return render(request, 'schools/edit_students.html', {
        'students': students,
        'school': school,
        'allow_pathways': school_allows_pathways(school),
    })


@login_required
def parent_dashboard(request):
    students = Student.objects.filter(parent_user=request.user).select_related(
        'classroom__level', 'stream', 'school'
    ).order_by('first_name', 'last_name')
    if not students.exists():
        return HttpResponseForbidden()

    term_order = {'Term 1': 1, 'Term 2': 2, 'Term 3': 3}

    def resolve_active_term_year(school):
        latest_term = (
            TermDate.objects.filter(school=school)
            .order_by('-year', '-term')
            .first()
        )
        if latest_term:
            return latest_term.term, latest_term.year
        latest_exam = (
            Exam.objects.filter(school=school)
            .order_by('-year', '-term', '-start_date')
            .first()
        )
        if latest_exam:
            return latest_exam.term, latest_exam.year
        now = timezone.now()
        return 'Term 1', now.year

    student_cards = []
    for student in students:
        school = student.school
        term, year = resolve_active_term_year(school)
        balance = student.balance(term, year)

        payments = list(
            Payment.objects.filter(student=student)
            .order_by('-date_paid', '-id')[:20]
        )

        class_level = student.classroom.level.name if student.classroom and student.classroom.level else None
        resolved_level = resolve_cbe_level(school, class_level)
        resolve_grade_points, _grade_list = _build_grade_resolver_for_class(
            school, student.classroom, resolved_level
        ) if student.classroom else (lambda pct: ('-', None), [])

        marks = (
            StudentMark.objects.filter(student=student, score__isnull=False)
            .select_related('marksheet__exam', 'marksheet')
        )
        exam_map = {}
        for m in marks:
            ms = cast(Any, m).marksheet
            ex = cast(Any, ms).exam
            key = ex.id
            entry = exam_map.setdefault(key, {
                'title': ex.title,
                'term': ex.term,
                'year': ex.year,
                'scores': [],
                'order': (ex.year, term_order.get(ex.term, 0), ex.title),
            })
            pct = _normalize_pct(m.score, ms.out_of)
            entry['scores'].append(pct)

        academic_history = []
        for entry in exam_map.values():
            scores = entry['scores']
            avg_pct = round(sum(scores) / len(scores), 2) if scores else 0.0
            grade, _pts = resolve_grade_points(avg_pct)
            academic_history.append({
                'title': entry['title'],
                'term': entry['term'],
                'year': entry['year'],
                'avg_pct': avg_pct,
                'grade': grade or '-',
            })
        academic_history.sort(key=lambda r: (r['year'], term_order.get(r['term'], 0), r['title']), reverse=True)

        student_cards.append({
            'student': student,
            'term': term,
            'year': year,
            'balance': balance,
            'payments': payments,
            'academic_history': academic_history,
        })

    announcements = Announcement.objects.filter(
        school_id__in={s.school_id for s in students}
    ).order_by('-created_at')[:20]

    return render(request, 'schools/parent_dashboard.html', {
        'student_cards': student_cards,
        'announcements': announcements,
    })


@login_required
def edit_student(request, student_id):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    if not school:
        return HttpResponseForbidden()
    try:
        student = Student.objects.get(id=student_id, school=school)
    except Student.DoesNotExist:
        return HttpResponseForbidden()

    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student, school=school)

        if form.is_valid():
            student = form.save()
            parent_username, parent_password, parent_error = _sync_parent_account_for_student(student)
            if parent_error:
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': {'parent_phone': [parent_error]}}, status=400)
                messages.error(request, parent_error)
                return redirect('edit_student', student_id=student.id)
            pathway_id = request.POST.get('pathway_id')
            level_name = student.classroom.level.name if student.classroom and student.classroom.level else None
            if school.system_type == 'CBE':
                try:
                    from academics.models import StudentPathway
                except Exception:
                    StudentPathway = None

                if level_name == 'Senior':
                    if not school_allows_pathways(school):
                        if StudentPathway:
                            StudentPathway.objects.filter(student=student).delete()
                    elif not pathway_id:
                        if is_ajax:
                            return JsonResponse({'success': False, 'errors': {'pathway_id': ['Pathway is required for Senior students.']}}, status=400)
                        messages.error(request, 'Pathway is required for Senior students.')
                        return redirect('edit_student', student_id=student.id)

                    if StudentPathway:
                        StudentPathway.objects.update_or_create(
                            student=student,
                            defaults={'pathway_id': pathway_id}
                        )
                else:
                    if StudentPathway:
                        StudentPathway.objects.filter(student=student).delete()
            if is_ajax:
                try:
                    photo_url = student.photo.url if student.photo else ''
                except Exception:
                    photo_url = ''
                student_data = {
                    'id': student.id,
                    'first_name': student.first_name,
                    'last_name': student.last_name,
                    'admission_number': student.admission_number,
                    'classroom': str(student.classroom) if student.classroom else '',
                    'classroom_level': student.classroom.level.name if student.classroom and student.classroom.level else '',
                    'stream': str(student.stream) if student.stream else '',
                    'parent_name': getattr(student, 'parent_name', '') or '',
                    'parent_phone': getattr(student, 'parent_phone', '') or '',
                    'parent_username': parent_username,
                    'parent_password': parent_password,
                    'photo_url': photo_url,
                }
                return JsonResponse({'success': True, 'student': student_data})

            if parent_username and parent_password:
                messages.success(request, f'Parent account updated. Username: {parent_username} | Password: {parent_password}')
            return redirect('edit_students')
        else:
            if is_ajax:
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = StudentForm(instance=student, school=school)

    try:
        from academics.models import StudentPathway
        student_pathway = StudentPathway.objects.filter(student=student).first()
    except Exception:
        student_pathway = None

    allow_pathways = school_allows_pathways(school)
    try:
        pathways = Pathway.objects.all().order_by('code') if allow_pathways else []
    except Exception:
        pathways = []
    level_name = student.classroom.level.name if student.classroom and student.classroom.level else None

    # If AJAX GET, return only the form partial for modal insertion
    if is_ajax and request.method == 'GET':
        return render(request, 'schools/_edit_student_form.html', {
            'form': form,
            'student': student,
            'school': school,
            'pathways': pathways,
            'student_pathway': student_pathway,
            'student_level': level_name,
            'allow_pathways': school_allows_pathways(school),
        })

    return render(request, 'schools/edit_student.html', {
        'form': form,
        'student': student,
        'school': school,
        'pathways': pathways,
        'student_pathway': student_pathway,
        'student_level': level_name,
        'allow_pathways': school_allows_pathways(school),
    })


@login_required
def view_student(request, student_id):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    try:
        student = Student.objects.get(id=student_id, school=school)
    except Student.DoesNotExist:
        return HttpResponseForbidden()

    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    try:
        from academics.models import StudentPathway
        student_pathway = StudentPathway.objects.filter(student=student).first()
    except Exception:
        student_pathway = None

    context = {
        'student': student,
        'student_pathway': student_pathway,
        'allow_pathways': school_allows_pathways(school),
    }

    if is_ajax and request.method == 'GET':
        return render(request, 'schools/_view_student.html', context)

    return render(request, 'schools/view_student.html', context)


@login_required
def class_lists(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    can_students = user_has_permission(request.user, school, 'students')
    if not (is_headteacher or is_superuser or is_teacher or can_students):
        return HttpResponseForbidden()

    ensure_cbe_learning_areas(school)

    # Return queryset objects so the template can iterate normally with
    # `{% for c in classes %}` and `{% for s in subjects %}`.
    classes_qs = ClassRoom.objects.filter(school=school)
    subjects_qs = filter_subjects_for_school(
        school,
        Subject.objects.filter(school=school).select_related('pathway', 'education_level')
    )
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        allowed_class_ids = _teacher_assigned_class_ids(teacher, school) | _teacher_class_teacher_class_ids(teacher, school)
        allowed_subject_ids = set(
            TeacherAssignment.objects.filter(
                teacher=teacher,
                classroom__school=school,
            ).values_list('subject_id', flat=True)
        )
        classes_qs = classes_qs.filter(id__in=allowed_class_ids)
        if allowed_subject_ids:
            subjects_qs = subjects_qs.filter(id__in=allowed_subject_ids)
        else:
            subjects_qs = subjects_qs.none()
    classes = classes_qs.order_by('name')
    subjects = subjects_qs.order_by('name')

    # Always render the full class lists page (remove modal/partial behavior)
    context = {
        'classes': classes,
        'subjects': subjects,
        'school': school,
        'allow_pathways': school_allows_pathways(school),
    }
    return render(request, 'schools/class_lists.html', context)


@login_required
def promote_students(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    classes = ClassRoom.objects.filter(school=school).order_by('name')
    return render(request, 'schools/promote_students.html', {'classes': classes})


@login_required
@require_http_methods(["POST"])
def move_students(request):
    """Move selected students to a target classroom.

    Expects JSON body: { "student_ids": [1,2,3], "target_class_id": 5 }
    """
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    student_ids = data.get('student_ids') or []
    target_class_id = data.get('target_class_id')

    if not student_ids or not target_class_id:
        return JsonResponse({'success': False, 'error': 'Missing parameters'}, status=400)

    # Ensure target classroom belongs to the headteacher's school
    try:
        target_cls = ClassRoom.objects.get(id=target_class_id, school=school)
    except ClassRoom.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid target class'}, status=400)

    # Update only students belonging to this school — iterate to record logs
    qs = Student.objects.filter(id__in=student_ids, school=school)
    moved = 0
    for s in qs.select_related('classroom'):
        old_cls = s.classroom
        s.classroom = target_cls
        s.save()
        PromotionLog.objects.create(
            student=s,
            from_class=old_cls,
            to_class=target_cls,
            performed_by=request.user,
            note=f"Moved via bulk move_students API"
        )
        moved += 1

    return JsonResponse({'success': True, 'moved': moved})


@login_required
@require_GET
def get_adjacent_class(request):
    """Return previous or next class id for a given class within the headteacher's school.

    Query params: class_id, direction=prev|next (default prev)
    """
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    class_id = request.GET.get('class_id')
    direction = request.GET.get('direction', 'prev')
    if not class_id:
        return JsonResponse({'id': None})

    # Determine ordering: prefer `order` field if present, else fallback to `id`
    cls_qs = ClassRoom.objects.filter(school=school)
    ordering_field = 'id'
    # detect 'order' field presence
    if hasattr(ClassRoom, 'order'):
        ordering_field = 'order'

    cls_list = list(cls_qs.order_by(ordering_field).values_list('id', flat=True))
    try:
        idx = cls_list.index(int(class_id))
    except Exception:
        return JsonResponse({'id': None})

    if direction == 'next':
        target = cls_list[idx+1] if idx+1 < len(cls_list) else None
    else:
        target = cls_list[idx-1] if idx-1 >= 0 else None

    return JsonResponse({'id': target})


@login_required
@require_http_methods(["POST"])
def promote_to_next(request):
    """Promote given students to the next class (based on ordering). If no next class, return that info."""
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    student_ids = data.get('student_ids') or []
    if not student_ids:
        return JsonResponse({'success': False, 'error': 'No students provided'}, status=400)

    moved = 0
    graduated = 0
    for s in Student.objects.filter(id__in=student_ids, school=school).select_related('classroom'):
        src = s.classroom
        if not src:
            continue
        # find next class for this student's current class
        resp = get_adjacent_class(request._request if hasattr(request, '_request') else request)
        # instead call logic directly: build ordered list
        cls_list = list(ClassRoom.objects.filter(school=school).order_by('order', 'name').values_list('id', flat=True))
        try:
            src_any = cast(Any, src)
            idx = cls_list.index(src_any.id)
            nxt = cls_list[idx+1] if idx+1 < len(cls_list) else None
        except Exception:
            nxt = None

        if nxt:
            tgt = ClassRoom.objects.get(id=nxt, school=school)
            s.classroom = tgt
            s.is_alumni = False
            s.save()
            PromotionLog.objects.create(student=s, from_class=src, to_class=tgt, performed_by=request.user, note='Promoted to next')
            moved += 1
        else:
            # no next class -> graduate
            s.classroom = None
            s.is_alumni = True
            s.save()
            PromotionLog.objects.create(student=s, from_class=src, to_class=None, performed_by=request.user, note='Graduated to alumni')
            graduated += 1

    return JsonResponse({'success': True, 'moved': moved, 'graduated': graduated})


@login_required
@require_http_methods(["POST"])
def graduate_students(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    student_ids = data.get('student_ids') or []
    if not student_ids:
        return JsonResponse({'success': False, 'error': 'No students provided'}, status=400)
    count = 0
    for s in Student.objects.filter(id__in=student_ids, school=school).select_related('classroom'):
        old = s.classroom
        s.classroom = None
        s.is_alumni = True
        s.save()
        PromotionLog.objects.create(student=s, from_class=old, to_class=None, performed_by=request.user, note='Graduated to alumni')
        count += 1
    return JsonResponse({'success': True, 'graduated': count})


@login_required
@require_http_methods(["POST"])
def undo_promotion(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    student_ids = data.get('student_ids') or []
    if not student_ids:
        return JsonResponse({'success': False, 'error': 'No students provided'}, status=400)

    undone = 0
    for s in Student.objects.filter(id__in=student_ids, school=school):
        last = PromotionLog.objects.filter(student=s).order_by('-timestamp').first()
        if not last:
            continue
        # revert
        s.classroom = last.from_class
        s.is_alumni = False if last.from_class else False
        s.save()
        PromotionLog.objects.create(student=s, from_class=last.to_class, to_class=last.from_class, performed_by=request.user, note='Undo promotion')
        undone += 1
    return JsonResponse({'success': True, 'undone': undone})


@login_required
@require_GET
def promotion_logs(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    student_id = request.GET.get('student_id')
    qs = PromotionLog.objects.select_related('student', 'from_class', 'to_class', 'performed_by')
    if student_id:
        qs = qs.filter(student_id=student_id)
    qs = qs.order_by('-timestamp')[:200]
    data = []
    for p in qs:
        data.append({
            'student_id': p.student.id,
            'student_name': str(p.student),
            'from_class': p.from_class.name if p.from_class else None,
            'to_class': p.to_class.name if p.to_class else None,
            'by': p.performed_by.username if p.performed_by else None,
            'timestamp': p.timestamp.isoformat(),
            'note': p.note,
        })
    return JsonResponse({'logs': data})


@login_required
@ensure_csrf_cookie
def enter_marks(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()

    ensure_cbe_learning_areas(school)

    user = request.user
    is_headteacher = hasattr(user, 'headteacher')
    is_superuser = user.is_superuser
    has_academics_role = user_has_permission(user, school, 'academics')
    if is_headteacher or is_superuser or has_academics_role:
        classes_qs = ClassRoom.objects.filter(school=school)
        subjects_qs = Subject.objects.filter(school=school).select_related('pathway', 'education_level')
    elif hasattr(user, 'teacher'):
        teacher = user.teacher
        assigned_class_ids = TeacherAssignment.objects.filter(
            teacher=teacher,
            classroom__school=school,
        ).values_list('classroom_id', flat=True).distinct()
        assigned_subject_ids = TeacherAssignment.objects.filter(
            teacher=teacher,
            classroom__school=school,
        ).values_list('subject_id', flat=True).distinct()
        classes_qs = ClassRoom.objects.filter(school=school, id__in=assigned_class_ids)
        subjects_qs = Subject.objects.filter(school=school, id__in=assigned_subject_ids).select_related('pathway', 'education_level')
    else:
        classes_qs = ClassRoom.objects.none()
        subjects_qs = Subject.objects.none()

    exams_qs = Exam.objects.filter(school=school).order_by('-year', '-id')
    selected_term = (request.GET.get('term') or '').strip()
    selected_exam_id = (request.GET.get('exam_id') or '').strip()
    selected_class_id = (request.GET.get('class_id') or '').strip()
    selected_subject_id = (request.GET.get('subject_id') or '').strip()
    selected_stream_id = (request.GET.get('stream_id') or '').strip()
    selected_out_of = (request.GET.get('out_of') or '100').strip() or '100'

    allowed_class_ids = {str(cid) for cid in classes_qs.values_list('id', flat=True)}
    allowed_subject_ids = {str(sid) for sid in subjects_qs.values_list('id', flat=True)}
    allowed_exam_ids = {str(eid) for eid in exams_qs.values_list('id', flat=True)}

    if selected_class_id and selected_class_id not in allowed_class_ids:
        selected_class_id = ''
    if selected_subject_id and selected_subject_id not in allowed_subject_ids:
        selected_subject_id = ''
    if selected_exam_id and selected_exam_id not in allowed_exam_ids:
        selected_exam_id = ''

    selected_exam = None
    if selected_exam_id:
        selected_exam = exams_qs.filter(id=selected_exam_id).first()
        if selected_exam and not selected_term:
            selected_term = cast(Any, selected_exam).term
    elif exams_qs.exists():
        selected_exam = exams_qs.first()
        if selected_exam:
            selected_exam_id = str(cast(Any, selected_exam).id)
            if not selected_term:
                selected_term = cast(Any, selected_exam).term

    if selected_term not in {'Term 1', 'Term 2', 'Term 3'}:
        selected_term = 'Term 1'

    context = {
        'school': school,
        'terms': ['Term 1', 'Term 2', 'Term 3'],
        'exams': exams_qs,
        'classes': classes_qs,
        'subjects': filter_subjects_for_school(
            school,
            subjects_qs
        ),
        'selected_term': selected_term,
        'selected_exam_id': selected_exam_id,
        'selected_class_id': selected_class_id,
        'selected_subject_id': selected_subject_id,
        'selected_stream_id': selected_stream_id,
        'selected_out_of': selected_out_of,
        'allow_pathways': school_allows_pathways(school),
    }
    return render(request, 'schools/enter_marks.html', context)


@login_required
def school_details(request):
    # Allow headteachers to edit their school; allow superusers to pass ?school_id=<id>
    school = None
    if request.user.is_superuser and request.GET.get('school_id'):
        try:
            school = School.objects.get(pk=int(request.GET.get('school_id')))
        except Exception:
            school = None
    if not school:
        school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()

    if request.method == 'POST':
        # permission: only superuser or headteacher of that school can update
        if not (request.user.is_superuser or (hasattr(request.user, 'headteacher') and request.user.headteacher.school == school)):
            return HttpResponseForbidden()

        form = None
        try:
            from .forms import SchoolDetailsForm
            form = SchoolDetailsForm(request.POST, request.FILES, instance=school)
        except Exception:
            form = None

        # handle base64 image uploads coming from client-side cropping modal
        def handle_base64(field_name, model_field_name):
            data = request.POST.get(field_name)
            if data and data.startswith('data:'):
                # decode
                import base64
                from django.core.files.base import ContentFile
                try:
                    header, encoded = data.split(',', 1)
                    decoded = base64.b64decode(encoded)
                    ext = 'png'
                    if header.startswith('data:image/'):
                        ext = header.split('/')[1].split(';')[0]
                    school_id = cast(Any, school).id
                    fname = f"{model_field_name}_{school_id}.{ext}"
                    getattr(school, model_field_name).save(fname, ContentFile(decoded), save=False)
                except Exception:
                    pass

        # If form valid, save; else if we have base64 fields, handle them and save
        if form and form.is_valid():
            obj = form.save(commit=False)
            # If files were uploaded via form.FILES they'll be attached; additionally handle any base64 fields
            handle_base64('logo_data', 'logo')
            handle_base64('stamp_data', 'stamp')
            handle_base64('signature_data', 'head_signature')
            obj.save()
            messages.success(request, 'School details updated.')
            school_id = cast(Any, school).id
            return redirect(request.path + (f"?school_id={school_id}" if request.user.is_superuser else ''))
        else:
            # attempt to save base64-only submissions (when form not provided)
            handle_base64('logo_data', 'logo')
            handle_base64('stamp_data', 'stamp')
            handle_base64('signature_data', 'head_signature')
            try:
                school.save()
                messages.success(request, 'School images updated.')
                school_id = cast(Any, school).id
                return redirect(request.path + (f"?school_id={school_id}" if request.user.is_superuser else ''))
            except Exception:
                messages.error(request, 'Please correct the errors below.')
    else:
        from .forms import SchoolDetailsForm
        form = SchoolDetailsForm(instance=school)

    return render(request, 'schools/school_details.html', {'school': school, 'form': form})


@login_required
def subject_allocation(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    school_obj = cast(Any, school)
    ensure_cbe_learning_areas(school)
    classes = ClassRoom.objects.filter(school=school).prefetch_related('streams')
    if school_obj.system_type == 'CBE':
        for c in classes:
            level_name = c.level.name if c.level else ''
            if level_name in ('Lower Primary', 'Upper Primary', 'Junior', 'Senior'):
                setattr(c, 'resolved_level_name', level_name)
            elif level_name == 'Primary' or not level_name:
                inferred = _resolve_primary_band_from_class_name(c.name)
                setattr(c, 'resolved_level_name', inferred or 'Lower Primary')
            else:
                setattr(c, 'resolved_level_name', level_name)
    subjects = filter_subjects_for_school(
        school,
        Subject.objects.filter(school=school).select_related('pathway', 'education_level')
    )
    pathways = Pathway.objects.all().order_by('code') if school_allows_pathways(school) else []

    return render(request, "schools/subject_allocation.html", {
        "classes": classes,
        "subjects": subjects,
        "school": school,
        "pathways": pathways,
        "allow_pathways": school_allows_pathways(school),
    })


@login_required
@require_GET
def load_students_for_subject(request):
    """
    Returns students for a class/stream. If `subject_id` is provided, returns only students
    allocated to that subject. Supports optional stream_id filtering.
    Results are limited to the logged-in user's school and role scope.
    """
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()

    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    stream_id = request.GET.get('stream_id')

    if not class_id:
        return JsonResponse({'students': []})

    classroom = ClassRoom.objects.filter(id=class_id, school=school).first()
    if not classroom:
        return JsonResponse({'students': []}, status=404)

    teacher_allowed_stream_ids = set()
    teacher_has_all_streams = True
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        class_pk = cast(Any, classroom).id
        allowed_class_ids = _teacher_assigned_class_ids(teacher, school) | _teacher_class_teacher_class_ids(teacher, school)
        if class_pk not in allowed_class_ids:
            return JsonResponse({'students': [], 'error': 'Not allowed for this class.'}, status=403)

        is_class_teacher_for_class = class_pk in _teacher_class_teacher_class_ids(teacher, school)
        if not is_class_teacher_for_class:
            assignment_qs = TeacherAssignment.objects.filter(
                teacher=teacher,
                classroom_id=class_pk,
                classroom__school=school,
            )
            if subject_id:
                assignment_qs = assignment_qs.filter(subject_id=subject_id)
                if not assignment_qs.exists():
                    return JsonResponse({'students': [], 'error': 'Not allowed for this subject in this class.'}, status=403)
            teacher_has_all_streams = assignment_qs.filter(stream__isnull=True).exists()
            teacher_allowed_stream_ids = set(
                assignment_qs.exclude(stream__isnull=True).values_list('stream_id', flat=True)
            )
            if stream_id and not teacher_has_all_streams:
                try:
                    if int(stream_id) not in teacher_allowed_stream_ids:
                        return JsonResponse({'students': [], 'error': 'Not allowed for selected stream.'}, status=403)
                except (TypeError, ValueError):
                    return JsonResponse({'students': [], 'error': 'Invalid stream selection.'}, status=400)

    students_qs = Student.objects.filter(
        school=school,
        classroom=classroom
    ).select_related('classroom', 'stream').order_by('first_name', 'last_name')

    if stream_id:
        students_qs = students_qs.filter(stream_id=stream_id)
    elif is_teacher and not (is_headteacher or is_superuser):
        if not teacher_has_all_streams and teacher_allowed_stream_ids:
            students_qs = students_qs.filter(stream_id__in=teacher_allowed_stream_ids)

    allocated_ids = set()
    if subject_id:
        allocated_ids = set(
            SubjectAllocation.objects.filter(
                subject_id=subject_id,
                student__school=school,
                student__classroom=classroom
            ).values_list('student_id', flat=True)
        )

    students_data = [
        {
            'id': cast(Any, s).id,
            'name': f"{s.first_name} {s.last_name}",
            'admission_number': s.admission_number,
            'gender': getattr(s, 'gender', ''),
            'stream': str(s.stream) if s.stream else '',
            'allocated': cast(Any, s).id in allocated_ids,
        }
        for s in students_qs
    ]

    return JsonResponse({'students': students_data})


def _get_students_for_class_and_subject(school, class_id, subject_id=None, stream_id=None):
    """Helper returning ordered queryset of students for exports/views."""
    qs = Student.objects.filter(school=school, classroom_id=class_id)
    if stream_id:
        qs = qs.filter(stream_id=stream_id)
    if subject_id:
        allocated_ids = SubjectAllocation.objects.filter(
            subject_id=subject_id,
            student__school=school,
            student__classroom_id=class_id
        ).values_list('student_id', flat=True)
        qs = qs.filter(id__in=allocated_ids)
    return qs.order_by('first_name', 'last_name')


def _class_list_access_scope(request, school, class_id, subject_id=None, stream_id=None):
    classroom = ClassRoom.objects.filter(id=class_id, school=school).first()
    if not classroom:
        return None, None, HttpResponse('Invalid class_id', status=400)

    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if is_headteacher or is_superuser:
        return classroom, None, None
    if not is_teacher:
        return None, None, HttpResponseForbidden()

    teacher = request.user.teacher
    class_pk = cast(Any, classroom).id
    allowed_class_ids = _teacher_assigned_class_ids(teacher, school) | _teacher_class_teacher_class_ids(teacher, school)
    if class_pk not in allowed_class_ids:
        return None, None, HttpResponseForbidden('Not allowed for this class.')

    is_class_teacher_for_class = class_pk in _teacher_class_teacher_class_ids(teacher, school)
    if is_class_teacher_for_class:
        return classroom, None, None

    assignment_qs = TeacherAssignment.objects.filter(
        teacher=teacher,
        classroom_id=class_pk,
        classroom__school=school,
    )
    if subject_id:
        assignment_qs = assignment_qs.filter(subject_id=subject_id)
    if not assignment_qs.exists():
        return None, None, HttpResponseForbidden('Not allowed for this subject/class.')

    has_all_streams = assignment_qs.filter(stream__isnull=True).exists()
    if has_all_streams:
        return classroom, None, None

    allowed_stream_ids = set(assignment_qs.exclude(stream__isnull=True).values_list('stream_id', flat=True))
    if stream_id:
        try:
            if int(stream_id) not in allowed_stream_ids:
                return None, None, HttpResponseForbidden('Not allowed for selected stream.')
        except (TypeError, ValueError):
            return None, None, HttpResponse('Invalid stream_id', status=400)
    return classroom, allowed_stream_ids, None


@login_required
@require_GET
def export_class_list_excel(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    stream_id = request.GET.get('stream_id')

    if not class_id:
        return HttpResponse('Missing class_id', status=400)

    class_obj, teacher_stream_scope, denied = _class_list_access_scope(
        request, school, class_id, subject_id, stream_id
    )
    if denied:
        return denied

    students = _get_students_for_class_and_subject(school, class_id, subject_id, stream_id)
    if teacher_stream_scope and not stream_id:
        students = students.filter(stream_id__in=teacher_stream_scope)

    wb = openpyxl.Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = 'Class List'

    class_name = class_obj.name if class_obj else f'class_{class_id}'
    class_display = class_name
    stream_obj = Stream.objects.filter(id=stream_id, classroom=class_obj).first() if stream_id and class_obj else None
    if stream_obj:
        class_display = f"{class_name} {stream_obj.name}"
    subject_obj = Subject.objects.filter(id=subject_id, school=school).first() if subject_id else None

    generated_on = timezone.localtime(timezone.now()).strftime('%d %b %Y, %I:%M %p')

    # Header block
    school_any = cast(Any, school)
    ws.append([school_any.name])
    ws.append(['WE PLAY, LEARN, GROW'])
    if school_any.phone:
        ws.append([school_any.phone])
    ws.append([f"Class List: {class_display}"])
    ws.append([])

    headers = ['SN', 'Adm No', 'Name', 'MARKS', 'OUT OF', '% (/100)']
    ws.append(headers)
    header_row_idx = ws.max_row

    for i, s in enumerate(students, start=1):
        ws.append([i, s.admission_number or '', f"{s.first_name} {s.last_name}", '', '', ''])

    # Summary rows
    boys_count = students.filter(gender__iexact='male').count()
    girls_count = students.filter(gender__iexact='female').count()
    ws.append([])
    ws.append(['Boys:', '', boys_count])
    ws.append(['Girls:', '', girls_count])

    # Styling
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    header_fill = PatternFill('solid', fgColor='E6F4F1')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            if cell.row == header_row_idx:
                cell.font = bold
                cell.alignment = center
                cell.fill = header_fill

    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(bold=True)
    ws['A4'].font = Font(bold=True)

    # Footer with page number and branding
    footer = cast(Any, ws.oddFooter)
    footer.left.text = f"Generated on {generated_on}"
    footer.center.text = "Page &P of &N"
    footer.right.text = "Skul Plus+"
    
    # Set page setup for portrait A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.print_options.horizontalCentered = False

    for col in ws.columns:
        max_length = 0
        col_idx = col[0].column
        if not col_idx:
            continue
        col_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    subject_name = f"_{subject_obj.name}" if subject_obj else ''

    filename = f"class_list_{class_name}{subject_name}.xlsx"
    resp = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_GET
def export_class_list_pdf(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    stream_id = request.GET.get('stream_id')

    if not class_id:
        return HttpResponse('Missing class_id', status=400)

    class_obj, teacher_stream_scope, denied = _class_list_access_scope(
        request, school, class_id, subject_id, stream_id
    )
    if denied:
        return denied

    students = _get_students_for_class_and_subject(school, class_id, subject_id, stream_id)
    if teacher_stream_scope and not stream_id:
        students = students.filter(stream_id__in=teacher_stream_scope)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)
    width, height = 595, 842  # A4 portrait in points
    top_margin = 20
    bottom_margin = 40
    left_margin = 40
    right_margin = 40
    row_height = 18
    table_width = width - left_margin - right_margin
    col_widths = [30, 80, 180, 70, 70, 85]
    col_positions = [left_margin]
    for w in col_widths:
        col_positions.append(col_positions[-1] + w)

    class_name = class_obj.name if class_obj else f'Class {class_id}'
    stream_obj = Stream.objects.filter(id=stream_id, classroom=class_obj).first() if stream_id and class_obj else None
    class_display = f"{class_name} {stream_obj.name}" if stream_obj else class_name

    generated_on = timezone.localtime(timezone.now()).strftime('%d %b %Y, %I:%M %p')

    def draw_header(page_y):
        # Logo
        school_any = cast(Any, school)
        if school_any.logo and hasattr(school_any.logo, 'path'):
            try:
                p.drawImage(school_any.logo.path, left_margin, page_y - 40, width=50, height=50, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        p.setFont('Helvetica-Bold', 12)
        p.drawCentredString(width / 2, page_y - 10, school_any.name.upper())
        p.setFont('Helvetica', 9)
        p.drawCentredString(width / 2, page_y - 24, 'WE PLAY, LEARN, GROW')
        if school_any.phone:
            p.setFont('Helvetica-Bold', 9)
            p.setFillColorRGB(0.1, 0.55, 0.1)
            p.drawCentredString(width / 2, page_y - 38, school_any.phone)
            p.setFillColorRGB(0, 0, 0)
        p.setFont('Helvetica-Bold', 10)
        p.drawString(left_margin, page_y - 62, 'Class List:')
        p.drawString(left_margin + 110, page_y - 62, class_display)
        return page_y - 80

    def draw_footer(page_num):
        p.setFont('Helvetica', 8)
        p.drawString(left_margin, bottom_margin - 10, f"Generated on {generated_on}")
        p.drawCentredString(width / 2, bottom_margin - 10, f"Page {page_num}")
        p.drawRightString(width - right_margin, bottom_margin - 10, 'Skul Plus+')

    def draw_row_grid(y, fill=False):
        if fill:
            p.setFillColorRGB(0.9, 0.97, 0.96)
            p.rect(left_margin, y - 14, table_width, 16, fill=1, stroke=0)
            p.setFillColorRGB(0, 0, 0)
        p.rect(left_margin, y - 14, table_width, 16, fill=0, stroke=1)
        for x in col_positions[1:-1]:
            p.line(x, y - 14, x, y + 2)

    def draw_table_header(y):
        p.setFont('Helvetica-Bold', 9)
        draw_row_grid(y, fill=True)
        p.drawString(col_positions[0] + 4, y - 10, 'SN')
        p.drawString(col_positions[1] + 4, y - 10, 'Adm No')
        p.drawString(col_positions[2] + 4, y - 10, 'Name')
        p.drawString(col_positions[3] + 4, y - 10, 'MARKS')
        p.drawString(col_positions[4] + 4, y - 10, 'OUT OF')
        p.drawString(col_positions[5] + 4, y - 10, '% (/100)')
        return y - 20

    page_num = 1
    y = height - top_margin
    y = draw_header(y)
    y = draw_table_header(y)

    # Table rows
    for i, s in enumerate(students, start=1):
        if y < bottom_margin + 60:
            draw_footer(page_num)
            p.showPage()
            page_num += 1
            y = height - top_margin
            y = draw_header(y)
            y = draw_table_header(y)
        p.setFont('Helvetica', 9)
        draw_row_grid(y)
        p.drawString(col_positions[0] + 4, y - 10, str(i))
        p.drawString(col_positions[1] + 4, y - 10, s.admission_number or '')
        p.drawString(col_positions[2] + 4, y - 10, f"{s.first_name} {s.last_name}")
        y -= row_height

    # Summary rows
    boys_count = students.filter(gender__iexact='male').count()
    girls_count = students.filter(gender__iexact='female').count()

    if y < bottom_margin + 60:
        draw_footer(page_num)
        p.showPage()
        page_num += 1
        y = height - top_margin
        y = draw_header(y)
        y = draw_table_header(y)

    p.setFont('Helvetica-Bold', 9)
    draw_row_grid(y)
    p.drawString(col_positions[1] + 4, y - 10, 'Boys:')
    p.drawString(col_positions[2] + 4, y - 10, str(boys_count))
    y -= row_height

    draw_row_grid(y)
    p.drawString(col_positions[1] + 4, y - 10, 'Girls:')
    p.drawString(col_positions[2] + 4, y - 10, str(girls_count))

    draw_footer(page_num)
    p.showPage()
    p.save()
    buffer.seek(0)

    class_name_safe = class_display.replace(' ', '_')
    filename = f"class_list_{class_name_safe}.pdf"
    return HttpResponse(buffer.read(), content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="{filename}"'
    })


@login_required
@require_GET
def print_class_list(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    stream_id = request.GET.get('stream_id')

    if not class_id:
        return HttpResponse('Missing class_id', status=400)

    class_obj, teacher_stream_scope, denied = _class_list_access_scope(
        request, school, class_id, subject_id, stream_id
    )
    if denied:
        return denied

    students = _get_students_for_class_and_subject(school, class_id, subject_id, stream_id)
    if teacher_stream_scope and not stream_id:
        students = students.filter(stream_id__in=teacher_stream_scope)

    class_name = class_obj.name if class_obj else f'Class {class_id}'
    stream_obj = Stream.objects.filter(id=stream_id, classroom=class_obj).first() if stream_id and class_obj else None
    subject_obj = Subject.objects.filter(id=subject_id, school=school).first() if subject_id else None
    class_display = f"{class_name} {stream_obj.name}" if stream_obj else class_name
    boys_count = students.filter(gender__iexact='male').count()
    girls_count = students.filter(gender__iexact='female').count()
    generated_on = timezone.localtime(timezone.now()).strftime('%d %b %Y, %I:%M %p')

    context = {
        'students': students,
        'class_name': class_display,
        'subject': subject_obj,
        'school': school,
        'generated_on': generated_on,
        'boys_count': boys_count,
        'girls_count': girls_count,
    }
    response = render(request, 'schools/class_list_print.html', context)
    response['X-Frame-Options'] = 'SAMEORIGIN'
    return response


@login_required
def save_subject_allocations(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)
    data = json.loads(request.body)

    class_id = data.get("class_id")
    subject_id = data.get("subject_id")
    stream_id = data.get("stream_id")
    selected_students = set(map(int, data.get("students", [])))
    allocate_all = bool(data.get('allocate_all'))

    subject_obj = Subject.objects.filter(id=subject_id, school=school).first() if subject_id else None
    if not subject_obj:
        return JsonResponse({"success": False, "error": "Invalid subject."}, status=400)

    classroom = ClassRoom.objects.filter(id=class_id, school=school).first()
    if not classroom:
        return JsonResponse({"success": False, "error": "Invalid class."}, status=400)

    students = Student.objects.filter(classroom_id=class_id, school=school)
    
    # Filter by stream if provided
    if stream_id:
        students = students.filter(stream_id=stream_id)

    # Get stream object if stream_id is provided
    stream_obj = None
    if stream_id:
        try:
            stream_obj = Stream.objects.get(id=stream_id)
        except Stream.DoesNotExist:
            pass

    def validate_cbe(student):
        school_any = cast(Any, school)
        if school_any.system_type != 'CBE':
            return None

        level_name_raw = classroom.level.name if classroom.level else None
        level_name = resolve_cbe_level(school, level_name_raw)
        if level_name == 'Primary' and classroom.name:
            inferred = _resolve_primary_band_from_class_name(classroom.name)
            if inferred:
                level_name = inferred
        subject_level = subject_obj.education_level.name if subject_obj.education_level else None

        if level_name and hasattr(school_any, 'allows_level') and not school_any.allows_level(level_name):
            return f"This school is not configured for {level_name} classes."

        if level_name and subject_level and subject_level not in (level_name, 'Primary'):
            return f"Subject '{subject_obj.name}' is not allowed for {level_name} level."

        if level_name == 'Senior':
            if not school_allows_pathways(school):
                return "Senior pathway features are disabled for this school."
            try:
                from academics.models import StudentPathway
                pathway = StudentPathway.objects.filter(student=student).first()
            except Exception:
                pathway = None

            if not pathway:
                return f"Student {student.admission_number} must select a pathway before subject registration."

            if subject_obj.pathway and pathway:
                subject_any = cast(Any, subject_obj)
                pathway_any = cast(Any, pathway)
                if pathway_any.pathway_id != subject_any.pathway_id:
                    return f"Student {student.admission_number} cannot register subjects outside their pathway."

        if level_name == 'Junior':
            if subject_level != 'Junior':
                return "Junior students can only register Junior learning areas."
            if not is_junior_subject_name(subject_obj.name):
                return "This subject is not part of the Junior learning areas."

        if level_name in ('Lower Primary', 'Upper Primary'):
            if subject_level != level_name:
                return f"{level_name} students can only register {level_name} learning areas."
            if not is_primary_subject_name(subject_obj.name):
                return "This subject is not part of the Primary learning areas."

        return None

    if allocate_all:
        for s in students:
            error = validate_cbe(s)
            if error:
                return JsonResponse({"success": False, "error": error}, status=400)
            try:
                SubjectAllocation.objects.get_or_create(
                    subject_id=subject_id,
                    student=s,
                    defaults={
                        'classroom': s.classroom,
                        'stream': stream_obj or s.stream,
                        'admission_number': s.admission_number or '',
                        'student_name': f"{s.first_name} {s.last_name}"
                    }
                )
            except Exception as exc:
                return JsonResponse({"success": False, "error": str(exc)}, status=400)
        return JsonResponse({"success": True})

    # Remove unselected (only within this class/stream)
    SubjectAllocation.objects.filter(
        subject_id=subject_id,
        student__in=students
    ).exclude(student_id__in=selected_students).delete()

    # Add selected
    for student_id in selected_students:
        try:
            s = Student.objects.get(id=student_id, school=school)
        except Student.DoesNotExist:
            continue
        error = validate_cbe(s)
        if error:
            return JsonResponse({"success": False, "error": error}, status=400)
        SubjectAllocation.objects.get_or_create(
            subject_id=subject_id,
            student=s,
            defaults={
                'classroom': s.classroom,
                'stream': stream_obj or s.stream,
                'admission_number': s.admission_number or '',
                'student_name': f"{s.first_name} {s.last_name}"
            }
        )

    return JsonResponse({"success": True})


@login_required
def classes_view(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)

    if request.method == 'POST':
        form = ClassRoomForm(request.POST, school=school)
        if form.is_valid():
            cls = form.save(commit=False)
            cls.school = school
            cls.save()
            return redirect('classes')
    else:
        form = ClassRoomForm(school=school)

    if 'level' in form.fields:
        school_any = cast(Any, school)
        allowed: list[str] = []
        if getattr(school_any, 'school_type', '') == 'CAMBRIDGE':
            allowed = ['Kindergarten', 'Lower Primary', 'Upper Primary', 'Lower Secondary', 'Upper Secondary (IGCSE)', 'A Level']
        elif school_any.system_type == 'CBE':
            if school_any.school_category == 'PRIMARY':
                allowed = ['Pre School', 'Lower Primary', 'Upper Primary']
            elif school_any.school_category == 'JUNIOR':
                allowed = ['Junior']
            elif school_any.school_category == 'SENIOR':
                allowed = ['Senior']
            elif school_any.school_category == 'COMPREHENSIVE':
                allowed = ['Pre School', 'Lower Primary', 'Upper Primary', 'Junior']
        if allowed:
            cast(Any, form.fields['level']).queryset = EducationLevel.objects.filter(name__in=allowed).order_by('name')
        else:
            cast(Any, form.fields['level']).queryset = EducationLevel.objects.all().order_by('name')

    # annotate student counts (student_set) for display
    from django.db.models import Count
    classes = ClassRoom.objects.filter(school=school).annotate(students_count=Count('student'))

    return render(request, 'schools/classes.html', {'classes': classes, 'form': form})


@login_required
def classes_management(request):
    """Headteacher-facing classes list + create + stream management (AJAX)"""
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    school_obj = cast(Any, school)

    # POST -> create new class or manage streams (AJAX preferred)
    if request.method == 'POST':
        # accept JSON or form-encoded
        try:
            data = json.loads(request.body) if request.body else request.POST
        except Exception:
            data = request.POST

        action = data.get('action')

        # Handle class creation
        if not action or action == 'create_class':
            name = (data.get('name') or '').strip()
            section = (data.get('section') or '').strip()
            level_id = data.get('level')
            class_teacher_id = data.get('class_teacher')

            if not name:
                return JsonResponse({'success': False, 'error': 'Name required'}, status=400)
            if ClassRoom.objects.filter(school=school, name__iexact=name, section__iexact=section).exists():
                return JsonResponse({'success': False, 'error': 'Class with this name/stream already exists.'}, status=400)

            cls = ClassRoom(name=name, section=section, school=school)
            if level_id:
                try:
                    level_obj = EducationLevel.objects.get(id=level_id)
                    if getattr(school_obj, 'school_type', '') == 'CAMBRIDGE':
                        allowed_cambridge = ['Kindergarten', 'Lower Primary', 'Upper Primary', 'Lower Secondary', 'Upper Secondary (IGCSE)', 'A Level']
                        if level_obj.name not in allowed_cambridge:
                            return JsonResponse({'success': False, 'error': f"{level_obj.name} level is not allowed for Cambridge schools."}, status=400)
                    elif school_obj.system_type == 'CBE' and hasattr(school_obj, 'allows_level') and not school_obj.allows_level(level_obj.name):
                        return JsonResponse({'success': False, 'error': f"{level_obj.name} level is not allowed for this school."}, status=400)
                    cast(Any, cls).level = level_obj
                except EducationLevel.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Invalid level selected.'}, status=400)
            if class_teacher_id:
                try:
                    teacher = Teacher.objects.get(id=int(class_teacher_id), school=school)
                    cls.class_teacher = teacher
                except (TypeError, ValueError, Teacher.DoesNotExist):
                    return JsonResponse({'success': False, 'error': 'Invalid class teacher selected.'}, status=400)

            cls.save()
            cls_any = cast(Any, cls)
            return JsonResponse({
                'success': True,
                'class': {
                    'id': cls_any.id,
                    'name': cls_any.name,
                    'section': cls_any.section,
                    'students_count': cls_any.student_set.count(),
                    'streams': []
                }
            })

        # Handle stream creation
        elif action == 'create_stream':
            class_id = data.get('class_id')
            stream_name = data.get('stream_name')
            stream_code = data.get('stream_code', f'STR_{stream_name}')

            if not class_id or not stream_name:
                return JsonResponse({'success': False, 'error': 'Class ID and stream name required'}, status=400)

            try:
                classroom = ClassRoom.objects.get(id=class_id, school=school)
            except ClassRoom.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Class not found'}, status=404)

            stream, created = Stream.objects.get_or_create(
                classroom=classroom,
                name=stream_name,
                defaults={'code': stream_code}
            )

            if not created:
                return JsonResponse({
                    'success': False,
                    'error': f'Stream "{stream_name}" already exists in this class'
                }, status=400)

            stream_any = cast(Any, stream)
            return JsonResponse({
                'success': True,
                'stream': {
                    'id': stream_any.id,
                    'name': stream_any.name,
                    'code': stream_any.code
                },
                'message': 'Stream created successfully'
            })

        # Handle stream deletion
        elif action == 'delete_stream':
            stream_id = data.get('stream_id')

            if not stream_id:
                return JsonResponse({'success': False, 'error': 'Stream ID required'}, status=400)

            try:
                stream = Stream.objects.get(id=stream_id, classroom__school=school)
                stream_name = stream.name
                stream.delete()
                return JsonResponse({
                    'success': True,
                    'message': f'Stream "{stream_name}" deleted successfully'
                })
            except Stream.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Stream not found'}, status=404)

        # Handle stream update/edit
        elif action == 'edit_stream':
            stream_id = data.get('stream_id')
            stream_name = data.get('stream_name')
            stream_code = data.get('stream_code')

            if not stream_id or not stream_name:
                return JsonResponse({'success': False, 'error': 'Stream ID and name required'}, status=400)

            try:
                stream = Stream.objects.get(id=stream_id, classroom__school=school)
                stream.name = stream_name
                if stream_code:
                    stream.code = stream_code
                stream.save()
                stream_any = cast(Any, stream)
                return JsonResponse({
                    'success': True,
                    'stream': {
                        'id': stream_any.id,
                        'name': stream_any.name,
                        'code': stream_any.code
                    },
                    'message': 'Stream updated successfully'
                })
            except Stream.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Stream not found'}, status=404)

        return JsonResponse({'success': False, 'error': 'Unknown action'}, status=400)

    # GET -> render list with streams
    classes = ClassRoom.objects.filter(school=school).annotate(
        students_count=Count('student'),
        streams_count=Count('streams')
    ).prefetch_related('streams')
    form = ClassRoomForm(school=school)
    if 'level' in form.fields:
        allowed: list[str] = []
        if getattr(school, 'school_type', '') == 'CAMBRIDGE':
            allowed = ['Kindergarten', 'Lower Primary', 'Upper Primary', 'Lower Secondary', 'Upper Secondary (IGCSE)', 'A Level']
        elif school.system_type == 'CBE':
            if school.school_category == 'PRIMARY':
                allowed = ['Pre School', 'Lower Primary', 'Upper Primary']
            elif school.school_category == 'JUNIOR':
                allowed = ['Junior']
            elif school.school_category == 'SENIOR':
                allowed = ['Senior']
            elif school.school_category == 'COMPREHENSIVE':
                allowed = ['Pre School', 'Lower Primary', 'Upper Primary', 'Junior']
        if allowed:
            cast(Any, form.fields['level']).queryset = EducationLevel.objects.filter(name__in=allowed).order_by('name')
        else:
            cast(Any, form.fields['level']).queryset = EducationLevel.objects.all().order_by('name')
    return render(request, 'schools/classes.html', {'classes': classes, 'form': form})


@login_required
def delete_class_headteacher(request, pk):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    school_class = get_object_or_404(ClassRoom, pk=pk, school=school)
    if request.method == 'POST':
        school_class.delete()
        return JsonResponse({'success': True})
    return HttpResponseForbidden()


@login_required
def export_classes_excel(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    classes = ClassRoom.objects.filter(school=school)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="classes.xlsx"'

    wb = openpyxl.Workbook()
    ws = cast(Worksheet, wb.active)
    ws.append(['Class Name', 'Section', 'Class Teacher', 'Students'])

    for c in classes:
        c_any = cast(Any, c)
        ws.append([
            c_any.name,
            c_any.section,
            str(c_any.class_teacher) if c_any.class_teacher else '',
            c_any.student_set.count()
        ])

    wb.save(response)
    return response


@login_required
def export_classes_pdf(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    classes = ClassRoom.objects.filter(school=school)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="classes.pdf"'

    p = canvas.Canvas(cast(IO[bytes], response))
    y = 800
    title = f"{school.name if school else ''} - Classes List"
    p.setFont('Helvetica-Bold', 14)
    p.drawString(50, y, title)
    y -= 30
    p.setFont('Helvetica', 11)
    for c in classes:
        c_any = cast(Any, c)
        line = f"{c_any.name} {c_any.section} | Teacher: {c_any.class_teacher or '-'} | Students: {c_any.student_set.count()}"
        p.drawString(50, y, line)
        y -= 18
        if y < 50:
            p.showPage()
            y = 800
    p.save()
    return response


@login_required
def students_page(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    students = Student.objects.filter(school=school).select_related(
        'classroom__level', 'stream'
    ).select_related('studentpathway__pathway').order_by('last_name', 'first_name')

    form = StudentForm(school=school)

    return render(request, 'schools/students_page.html', {
        'students': students,
        'form': form,
        'school': school,
        'allow_pathways': school_allows_pathways(school),
    })


@login_required
def admit_student_ajax(request):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    form = StudentForm(request.POST, request.FILES, school=school)

    if form.is_valid():
        student = form.save(commit=False)
        student.school = school
        student.save()
        parent_username, parent_password, parent_error = _sync_parent_account_for_student(student)
        if parent_error:
            return JsonResponse({'success': False, 'errors': {'parent_phone': [parent_error]}}, status=400)
        try:
            photo_url = student.photo.url if student.photo else ''
        except Exception:
            photo_url = ''

        student_data = {
            'id': student.id,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'admission_number': student.admission_number,
            'classroom': str(student.classroom) if student.classroom else '',
            'parent_name': getattr(student, 'parent_name', '') or '',
            'parent_phone': getattr(student, 'parent_phone', '') or '',
            'parent_username': parent_username,
            'parent_password': parent_password,
            'photo_url': photo_url,
        }
        return JsonResponse({'success': True, 'student': student_data})

    return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@login_required
def delete_class(request, class_id):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    try:
        cls = ClassRoom.objects.get(id=class_id, school=school)
    except ClassRoom.DoesNotExist:
        return HttpResponseForbidden()

    if request.method == 'POST':
        cls.delete()
        return redirect('classes')

    # simple confirm page
    return render(request, 'schools/delete_class_confirm.html', {'class': cls})


@login_required
def delete_student(request, student_id):
    school, denied = _require_school_permission(request, 'students')
    if denied:
        return denied
    try:
        student = Student.objects.get(id=student_id, school=school)
    except Student.DoesNotExist:
        return HttpResponseForbidden()

    if request.method == 'POST':
        student.delete()
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({'success': True})
        return redirect('students_page')

    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if is_ajax and request.method == 'GET':
        return render(request, 'schools/_delete_student_confirm.html', {'student': student})

    return render(request, 'schools/delete_student_confirm.html', {'student': student})


@login_required
def exams(request):
    # Render exams list for settings/exams/ (keeps parity with exams_management)
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    exams = Exam.objects.filter(school=school).order_by('-year', 'term')
    return render(request, 'schools/exams.html', {'exams': exams})


@login_required
def grading(request):
    # Manage grade scales per-school (AJAX-enabled)
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    read_only = False
    is_cambridge = getattr(school, 'school_type', '') == 'CAMBRIDGE'
    show_junior_section = bool((not is_cambridge) and school and school.allows_level('Junior'))

    section_prefixes = {
        'LOWER_PRIMARY': 'L',
        'UPPER_PRIMARY': 'U',
        'JUNIOR': 'J',
        'GENERAL': '',
    }

    def pack_grade(section: str, grade_label: str) -> str:
        prefix = section_prefixes.get(section, '')
        return f"{prefix}{grade_label}"

    def unpack_grade(raw_grade: str) -> tuple[str, str]:
        if raw_grade.startswith('L'):
            return ('LOWER_PRIMARY', raw_grade[1:])
        if raw_grade.startswith('U'):
            return ('UPPER_PRIMARY', raw_grade[1:])
        if raw_grade.startswith('J'):
            return ('JUNIOR', raw_grade[1:])
        return ('GENERAL', raw_grade)

    # POST - create new grade scale (AJAX or normal)
    if request.method == 'POST':
        if read_only:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Grade scales are read-only for Lower/Upper Primary.'})
            messages.error(request, 'Grade scales are read-only for Lower/Upper Primary.')
            return redirect('grading')
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        try:
            data = json.loads(request.body) if is_ajax else request.POST
            min_score = float(data.get('min_score') or 0)
            max_score = float(data.get('max_score') or 0)
            grade = (data.get('grade') or '').strip()
            points = int(data.get('points') or 0)
            section = (data.get('section') or 'GENERAL').strip()
            if is_cambridge:
                section = 'GENERAL'
            if min_score < 0 or max_score < 0 or max_score < min_score:
                raise ValueError('Invalid score range')
            if not grade:
                raise ValueError('Grade label required')
            if points < 0:
                raise ValueError('Points must be zero or greater')
            if section not in section_prefixes:
                raise ValueError('Invalid grading section')
            if len(pack_grade(section, grade)) > 5:
                raise ValueError('Grade label too long for this section')

            gs = GradeScale.objects.create(
                school=school,
                min_score=min_score,
                max_score=max_score,
                grade=pack_grade(section, grade),
                points=points,
            )
            if is_ajax:
                gs_any = cast(Any, gs)
                return JsonResponse({'success': True, 'grade': {'id': gs_any.id, 'min_score': gs_any.min_score, 'max_score': gs_any.max_score, 'grade': gs_any.grade, 'points': gs_any.points}})
            messages.success(request, 'Grade scale created')
            return redirect('grading')
        except Exception as e:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'Error: {e}')
            return redirect('grading')

    # Seed section scales from fixed defaults if missing, so each section is editable in-place.
    if is_cambridge:
        if not GradeScale.objects.filter(school=school).exists():
            for min_score, max_score, grade_label, points in _default_cambridge_bands(school):
                GradeScale.objects.create(
                    school=school,
                    min_score=float(min_score),
                    max_score=float(max_score),
                    grade=str(grade_label),
                    points=int(points),
                )
    else:
        default_section_bands = {
            'LOWER_PRIMARY': LOWER_PRIMARY_LEVEL_BANDS,
            'UPPER_PRIMARY': UPPER_PRIMARY_LEVEL_BANDS,
        }
        if show_junior_section:
            default_section_bands['JUNIOR'] = JUNIOR_LEVEL_BANDS

        for section_key, bands in default_section_bands.items():
            section_prefix = section_prefixes.get(section_key, '')
            exists = GradeScale.objects.filter(school=school, grade__startswith=section_prefix).exists()
            if not exists:
                for min_score, max_score, level, _points in bands:
                    packed_grade = pack_grade(section_key, level)
                    if len(packed_grade) > 5:
                        continue
                    GradeScale.objects.create(
                        school=school,
                        min_score=float(min_score),
                        max_score=float(max_score),
                        grade=packed_grade,
                        points=int(_points),
                    )

    raw_scales = GradeScale.objects.filter(school=school).order_by('-min_score')
    sectioned_scales = {
        'LOWER_PRIMARY': [],
        'UPPER_PRIMARY': [],
        'JUNIOR': [],
        'GENERAL': [],
    }
    for s in raw_scales:
        s_any = cast(Any, s)
        section, label = unpack_grade(s_any.grade or '')
        sectioned_scales.setdefault(section, []).append({
            'id': s_any.id,
            'grade': label,
            'raw_grade': s_any.grade,
            'min_score': s_any.min_score,
            'max_score': s_any.max_score,
            'points': s_any.points,
            'section': section,
        })

    return render(request, 'schools/grading.html', {
        'scales': raw_scales,
        'sectioned_scales': sectioned_scales,
        'read_only': read_only,
        'show_junior_section': show_junior_section,
        'junior_bands': JUNIOR_LEVEL_BANDS,
        'lower_primary_bands': LOWER_PRIMARY_LEVEL_BANDS,
        'upper_primary_bands': UPPER_PRIMARY_LEVEL_BANDS,
        'is_cambridge': is_cambridge,
        'cambridge_scheme': getattr(school, 'cambridge_grading_system', 'CAMB_9_1'),
    })


@login_required
def cbc_comments_report(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()

    if request.method == 'POST':
        result = _seed_missing_primary_comments(school)
        messages.success(request, f"Seeded {result['added']} comments.")
        return redirect('cbc_comments_report')

    report = _build_primary_comment_report(school)
    return render(request, 'schools/cbc_comments_report.html', report)


@login_required
@require_POST
def edit_grade_scale(request, pk):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    if school and school.school_type != 'CAMBRIDGE' and school.school_category == 'PRIMARY':
        return JsonResponse({'success': False, 'error': 'Grade scales are read-only for Lower/Upper Primary.'})
    try:
        gs = get_object_or_404(GradeScale, pk=pk, school=school)
        section_prefixes = {
            'LOWER_PRIMARY': 'L',
            'UPPER_PRIMARY': 'U',
            'JUNIOR': 'J',
            'GENERAL': '',
        }
        data = json.loads(request.body)
        min_score = float(data.get('min_score') or 0)
        max_score = float(data.get('max_score') or 0)
        grade = (data.get('grade') or '').strip()
        points = int(data.get('points') or 0)
        section = (data.get('section') or 'GENERAL').strip()
        if school and school.school_type == 'CAMBRIDGE':
            section = 'GENERAL'
        if min_score < 0 or max_score < 0 or max_score < min_score:
            return JsonResponse({'success': False, 'error': 'Invalid score range'})
        if not grade:
            return JsonResponse({'success': False, 'error': 'Grade label required'})
        if points < 0:
            return JsonResponse({'success': False, 'error': 'Points must be zero or greater'})
        if section not in section_prefixes:
            return JsonResponse({'success': False, 'error': 'Invalid grading section'})
        packed_grade = f"{section_prefixes[section]}{grade}"
        if len(packed_grade) > 5:
            return JsonResponse({'success': False, 'error': 'Grade label too long for this section'})
        gs.min_score = min_score
        gs.max_score = max_score
        gs.grade = packed_grade
        gs.points = points
        gs.save()
        gs_any = cast(Any, gs)
        return JsonResponse({'success': True, 'grade': {'id': gs_any.id, 'min_score': gs_any.min_score, 'max_score': gs_any.max_score, 'grade': gs_any.grade, 'points': gs_any.points}})
    except GradeScale.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def delete_grade_scale(request, pk):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    if school and school.school_type != 'CAMBRIDGE' and school.school_category == 'PRIMARY':
        return JsonResponse({'success': False, 'error': 'Grade scales are read-only for Lower/Upper Primary.'})
    try:
        gs = get_object_or_404(GradeScale, pk=pk, school=school)
        gs.delete()
        return JsonResponse({'success': True})
    except GradeScale.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'})


@login_required
def subjects(request):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()

    school = get_user_school(request.user)
    ensure_cbe_learning_areas(school)
    from schools.models import MarkSheet

    # handle subject creation via AJAX POST
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            subj = form.save(commit=False)
            subj.school = school
            subj.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'id': subj.id})
            return redirect('subjects')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    form = SubjectForm()

    # Include normal school-visible subjects plus any legacy subjects already used in marksheets.
    base_subjects = Subject.objects.filter(school=school).select_related('pathway', 'education_level')
    visible_ids = set(
        filter_subjects_for_school(school, base_subjects).values_list('id', flat=True)
    )
    in_use_ids = set(
        MarkSheet.objects.filter(exam__school=school).values_list('subject_id', flat=True).distinct()
    )
    subject_ids = visible_ids.union(in_use_ids)
    subjects_qs = Subject.objects.filter(id__in=subject_ids).select_related('pathway', 'education_level').annotate(
        class_count=Value(0, output_field=IntegerField())
    ).order_by('name')

    return render(request, 'schools/subjects.html', {
        'subjects': subjects_qs,
        'form': form,
        'school': school,
        'allow_pathways': school_allows_pathways(school),
    })


@login_required
def edit_subject(request, subject_id):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    subject = get_object_or_404(Subject, id=subject_id, school=school)

    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    # GET: return JSON with subject data for client-side population
    subject_any = cast(Any, subject)
    return JsonResponse({
        'id': subject_any.id,
        'code': subject_any.code,
        'name': subject_any.name,
        'short_name': subject_any.short_name or '',
        'subject_category': subject_any.subject_category or '',
    })


@login_required
def delete_subject(request, subject_id):
    if not hasattr(request.user, 'headteacher'):
        return HttpResponseForbidden()
    if request.method != 'POST':
        return HttpResponseForbidden()
    school = get_user_school(request.user)
    subject = get_object_or_404(Subject, id=subject_id, school=school)
    subject.delete()
    return JsonResponse({'success': True})


@login_required
def merit_lists(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    has_academics_role = user_has_permission(request.user, school, 'academics')
    if not (is_headteacher or is_superuser or is_teacher or has_academics_role):
        return HttpResponseForbidden()
    if is_teacher and not request.user.teacher.is_class_teacher:
        return HttpResponseForbidden('Merit lists are available to class teachers only.')

    ensure_cbe_learning_areas(school)
    classes_qs = ClassRoom.objects.filter(school=school)
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        classes_qs = classes_qs.filter(id__in=_teacher_class_teacher_class_ids(teacher, school))
    context = {
        'school': school,
        'classes': classes_qs.order_by('order', 'name'),
        'exams': Exam.objects.filter(school=school).order_by('-year', 'term', 'title'),
        'terms': ['Term 1', 'Term 2', 'Term 3'],
    }
    return render(request, 'schools/merit_lists.html', context)

# AJAX endpoint to get only exams with marks for a class
from django.views.decorators.http import require_GET
@require_GET
def exams_with_marks_for_class(request):
    school = get_user_school(request.user)
    class_id = request.GET.get('class_id')
    if not class_id:
        return JsonResponse({'exams': []})
    from schools.models import MarkSheet, Exam
    import logging
    logger = logging.getLogger("cbc_debug")
    marksheets = MarkSheet.objects.filter(school_class=class_id, school=school)
    logger.info(f"[AJAX exams-for-class] MarkSheet count for class_id={class_id}, school={school}: {marksheets.count()}")
    exam_ids = marksheets.values_list('exam_id', flat=True).distinct()
    logger.info(f"[AJAX exams-for-class] Exam IDs: {list(exam_ids)}")
    exams = Exam.objects.filter(id__in=exam_ids, school=school).order_by('-year', 'term', 'title')
    # Fallback: if no marksheets exist yet for this class, still allow selecting any school exam.
    if not exams.exists():
        exams = Exam.objects.filter(school=school).order_by('-year', 'term', 'title')
    logger.info(f"[AJAX exams-for-class] Exams returned: {[e.id for e in exams]}")
    exam_list = [
        {'id': e.id, 'title': e.title, 'term': e.term, 'year': e.year}
        for e in exams
    ]
    return JsonResponse({'exams': exam_list})

# generate_merit_list
@login_required
def merit_lists_data(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()
    if is_teacher and not request.user.teacher.is_class_teacher:
        return HttpResponseForbidden('Merit lists are available to class teachers only.')

    class_id = request.GET.get('class_id')
    exam_id = request.GET.get('exam_id')
    stream_id = request.GET.get('stream_id')
    term = (request.GET.get('term') or '').strip()

    # DEBUG: Log CBC detection variables
    import logging
    logger = logging.getLogger("cbc_debug")

    if not class_id or not exam_id:
        return JsonResponse({'success': False, 'error': 'Please select both class and exam.'}, status=400)

    ensure_cbe_learning_areas(school)

    classroom = get_object_or_404(ClassRoom, id=class_id, school=school)
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
            teacher, school, cast(Any, classroom).id
        )
        if not allowed:
            return JsonResponse({'success': False, 'error': 'Not allowed for this class.'}, status=403)
        if stream_id and not all_streams:
            try:
                if int(stream_id) not in allowed_stream_ids:
                    return JsonResponse({'success': False, 'error': 'Not allowed for selected stream.'}, status=403)
            except (TypeError, ValueError):
                return JsonResponse({'success': False, 'error': 'Invalid stream selection.'}, status=400)
    exam = get_object_or_404(Exam, id=exam_id, school=school)
    ensure_cbe_learning_areas(school)
    class_level = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    school_any = cast(Any, school)
    # Force CBC grading for CBE or COMPREHENSIVE schools
    is_cbc_school = (getattr(school_any, 'system_type', None) == 'CBE' or getattr(school_any, 'school_category', None) == 'COMPREHENSIVE')
    is_junior = is_cbc_school and resolved_level == 'Junior'
    is_primary = is_cbc_school and resolved_level in ('Lower Primary', 'Upper Primary')
    show_junior_pathway = False

    term_filter = term or exam.term
    students_qs = Student.objects.filter(school=school, classroom=classroom).select_related('stream')
    if stream_id:
        students_qs = students_qs.filter(stream_id=stream_id)
    elif is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        _allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
            teacher, school, cast(Any, classroom).id
        )
        if not all_streams and allowed_stream_ids:
            students_qs = students_qs.filter(stream_id__in=allowed_stream_ids)
    students = list(students_qs.order_by('admission_number', 'last_name', 'first_name'))
    student_ids = [cast(Any, s).id for s in students]
    if not student_ids:
        return JsonResponse({'success': False, 'error': 'No students found for this selection.'})

    mark_sheets = MarkSheet.objects.filter(
        school_class=classroom,
        exam=exam,
        status='published',
    ).select_related('subject')
    if term_filter:
        mark_sheets = mark_sheets.filter(term=term_filter)

    subject_objs = []
    seen_subjects = set()
    for ms in mark_sheets:
        subj = cast(Any, ms).subject
        if not subj:
            continue
        sid = cast(Any, subj).id
        if sid in seen_subjects:
            continue
        seen_subjects.add(sid)
        short_code = cast(Any, subj).short_name or cast(Any, subj).code or cast(Any, subj).name
        subject_objs.append({
            'id': sid,
            'name': cast(Any, subj).name,
            'code': str(short_code).strip(),
        })
    subject_objs.sort(key=lambda x: x['code'])
    subject_ids = [s['id'] for s in subject_objs]

    def _section_for_class_level() -> str:
        # Explicit level mapping first
        if resolved_level == 'Lower Primary':
            return 'LOWER_PRIMARY'
        if resolved_level == 'Upper Primary':
            return 'UPPER_PRIMARY'
        if resolved_level == 'Junior':
            return 'JUNIOR'

        # Fallback: infer from classroom name (e.g. Grade 1..9)
        class_name = str(getattr(cast(Any, classroom), 'name', '') or '')
        match = re.search(r'(\d+)', class_name)
        if match:
            grade_num = int(match.group(1))
            if 1 <= grade_num <= 3:
                return 'LOWER_PRIMARY'
            if 4 <= grade_num <= 6:
                return 'UPPER_PRIMARY'
            if 7 <= grade_num <= 9:
                return 'JUNIOR'
        return 'GENERAL'

    def _band_section(raw_grade: str) -> tuple[str, str]:
        if raw_grade.startswith('L'):
            return ('LOWER_PRIMARY', raw_grade[1:])
        if raw_grade.startswith('U'):
            return ('UPPER_PRIMARY', raw_grade[1:])
        if raw_grade.startswith('J'):
            return ('JUNIOR', raw_grade[1:])
        return ('GENERAL', raw_grade)

    selected_section = _section_for_class_level()
    all_grade_bands = list(
        GradeScale.objects.filter(school=school).order_by('-min_score').values('min_score', 'max_score', 'grade', 'points')
    )
    section_grade_bands = []
    general_grade_bands = []
    for band in all_grade_bands:
        sec, label = _band_section(str(band['grade'] or ''))
        band_entry = {
            'min_score': float(band['min_score'] or 0),
            'max_score': float(band['max_score'] or 0),
            'grade': label,
            'points': int(band.get('points') or 0),
        }
        if sec == selected_section:
            section_grade_bands.append(band_entry)
        elif sec == 'GENERAL':
            general_grade_bands.append(band_entry)

    def resolve_level_points(percentage_value: float) -> tuple[str, int]:
        # Use selected section first, then GENERAL within same school.
        bands = section_grade_bands if section_grade_bands else general_grade_bands
        if not bands:
            return ('-', 0)

        # 1) Exact range match (preferred)
        for band in bands:
            if band['min_score'] <= percentage_value <= band['max_score']:
                return (str(band['grade']), int(band['points']))

        # 2) VBA-like fallback by descending Min: first Min <= score
        #    This prevents ungraded marks when there are small config gaps.
        sorted_by_min = sorted(bands, key=lambda b: float(b['min_score']), reverse=True)
        for band in sorted_by_min:
            if percentage_value >= float(band['min_score']):
                return (str(band['grade']), int(band['points']))

        # 3) If score is below all mins (e.g., 0 when lowest min is 1), use lowest band.
        lowest_band = sorted_by_min[-1]
        return (str(lowest_band['grade']), int(lowest_band['points']))

    marks = StudentMark.objects.filter(
        marksheet__in=mark_sheets,
        student_id__in=student_ids,
        marksheet__subject_id__in=subject_ids,
        score__isnull=False,
    ).select_related('marksheet', 'marksheet__subject')

    marks_map = defaultdict(dict)
    subject_percentages = defaultdict(list)
    for m in marks:
        m_any = cast(Any, m)
        sid = m_any.student_id
        sub_id = m_any.marksheet.subject_id
        score = float(m_any.score or 0)
        out_of = float(m_any.marksheet.out_of or 0)
        pct = round((score / out_of) * 100, 1) if out_of > 0 else 0
        lvl, pts = resolve_level_points(pct)
        marks_map[sid][sub_id] = {
            'mark': round(pct, 1),
            'level': lvl,
            'points': pts,
            'raw': round(score, 1),
            'out_of': out_of,
        }
        subject_percentages[sub_id].append(pct)

    rows = []
    for s in students:
        s_any = cast(Any, s)
        sub_values = marks_map.get(s_any.id, {})
        total_marks = 0.0
        total_out = 0.0
        for rec in sub_values.values():
            total_marks += float(rec.get('raw') or 0)
            total_out += float(rec.get('out_of') or 0)
        total_pct = round((total_marks / total_out) * 100, 1) if total_out > 0 else 0
        total_lvl, _total_level_points = resolve_level_points(total_pct) if total_out > 0 else ('-', 0)
        total_points = sum(int(v.get('points') or 0) for v in sub_values.values())

        row_subjects = {}
        for subj in subject_objs:
            row_subjects[str(subj['id'])] = sub_values.get(subj['id'], {'mark': '-', 'level': '-', 'points': 0, 'raw': 0, 'out_of': 0})

        rows.append({
            'adm': s_any.admission_number or '',
            'name': f"{s_any.first_name} {s_any.last_name}".strip(),
            'stream': cast(Any, s_any.stream).name if s_any.stream else '',
            'subjects': row_subjects,
            'total': round(total_marks, 1),
            'total_out': round(total_out, 1),
            'total_points': total_points,
            'total_level': total_lvl,
            'total_pct': total_pct,
        })

    rows.sort(key=lambda r: (r['total_points'], r['total'], r['total_pct']), reverse=True)
    prev_total = None
    prev_pct = None
    rank = 0
    for i, r in enumerate(rows, start=1):
        if prev_total == r['total'] and prev_pct == r['total_pct']:
            r['pos'] = rank
        else:
            rank = i
            r['pos'] = rank
            prev_total = r['total']
            prev_pct = r['total_pct']

    sub_mss = {}
    sub_lvl = {}
    for subj in subject_objs:
        vals = subject_percentages.get(subj['id'], [])
        if vals:
            mss = round(sum(vals) / len(vals), 1)
            sub_mss[str(subj['id'])] = mss
            lvl, pts = resolve_level_points(mss)
            sub_lvl[str(subj['id'])] = lvl
        else:
            sub_mss[str(subj['id'])] = '-'
            sub_lvl[str(subj['id'])] = '-'

    distribution_levels = [b['grade'] for b in section_grade_bands] if section_grade_bands else [b['grade'] for b in general_grade_bands]
    level_distribution = {lvl: 0 for lvl in distribution_levels}
    for r in rows:
        lvl = r['total_level']
        if lvl and lvl != '-':
            if lvl not in level_distribution:
                level_distribution[lvl] = 0
            level_distribution[lvl] += 1

    class_mean_pct = round(sum([r['total_pct'] for r in rows]) / len(rows), 2) if rows else 0
    class_mean_level, _cm_pts = resolve_level_points(class_mean_pct) if rows else ('-', 0)

    stream_tables = []
    if not stream_id:
        rows_by_stream = defaultdict(list)
        for r in rows:
            rows_by_stream[r.get('stream') or 'No Stream'].append(dict(r))

        for stream_name in sorted(rows_by_stream.keys()):
            stream_rows = rows_by_stream[stream_name]
            stream_rows.sort(key=lambda r: (r['total_points'], r['total'], r['total_pct']), reverse=True)
            s_prev_total = None
            s_prev_pct = None
            s_rank = 0
            for i, r in enumerate(stream_rows, start=1):
                if s_prev_total == r['total'] and s_prev_pct == r['total_pct']:
                    r['pos'] = s_rank
                else:
                    s_rank = i
                    r['pos'] = s_rank
                    s_prev_total = r['total']
                    s_prev_pct = r['total_pct']

            stream_sub_vals = defaultdict(list)
            for r in stream_rows:
                for subj in subject_objs:
                    rec = (r.get('subjects') or {}).get(str(subj['id']), {})
                    mark_val = rec.get('mark')
                    if isinstance(mark_val, (int, float)):
                        stream_sub_vals[str(subj['id'])].append(float(mark_val))

            stream_sub_mss = {}
            stream_sub_lvl = {}
            for subj in subject_objs:
                sid = str(subj['id'])
                vals = stream_sub_vals.get(sid, [])
                if vals:
                    mss = round(sum(vals) / len(vals), 1)
                    stream_sub_mss[sid] = mss
                    lvl, _pts = resolve_level_points(mss)
                    stream_sub_lvl[sid] = lvl
                else:
                    stream_sub_mss[sid] = '-'
                    stream_sub_lvl[sid] = '-'

            stream_level_distribution = {lvl: 0 for lvl in distribution_levels}
            for r in stream_rows:
                lvl = r.get('total_level')
                if lvl and lvl != '-':
                    if lvl not in stream_level_distribution:
                        stream_level_distribution[lvl] = 0
                    stream_level_distribution[lvl] += 1

            stream_mean_pct = round(sum([r['total_pct'] for r in stream_rows]) / len(stream_rows), 2) if stream_rows else 0
            stream_mean_level, _spts = resolve_level_points(stream_mean_pct) if stream_rows else ('-', 0)

            stream_tables.append({
                'stream': stream_name,
                'students': stream_rows,
                'sub_mss': stream_sub_mss,
                'sub_lvl': stream_sub_lvl,
                'level_distribution': stream_level_distribution,
                'class_mean': {'score': stream_mean_pct, 'level': stream_mean_level},
            })

    return JsonResponse({
        'success': True,
        'school': {
            'name': cast(Any, school).name,
            'motto': getattr(cast(Any, school), 'motto', '') or '',
            'address': getattr(cast(Any, school), 'address', '') or '',
            'phone': getattr(cast(Any, school), 'phone', '') or '',
            'email': getattr(cast(Any, school), 'email', '') or '',
            'logo_url': (cast(Any, school).logo.url if getattr(cast(Any, school), 'logo', None) else ''),
        },
        'report': {
            'class': classroom.name,
            'exam': exam.title,
            'term': term_filter,
            'title': f"CLASS {str(classroom.name).upper()} - {term_filter} {exam.title}",
        },
        'subjects': subject_objs,
        'students': rows,
        'sub_mss': sub_mss,
        'sub_lvl': sub_lvl,
        'level_distribution': dict(level_distribution),
        'level_order': distribution_levels,
        'class_mean': {'score': class_mean_pct, 'level': class_mean_level},
        'stream_tables': stream_tables,
        'summary': {
            'class': classroom.name,
            'exam': exam.title,
            'population': len(rows),
            'subjects': len(subject_objs),
        },
    })

    # DEBUG: Log CBC detection variables
    logger.warning(f"CBC DEBUG: class_level={class_level}, resolved_level={resolved_level}, system_type={getattr(school_any, 'system_type', None)}, is_primary={is_primary}, is_junior={is_junior}")
    ensure_cbe_learning_areas(school)
    class_level = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    school_any = cast(Any, school)
    is_junior = school_any.system_type == 'CBE' and resolved_level == 'Junior'
    is_primary = school_any.system_type == 'CBE' and resolved_level in ('Lower Primary', 'Upper Primary')
    show_junior_pathway = False

    class_level = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    school_any = cast(Any, school)
    is_junior = school_any.system_type == 'CBE' and resolved_level == 'Junior'
    is_primary = school_any.system_type == 'CBE' and resolved_level in ('Lower Primary', 'Upper Primary')
    show_junior_pathway = False

    term_filter = term or exam.term

    # (Removed unreachable/leftover code after merit_lists_data replacement)
    # (Removed all unreachable/indented code after new merit_lists_data logic)

    subject_scores_by_student = defaultdict(dict)
    student_percentage_sum = defaultdict(float)
    student_subject_counts = defaultdict(int)
    student_points_total = defaultdict(int)
    subject_points_by_student = defaultdict(dict)

    for mark in marks:
        mark_any = cast(Any, mark)
        subject = mark_any.marksheet.subject
        out_of_val = mark_any.marksheet.out_of or 0
        score_val = float(mark_any.score)
        percentage = round((score_val / float(out_of_val) * 100), 1) if out_of_val else None

        entry = {
            'score': score_val,
            'percentage': percentage,
        }

        if is_primary:
            level = get_primary_level(percentage, resolved_level)
            points = get_primary_points(level)
            entry.update({'level': level, 'points': points})
            student_points_total[mark_any.student_id] += points
            if subject:
                subject_points_by_student[mark_any.student_id][subject.name] = points
        elif is_junior:
            level = get_junior_level(percentage)
            points = get_junior_points(level)
            entry.update({'level': level, 'points': points})
            student_points_total[mark_any.student_id] += points
            if subject:
                subject_points_by_student[mark_any.student_id][subject.name] = points
        else:
            entry['grade'] = resolve_grade(percentage) if percentage is not None else '--'

        subject_scores_by_student[mark_any.student_id][mark_any.marksheet.subject_id] = entry

        if percentage is not None:
            student_percentage_sum[mark_any.student_id] += percentage
            student_subject_counts[mark_any.student_id] += 1

    results = []
    prev_avg = prev_total = None
    position = 0
    for idx, row in enumerate(marks_qs, start=1):
        student = students_map.get(row['student_id'])
        if not student:
            continue
        avg_score = float(row['average'] or 0)
        total_score = float(row['total'] or 0)
        if prev_avg == avg_score and prev_total == total_score:
            pass
        else:
            position = idx
            prev_avg = avg_score
            prev_total = total_score

        subjects_map = subject_scores_by_student.get(row['student_id'], {})

        average_percentage = None
        if student_subject_counts[row['student_id']]:
            average_percentage = round(
                student_percentage_sum[row['student_id']] / student_subject_counts[row['student_id']],
                2
            )

        row_payload = {
            'position': position,
            'admission': student.admission_number,
            'student': f"{student.first_name} {student.last_name}".upper(),
            'stream': 'M' if student.gender == 'Male' else 'W',
            'subjects_map': subjects_map,
            'total': round(total_score, 2),
            'average': round(avg_score, 2),
        }

        if is_primary:
            avg_level = get_primary_level(average_percentage, resolved_level)
            row_payload.update({
                'average_level': avg_level or '--',
                'average_percentage': average_percentage,
                'recommended_interest': recommend_primary_interest(
                    subject_points_by_student.get(row['student_id'], {})
                ),
            })
        elif is_junior:
            total_points = student_points_total.get(row['student_id'], 0)
            avg_level = get_junior_level(average_percentage)
            row_payload.update({
                'average_level': avg_level or '--',
                'average_percentage': average_percentage,
                'total_points': total_points,
            })
            if show_junior_pathway:
                row_payload['recommended_pathway'] = recommend_junior_pathway(
                    subject_points_by_student.get(row['student_id'], {})
                )
        else:
            row_payload['grade'] = resolve_grade(avg_score)

        results.append(row_payload)

    # Calculate subject means and grade distribution
    subject_means_qs = StudentMark.objects.filter(
        marksheet__in=mark_sheets, score__isnull=False,
    ).values(
        'marksheet__subject_id', 'marksheet__subject__short_name', 'marksheet__out_of'
    ).annotate(mean_score=Avg('score')).order_by('marksheet__subject__short_name')

    subject_means = []
    for row in subject_means_qs:
        out_of_val = row['marksheet__out_of'] or 0
        mean_score = float(row['mean_score'] or 0)
        mean_percentage = round((mean_score / float(out_of_val) * 100), 1) if out_of_val else None
        mean_payload = {
            'short_name': row['marksheet__subject__short_name'],
            'mean_score': round(mean_score, 1),
            'mean_percentage': mean_percentage,
        }
        if is_primary:
            level = get_primary_level(mean_percentage, resolved_level)
            mean_payload.update({
                'level': level or '--',
                'points': get_primary_points(level),
            })
        elif is_junior:
            level = get_junior_level(mean_percentage)
            mean_payload.update({
                'level': level or '--',
                'points': get_junior_points(level),
            })
        else:
            mean_payload['grade'] = resolve_grade(mean_percentage) if mean_percentage is not None else '--'
        subject_means.append(mean_payload)

    # Calculate grade distribution
    grade_distribution = {}
    for result in results:
        if is_primary:
            level = result.get('average_level')
            if level and level != '--':
                grade_distribution[level] = grade_distribution.get(level, 0) + 1
        elif is_junior:
            level = result.get('average_level')
            if level and level != '--':
                grade_distribution[level] = grade_distribution.get(level, 0) + 1
        else:
            grade = result.get('grade')
            if grade and grade != '--':
                grade_distribution[grade] = grade_distribution.get(grade, 0) + 1

    # Calculate class mean grade
    if results:
        if is_primary:
            avg_values = [r.get('average_percentage') for r in results if r.get('average_percentage') is not None]
            class_mean_score = (sum(avg_values) / len(avg_values)) if avg_values else None
            class_mean_grade = get_primary_level(class_mean_score, resolved_level) if class_mean_score is not None else '--'
        elif is_junior:
            avg_values = [r.get('average_percentage') for r in results if r.get('average_percentage') is not None]
            class_mean_score = (sum(avg_values) / len(avg_values)) if avg_values else None
            class_mean_grade = get_junior_level(class_mean_score) if class_mean_score is not None else '--'
        else:
            total_grades = sum([r['average'] for r in results])
            class_mean_score = total_grades / len(results)
            class_mean_grade = resolve_grade(class_mean_score)
    else:
        class_mean_grade = '--'

    # Build PDF with custom footer
    response = HttpResponse(content_type='application/pdf')
    inline = (request.GET.get('inline') or '').strip() == '1'
    disposition = 'inline' if inline else 'attachment'
    response['Content-Disposition'] = f'{disposition}; filename="merit_list_{classroom.name}_{exam.title}.pdf"'

    buffer = io.BytesIO()
    
    # Create custom footer function
    generated_time = datetime.now().strftime('%d-%b-%Y %I:%M %p')
    
    def add_page_footer(canvas_obj, doc):
        canvas_obj.saveState()
        
        # Left - timestamp
        canvas_obj.setFont("Helvetica-Oblique", 9)
        canvas_obj.drawString(0.5*inch, 0.3*inch, f"Generated on {generated_time}")
        
        # Center - page number
        canvas_obj.setFont("Helvetica", 9)
        page_num_text = f"{doc.page}"
        canvas_obj.drawCentredString(5.5*inch, 0.3*inch, page_num_text)
        
        # Right - system name
        canvas_obj.setFont("Helvetica-Oblique", 9)
        system_text = "SKUL PLUS +(student information reporting management system)"
        canvas_obj.drawRightString(10.5*inch, 0.3*inch, system_text)
        
        canvas_obj.restoreState()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4), 
        topMargin=0.3*inch, 
        bottomMargin=0.6*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Create center-aligned style for header
    from reportlab.lib.enums import TA_CENTER
    center_style = styles['Normal'].clone('center_style')
    center_style.alignment = TA_CENTER
    title_style = styles['Title'].clone('title_center')
    title_style.alignment = TA_CENTER
    heading_style = styles['Heading2'].clone('heading_center')
    heading_style.alignment = TA_CENTER

    # Header - Logo on the left, text centered
    logo_cell = ''
    school_any = cast(Any, school)
    if school_any.logo:
        try:
            logo = Image(school_any.logo.path, width=0.9*inch, height=0.9*inch)
            logo.hAlign = 'LEFT'
            logo_cell = logo
        except:
            logo_cell = ''

    header_text = [
        Paragraph(f"<b>{school_any.name}</b>", title_style),
    ]
    if school_any.address:
        header_text.append(Paragraph(f"{school_any.address}", center_style))
    header_text.append(Paragraph(
        f"<b>CLASS {classroom.name.upper()} - {term_filter.upper()} {exam.year} {exam.title.upper()}</b>",
        heading_style,
    ))

    header_table = Table([
        [logo_cell, header_text]
    ], colWidths=[0.9*inch, 9.6*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (0, 0), 10),
        ('RIGHTPADDING', (0, 0), (0, 0), 4),
        ('LEFTPADDING', (1, 0), (1, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.08*inch))

    # Table headers: POS, ADM, NAME, STRM, Subject columns, totals
    table_headers = ['POS', 'ADM', 'NAME', 'STRM']
    if is_primary:
        for subj in subjects_list:
            table_headers.append(subj['subject__short_name'] or subj['subject__name'][:3].upper())
            table_headers.append('LVL')
        table_headers.extend(['AVG LVL', 'INTEREST'])
    elif is_junior:
        for subj in subjects_list:
            table_headers.append(subj['subject__short_name'] or subj['subject__name'][:3].upper())
            table_headers.append('LVL')
            table_headers.append('PTS')
        table_headers.extend(['TOTAL PTS', 'AVG LVL'])
        if show_junior_pathway:
            table_headers.append('REC PATHWAY')
    else:
        for subj in subjects_list:
            table_headers.append(subj['subject__short_name'] or subj['subject__name'][:3].upper())
            table_headers.append('LVL')
        table_headers.extend(['TOTAL', 'LVL'])
    
    table_data = [table_headers]
    
    for row in results:
        row_data = [
            str(row['position']),
            row['admission'],
            row['student'],
            row['stream'],
        ]
        if is_primary:
            for subj in subjects_list:
                subj_data = row['subjects_map'].get(subj['subject_id'])
                if subj_data and subj_data.get('score') is not None:
                    row_data.append(str(int(round(subj_data['score']))))
                    row_data.append(subj_data.get('level') or '--')
                else:
                    row_data.append('-')
                    row_data.append('-')
            row_data.append(row.get('average_level', '--'))
            row_data.append(row.get('recommended_interest', 'GENERAL'))
        elif is_junior:
            for subj in subjects_list:
                subj_data = row['subjects_map'].get(subj['subject_id'])
                if subj_data and subj_data.get('score') is not None:
                    row_data.append(str(int(round(subj_data['score']))))
                    row_data.append(subj_data.get('level') or '--')
                    row_data.append(str(subj_data.get('points', 0)))
                else:
                    row_data.append('-')
                    row_data.append('-')
                    row_data.append('-')
            row_data.append(str(int(round(row.get('total_points', 0)))))
            row_data.append(row.get('average_level', '--'))
            if show_junior_pathway:
                row_data.append(row.get('recommended_pathway', 'GENERAL'))
        else:
            for subj in subjects_list:
                subj_data = row['subjects_map'].get(subj['subject_id'])
                if subj_data and subj_data['score'] is not None:
                    row_data.append(str(int(round(subj_data['score']))))
                    row_data.append(subj_data['grade'])
                else:
                    row_data.append('-')
                    row_data.append('-')
            row_data.append(str(int(round(row['total']))))
            row_data.append(row['grade'])
        table_data.append(row_data)

    # Add subject means row
    subject_means_row = ['', '', 'SUB MSS', '']
    if is_primary:
        for subj_mean in subject_means:
            subject_means_row.append(str(subj_mean['mean_score']))
            subject_means_row.append('-')
        subject_means_row.extend(['', ''])
        table_data.append(subject_means_row)

        subject_level_row = ['', '', 'SUB LVL', '']
        for subj_mean in subject_means:
            subject_level_row.append('-')
            subject_level_row.append(subj_mean.get('level', '--'))
        subject_level_row.extend(['', ''])
        table_data.append(subject_level_row)
    elif is_junior:
        for subj_mean in subject_means:
            subject_means_row.append(str(subj_mean['mean_score']))
            subject_means_row.append('-')
            subject_means_row.append('-')
        subject_means_row.extend(['', ''])
        if show_junior_pathway:
            subject_means_row.append('')
        table_data.append(subject_means_row)

        subject_level_row = ['', '', 'SUB LVL', '']
        for subj_mean in subject_means:
            subject_level_row.append('-')
            subject_level_row.append(subj_mean.get('level', '--'))
            subject_level_row.append('-')
        subject_level_row.extend(['', ''])
        if show_junior_pathway:
            subject_level_row.append('')
        table_data.append(subject_level_row)

        subject_points_row = ['', '', 'SUB PTS', '']
        for subj_mean in subject_means:
            subject_points_row.append('-')
            subject_points_row.append('-')
            subject_points_row.append(str(subj_mean.get('points', 0)))
        subject_points_row.extend(['', ''])
        if show_junior_pathway:
            subject_points_row.append('')
        table_data.append(subject_points_row)
    else:
        for subj_mean in subject_means:
            subject_means_row.append(str(subj_mean['mean_score']))
            subject_means_row.append('-')
        subject_means_row.extend(['', ''])
        table_data.append(subject_means_row)

        # Add subject level row
        subject_level_row = ['', '', 'SUB LVL', '']
        for subj_mean in subject_means:
            subject_level_row.append('-')
            subject_level_row.append(subj_mean['grade'])
        subject_level_row.extend(['', ''])
        table_data.append(subject_level_row)

    table = Table(table_data)
    base_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        # Highlight top 3 students
        ('BACKGROUND', (0, 1), (-1, 1), colors.lightyellow),
        ('BACKGROUND', (0, 2), (-1, 2), colors.lightyellow),
        ('BACKGROUND', (0, 3), (-1, 3), colors.lightyellow),
    ]

    # Style LVL columns and subject mark columns
    lvl_bg = colors.HexColor('#FFFBEA')
    lvl_text = colors.HexColor('#1D4ED8')
    subject_count = len(subjects_list)
    start_col = 4
    if is_junior:
        for idx in range(subject_count):
            mark_col = start_col + idx * 3
            lvl_col = mark_col + 1
            base_styles.append(('FONTNAME', (mark_col, 1), (mark_col, -1), 'Helvetica-Bold'))
            base_styles.append(('BACKGROUND', (lvl_col, 1), (lvl_col, -1), lvl_bg))
            base_styles.append(('TEXTCOLOR', (lvl_col, 1), (lvl_col, -1), lvl_text))

        total_points_col = start_col + subject_count * 3
        avg_level_col = total_points_col + 1
        base_styles.append(('BACKGROUND', (avg_level_col, 1), (avg_level_col, -1), lvl_bg))
        base_styles.append(('TEXTCOLOR', (avg_level_col, 1), (avg_level_col, -1), lvl_text))
    elif is_primary:
        for idx in range(subject_count):
            mark_col = start_col + idx * 2
            lvl_col = mark_col + 1
            base_styles.append(('FONTNAME', (mark_col, 1), (mark_col, -1), 'Helvetica-Bold'))
            base_styles.append(('BACKGROUND', (lvl_col, 1), (lvl_col, -1), lvl_bg))
            base_styles.append(('TEXTCOLOR', (lvl_col, 1), (lvl_col, -1), lvl_text))

        avg_level_col = start_col + subject_count * 2
        base_styles.append(('BACKGROUND', (avg_level_col, 1), (avg_level_col, -1), lvl_bg))
        base_styles.append(('TEXTCOLOR', (avg_level_col, 1), (avg_level_col, -1), lvl_text))
    else:
        for idx in range(subject_count):
            mark_col = start_col + idx * 2
            lvl_col = mark_col + 1
            base_styles.append(('FONTNAME', (mark_col, 1), (mark_col, -1), 'Helvetica-Bold'))
            base_styles.append(('BACKGROUND', (lvl_col, 1), (lvl_col, -1), lvl_bg))
            base_styles.append(('TEXTCOLOR', (lvl_col, 1), (lvl_col, -1), lvl_text))

        total_col = start_col + subject_count * 2
        total_lvl_col = total_col + 1
        base_styles.append(('BACKGROUND', (total_lvl_col, 1), (total_lvl_col, -1), lvl_bg))
        base_styles.append(('TEXTCOLOR', (total_lvl_col, 1), (total_lvl_col, -1), lvl_text))

    table.setStyle(TableStyle(base_styles))
    elements.append(table)
    elements.append(Spacer(1, 0.15*inch))

    # Create center-aligned style
    center_heading_style = styles['Heading3'].clone('center_heading')
    center_heading_style.alignment = TA_CENTER
    center_normal_style = styles['Normal'].clone('center_normal')
    center_normal_style.alignment = TA_CENTER

    # Level distribution table - center aligned
    elements.append(Paragraph("<b>LEVEL DISTRIBUTION</b>", center_heading_style))
    grade_dist_headers = get_primary_level_order(resolved_level) if is_primary else (JUNIOR_LEVEL_ORDER if is_junior else ['A+', 'A', 'B+', 'B', 'C', 'D', 'E'])
    grade_dist_data = [grade_dist_headers]
    grade_dist_values = [str(grade_distribution.get(g, 0)) for g in grade_dist_headers]
    grade_dist_data.append(grade_dist_values)
    
    grade_table = Table(grade_dist_data)
    grade_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    grade_table.hAlign = 'CENTER'
    elements.append(grade_table)
    elements.append(Spacer(1, 0.1*inch))

    # Class mean - center aligned
    elements.append(Paragraph(f"<b>CLASS MEAN: {class_mean_grade}</b>", center_normal_style))

    doc.build(elements, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


@login_required
def export_merit_list_excel(request):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()
    if is_teacher and not request.user.teacher.is_class_teacher:
        return HttpResponseForbidden('Merit lists are available to class teachers only.')

    class_id = request.GET.get('class_id')
    exam_id = request.GET.get('exam_id')
    stream_id = request.GET.get('stream_id')
    term = (request.GET.get('term') or '').strip()

    if not class_id or not exam_id:
        return HttpResponse('Missing parameters', status=400)

    classroom = get_object_or_404(ClassRoom, id=class_id, school=school)
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
            teacher, school, cast(Any, classroom).id
        )
        if not allowed:
            return HttpResponseForbidden('Not allowed for this class.')
        if stream_id and not all_streams:
            try:
                if int(stream_id) not in allowed_stream_ids:
                    return HttpResponseForbidden('Not allowed for selected stream.')
            except (TypeError, ValueError):
                return HttpResponse('Invalid stream_id', status=400)
    exam = get_object_or_404(Exam, id=exam_id, school=school)

    ensure_cbe_learning_areas(school)
    class_level = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    school_any = cast(Any, school)
    is_junior = school_any.system_type == 'CBE' and resolved_level == 'Junior'
    is_primary = school_any.system_type == 'CBE' and resolved_level in ('Lower Primary', 'Upper Primary')
    show_junior_pathway = False

    # Fetch data using same logic
    term_filter = term or exam.term
    mark_sheets = MarkSheet.objects.filter(
        school_class=classroom, exam=exam, status='published',
    )
    if term_filter:
        mark_sheets = mark_sheets.filter(term=term_filter)
    
    mark_sheets = mark_sheets.select_related('subject')
    subjects_list = list(
        mark_sheets.values('subject_id', 'subject__short_name', 'subject__name', 'out_of').order_by('subject__short_name')
    )

    marks_qs = StudentMark.objects.filter(
        marksheet__in=mark_sheets, score__isnull=False,
    ).values('student_id').annotate(
        total=Sum('score'), subjects=Count('id'), average=Avg('score'),
    ).order_by('-average', '-total', 'student_id')
    if stream_id:
        marks_qs = marks_qs.filter(student__stream_id=stream_id)
    elif is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        _allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
            teacher, school, cast(Any, classroom).id
        )
        if not all_streams and allowed_stream_ids:
            marks_qs = marks_qs.filter(student__stream_id__in=allowed_stream_ids)

    student_ids = [row['student_id'] for row in marks_qs]
    students_map = {cast(Any, s).id: s for s in Student.objects.filter(id__in=student_ids, school=school)}

    grade_bands = []
    if not is_junior and not is_primary:
        grade_bands = list(
            GradeScale.objects.filter(school=school).order_by('-min_score')
            .values('min_score', 'max_score', 'grade')
        )

    def resolve_grade(score):
        if score is None:
            return '--'
        for band in grade_bands:
            if band['min_score'] <= score <= band['max_score']:
                return band['grade']
        return '--'

    marks = StudentMark.objects.filter(
        marksheet__in=mark_sheets, score__isnull=False,
    ).select_related('marksheet', 'marksheet__subject', 'student', 'student__stream')
    if stream_id:
        marks = marks.filter(student__stream_id=stream_id)
    elif is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        _allowed, all_streams, allowed_stream_ids = _teacher_class_teacher_scope_for_class(
            teacher, school, cast(Any, classroom).id
        )
        if not all_streams and allowed_stream_ids:
            marks = marks.filter(student__stream_id__in=allowed_stream_ids)

    subject_scores_by_student = defaultdict(dict)
    student_percentage_sum = defaultdict(float)
    student_subject_counts = defaultdict(int)
    student_points_total = defaultdict(int)
    subject_points_by_student = defaultdict(dict)

    for mark in marks:
        mark_any = cast(Any, mark)
        subject = mark_any.marksheet.subject
        out_of_val = mark_any.marksheet.out_of or 0
        score_val = float(mark_any.score)
        percentage = round((score_val / float(out_of_val) * 100), 1) if out_of_val else None

        entry = {
            'score': score_val,
            'percentage': percentage,
        }

        if is_primary:
            level = get_primary_level(percentage, resolved_level)
            points = get_primary_points(level)
            entry.update({'level': level, 'points': points})
            student_points_total[mark_any.student_id] += points
            if subject:
                subject_points_by_student[mark_any.student_id][subject.name] = points
        elif is_junior:
            level = get_junior_level(percentage)
            points = get_junior_points(level)
            entry.update({'level': level, 'points': points})
            student_points_total[mark_any.student_id] += points
            if subject:
                subject_points_by_student[mark_any.student_id][subject.name] = points
        else:
            entry['grade'] = resolve_grade(percentage) if percentage is not None else '--'

        subject_scores_by_student[mark_any.student_id][mark_any.marksheet.subject_id] = entry

        if percentage is not None:
            student_percentage_sum[mark_any.student_id] += percentage
            student_subject_counts[mark_any.student_id] += 1

    results = []
    prev_avg = prev_total = None
    position = 0
    for idx, row in enumerate(marks_qs, start=1):
        student = students_map.get(row['student_id'])
        if not student:
            continue
        avg_score = float(row['average'] or 0)
        total_score = float(row['total'] or 0)
        if prev_avg == avg_score and prev_total == total_score:
            pass
        else:
            position = idx
            prev_avg = avg_score
            prev_total = total_score

        subjects_map = subject_scores_by_student.get(row['student_id'], {})

        average_percentage = None
        if student_subject_counts[row['student_id']]:
            average_percentage = round(
                student_percentage_sum[row['student_id']] / student_subject_counts[row['student_id']],
                2
            )

        row_payload = {
            'position': position,
            'admission': student.admission_number,
            'student': f"{student.first_name} {student.last_name}".upper(),
            'stream': 'M' if student.gender == 'Male' else 'W',
            'subjects_map': subjects_map,
            'total': round(total_score, 2),
            'average': round(avg_score, 2),
        }

        if is_primary:
            avg_level = get_primary_level(average_percentage, resolved_level)
            row_payload.update({
                'average_level': avg_level or '--',
                'average_percentage': average_percentage,
                'recommended_interest': recommend_primary_interest(
                    subject_points_by_student.get(row['student_id'], {})
                ),
            })
        elif is_junior:
            total_points = student_points_total.get(row['student_id'], 0)
            avg_level = get_junior_level(average_percentage)
            row_payload.update({
                'average_level': avg_level or '--',
                'average_percentage': average_percentage,
                'total_points': total_points,
                'recommended_pathway': recommend_junior_pathway(
                    subject_points_by_student.get(row['student_id'], {})
                ),
            })
        else:
            row_payload['grade'] = resolve_grade(avg_score)

        results.append(row_payload)

    # Subject means, grade distribution, class mean
    subject_means_qs = StudentMark.objects.filter(
        marksheet__in=mark_sheets, score__isnull=False,
    ).values(
        'marksheet__subject_id', 'marksheet__subject__short_name', 'marksheet__out_of'
    ).annotate(mean_score=Avg('score')).order_by('marksheet__subject__short_name')

    subject_means = []
    for row in subject_means_qs:
        out_of_val = row['marksheet__out_of'] or 0
        mean_score = float(row['mean_score'] or 0)
        mean_percentage = round((mean_score / float(out_of_val) * 100), 1) if out_of_val else None
        mean_payload = {
            'short_name': row['marksheet__subject__short_name'],
            'mean_score': round(mean_score, 1),
        }
        if is_primary:
            level = get_primary_level(mean_percentage, resolved_level)
            mean_payload.update({
                'level': level or '--',
                'points': get_primary_points(level),
            })
        elif is_junior:
            level = get_junior_level(mean_percentage)
            mean_payload.update({
                'level': level or '--',
                'points': get_junior_points(level),
            })
        else:
            mean_payload['grade'] = resolve_grade(mean_percentage) if mean_percentage is not None else '--'
        subject_means.append(mean_payload)

    grade_distribution = {}
    for result in results:
        if is_primary:
            level = result.get('average_level')
            if level and level != '--':
                grade_distribution[level] = grade_distribution.get(level, 0) + 1
        elif is_junior:
            level = result.get('average_level')
            if level and level != '--':
                grade_distribution[level] = grade_distribution.get(level, 0) + 1
        else:
            grade = result.get('grade')
            if grade and grade != '--':
                grade_distribution[grade] = grade_distribution.get(grade, 0) + 1

    if results:
        if is_primary:
            avg_values = [r.get('average_percentage') for r in results if r.get('average_percentage') is not None]
            class_mean_score = (sum(avg_values) / len(avg_values)) if avg_values else None
            class_mean_grade = get_primary_level(class_mean_score, resolved_level) if class_mean_score is not None else '--'
        elif is_junior:
            avg_values = [r.get('average_percentage') for r in results if r.get('average_percentage') is not None]
            class_mean_score = (sum(avg_values) / len(avg_values)) if avg_values else None
            class_mean_grade = get_junior_level(class_mean_score) if class_mean_score is not None else '--'
        else:
            total_grades = sum([r['average'] for r in results])
            class_mean_score = total_grades / len(results)
            class_mean_grade = resolve_grade(class_mean_score)
    else:
        class_mean_grade = '--'

    # Build Excel
    wb = openpyxl.Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Merit List"

    # Footer (print footer)
    generated_time = timezone.localtime(timezone.now()).strftime('%d-%b-%Y %I:%M %p')
    footer = cast(Any, ws.oddFooter)
    footer.left.text = f"Generated on {generated_time}"
    footer.center.text = "&P"
    footer.right.text = "SKUL PLUS +(student information reporting management system)"

    # Styling
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    lvl_font = Font(color="1D4ED8")
    header_fill = PatternFill(start_color="777777", end_color="777777", fill_type="solid")
    lvl_fill = PatternFill(start_color="FFFBEA", end_color="FFFBEA", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Columns setup
    subject_count = len(subjects_list)
    subject_cols_per = 3 if is_junior else 2
    trailing_cols = (3 if show_junior_pathway else 2) if is_junior else (2 if is_primary else 2)
    total_cols = 4 + subject_count * subject_cols_per + trailing_cols
    last_col_letter = get_column_letter(total_cols)

    # Header rows
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = school_any.name
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = school_any.address
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A4:{last_col_letter}4')
    ws['A4'] = f"CLASS {classroom.name.upper()} - {term_filter.upper()} {exam.year} {exam.title.upper()}"
    ws['A4'].font = title_font
    ws['A4'].alignment = Alignment(horizontal='center')

    # Table headers
    header_row = 6
    headers = ['POS', 'ADM', 'NAME', 'STRM']
    if is_primary:
        for subj in subjects_list:
            headers.append(subj['subject__short_name'] or subj['subject__name'][:3].upper())
            headers.append('LVL')
        headers.extend(['AVG LVL', 'INTEREST'])
    elif is_junior:
        for subj in subjects_list:
            headers.append(subj['subject__short_name'] or subj['subject__name'][:3].upper())
            headers.append('LVL')
            headers.append('PTS')
        headers.extend(['TOTAL PTS', 'AVG LVL'])
        if show_junior_pathway:
            headers.append('REC PATH')
    else:
        for subj in subjects_list:
            headers.append(subj['subject__short_name'] or subj['subject__name'][:3].upper())
            headers.append('LVL')
        headers.extend(['TOTAL', 'LVL'])

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = table_header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = header_row + 1
    start_col = 4
    total_col = start_col + subject_count * subject_cols_per + 1
    total_lvl_col = total_col + 1
    if is_junior and show_junior_pathway:
        rec_path_col = total_lvl_col + 1
    for result in results:
        ws.cell(row=row, column=1, value=result['position']).border = border
        ws.cell(row=row, column=2, value=result['admission']).border = border
        ws.cell(row=row, column=3, value=result['student']).border = border
        ws.cell(row=row, column=4, value=result['stream']).border = border

        col = start_col + 1
        for subj in subjects_list:
            subj_data = result['subjects_map'].get(subj['subject_id'])
            mark_col = col
            lvl_col = col + 1
            if subj_data and subj_data.get('score') is not None:
                mark_cell = ws.cell(row=row, column=mark_col, value=int(round(subj_data['score'])))
                mark_cell.font = bold_font
                mark_cell.border = border
                lvl_cell = ws.cell(row=row, column=lvl_col, value=subj_data.get('level') if (is_junior or is_primary) else subj_data.get('grade'))
            else:
                mark_cell = ws.cell(row=row, column=mark_col, value='-')
                mark_cell.font = bold_font
                mark_cell.border = border
                lvl_cell = ws.cell(row=row, column=lvl_col, value='-')
            lvl_cell.fill = lvl_fill
            lvl_cell.font = lvl_font
            lvl_cell.border = border
            mark_cell.alignment = Alignment(horizontal='center')
            lvl_cell.alignment = Alignment(horizontal='center')
            if is_junior:
                pts_cell = ws.cell(row=row, column=col + 2, value=subj_data.get('points', 0) if subj_data else '-')
                pts_cell.border = border
                pts_cell.alignment = Alignment(horizontal='center')
                col += 3
            else:
                col += 2

        if is_primary:
            total_cell = ws.cell(row=row, column=total_col, value=result.get('average_level', '--'))
            total_cell.border = border
            total_cell.alignment = Alignment(horizontal='center')

            total_lvl_cell = ws.cell(row=row, column=total_lvl_col, value=result.get('recommended_interest', 'GENERAL'))
            total_lvl_cell.border = border
            total_lvl_cell.alignment = Alignment(horizontal='center')
        elif is_junior:
            total_cell = ws.cell(row=row, column=total_col, value=int(round(result.get('total_points', 0))))
            total_cell.border = border
            total_cell.alignment = Alignment(horizontal='center')

            total_lvl_cell = ws.cell(row=row, column=total_lvl_col, value=result.get('average_level', '--'))
            total_lvl_cell.fill = lvl_fill
            total_lvl_cell.font = lvl_font
            total_lvl_cell.border = border
            total_lvl_cell.alignment = Alignment(horizontal='center')
            if show_junior_pathway:
                rec_cell = ws.cell(row=row, column=rec_path_col, value=result.get('recommended_pathway', 'GENERAL'))
                rec_cell.border = border
                rec_cell.alignment = Alignment(horizontal='center')
        else:
            total_cell = ws.cell(row=row, column=total_col, value=int(round(result['total'])))
            total_cell.border = border
            total_cell.alignment = Alignment(horizontal='center')

            total_lvl_cell = ws.cell(row=row, column=total_lvl_col, value=result['grade'])
            total_lvl_cell.fill = lvl_fill
            total_lvl_cell.font = lvl_font
            total_lvl_cell.border = border
            total_lvl_cell.alignment = Alignment(horizontal='center')

        row += 1

    # Subject mean row
    ws.cell(row=row, column=3, value='SUB MSS').font = bold_font
    col = start_col + 1
    for subj_mean in subject_means:
        mark_cell = ws.cell(row=row, column=col, value=subj_mean['mean_score'])
        mark_cell.font = bold_font
        mark_cell.alignment = Alignment(horizontal='center')
        mark_cell.border = border

        lvl_cell = ws.cell(row=row, column=col + 1, value='-')
        lvl_cell.alignment = Alignment(horizontal='center')
        lvl_cell.border = border
        lvl_cell.fill = lvl_fill
        lvl_cell.font = lvl_font
        if is_junior:
            pts_cell = ws.cell(row=row, column=col + 2, value='-')
            pts_cell.alignment = Alignment(horizontal='center')
            pts_cell.border = border
            col += 3
        else:
            col += 2

    for col_idx in range(1, total_cols + 1):
        ws.cell(row=row, column=col_idx).border = border
    row += 1

    # Subject level row
    ws.cell(row=row, column=3, value='SUB LVL').font = bold_font
    col = start_col + 1
    for subj_mean in subject_means:
        mark_cell = ws.cell(row=row, column=col, value='-')
        mark_cell.font = bold_font
        mark_cell.alignment = Alignment(horizontal='center')
        mark_cell.border = border

        lvl_cell = ws.cell(row=row, column=col + 1, value=subj_mean['level'] if (is_junior or is_primary) else subj_mean['grade'])
        lvl_cell.fill = lvl_fill
        lvl_cell.font = lvl_font
        lvl_cell.alignment = Alignment(horizontal='center')
        lvl_cell.border = border
        if is_junior:
            pts_cell = ws.cell(row=row, column=col + 2, value='-')
            pts_cell.alignment = Alignment(horizontal='center')
            pts_cell.border = border
            col += 3
        else:
            col += 2

    for col_idx in range(1, total_cols + 1):
        ws.cell(row=row, column=col_idx).border = border
    row += 1

    if is_junior:
        ws.cell(row=row, column=3, value='SUB PTS').font = bold_font
        col = start_col + 1
        for subj_mean in subject_means:
            mark_cell = ws.cell(row=row, column=col, value='-')
            mark_cell.font = bold_font
            mark_cell.alignment = Alignment(horizontal='center')
            mark_cell.border = border

            lvl_cell = ws.cell(row=row, column=col + 1, value='-')
            lvl_cell.fill = lvl_fill
            lvl_cell.font = lvl_font
            lvl_cell.alignment = Alignment(horizontal='center')
            lvl_cell.border = border

            pts_cell = ws.cell(row=row, column=col + 2, value=subj_mean.get('points', 0))
            pts_cell.alignment = Alignment(horizontal='center')
            pts_cell.border = border
            col += 3

        for col_idx in range(1, total_cols + 1):
            ws.cell(row=row, column=col_idx).border = border
        row += 1

    row += 1

    # Level distribution
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    ws.cell(row=row, column=1, value='LEVEL DISTRIBUTION').font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1

    grade_headers = get_primary_level_order(resolved_level) if is_primary else (JUNIOR_LEVEL_ORDER if is_junior else ['A+', 'A', 'B+', 'B', 'C', 'D', 'E'])
    start_dist_col = max(1, (total_cols - len(grade_headers)) // 2 + 1)
    for idx, grade in enumerate(grade_headers):
        cell = ws.cell(row=row, column=start_dist_col + idx, value=grade)
        cell.font = table_header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    for idx, grade in enumerate(grade_headers):
        cell = ws.cell(row=row, column=start_dist_col + idx, value=grade_distribution.get(grade, 0))
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 2

    # Class mean
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    ws.cell(row=row, column=1, value=f"CLASS MEAN: {class_mean_grade}").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    for col_idx in range(1, total_cols + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="merit_list_{classroom.name}_{exam.title}.xlsx"'
    wb.save(response)
    return response


@login_required
def subject_analysis(request):
    return _analysis_page(request, default_scope='subject')


@login_required
def full_class_analysis(request):
    return _analysis_page(request, default_scope='class')


def _analysis_page(request, default_scope='class'):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    has_academics_role = user_has_permission(request.user, school, 'academics')
    if not (is_headteacher or is_superuser or is_teacher or has_academics_role):
        return HttpResponseForbidden()

    analysis_scope = (request.GET.get('scope') or default_scope or 'class').strip().lower()
    if analysis_scope not in {'class', 'subject'}:
        analysis_scope = 'class'

    classes_qs = ClassRoom.objects.filter(school=school)
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        if analysis_scope == 'subject':
            allowed_class_ids = _teacher_assigned_class_ids(teacher, school)
        else:
            allowed_class_ids = _teacher_class_teacher_class_ids(teacher, school)
        classes_qs = classes_qs.filter(id__in=allowed_class_ids)

    ensure_cbe_learning_areas(school)
    context = {
        'school': school,
        'classes': classes_qs.order_by('name'),
        'exams': Exam.objects.filter(school=school).order_by('-year', 'term', 'title'),
        'terms': ['Term 1', 'Term 2', 'Term 3'],
        'analysis_scope': analysis_scope,
    }
    return render(request, 'schools/subject_analysis.html', context)


def _grading_section_for_classroom(classroom: ClassRoom, resolved_level: str | None) -> str:
    if resolved_level == 'Lower Primary':
        return 'LOWER_PRIMARY'
    if resolved_level == 'Upper Primary':
        return 'UPPER_PRIMARY'
    if resolved_level == 'Junior':
        return 'JUNIOR'

    class_name = str(getattr(cast(Any, classroom), 'name', '') or '')
    match = re.search(r'(\d+)', class_name)
    if match:
        grade_num = int(match.group(1))
        if 1 <= grade_num <= 3:
            return 'LOWER_PRIMARY'
        if 4 <= grade_num <= 6:
            return 'UPPER_PRIMARY'
        if 7 <= grade_num <= 9:
            return 'JUNIOR'
    return 'GENERAL'


def _split_section_from_grade(raw_grade: str) -> tuple[str, str]:
    if raw_grade.startswith('L'):
        return ('LOWER_PRIMARY', raw_grade[1:])
    if raw_grade.startswith('U'):
        return ('UPPER_PRIMARY', raw_grade[1:])
    if raw_grade.startswith('J'):
        return ('JUNIOR', raw_grade[1:])
    return ('GENERAL', raw_grade)


def _build_grade_resolver_for_class(school: School, classroom: ClassRoom, resolved_level: str | None):
    is_cambridge = getattr(school, 'school_type', '') == 'CAMBRIDGE'
    selected_section = 'GENERAL' if is_cambridge else _grading_section_for_classroom(classroom, resolved_level)
    all_grade_bands = list(
        GradeScale.objects.filter(school=school).order_by('-min_score').values('min_score', 'max_score', 'grade', 'points')
    )

    section_grade_bands = []
    general_grade_bands = []
    for band in all_grade_bands:
        sec, label = _split_section_from_grade(str(band['grade'] or ''))
        entry = {
            'min_score': float(band['min_score'] or 0),
            'max_score': float(band['max_score'] or 0),
            'grade': str(label),
            'points': int(band.get('points') or 0),
        }
        if sec == selected_section:
            section_grade_bands.append(entry)
        elif sec == 'GENERAL':
            general_grade_bands.append(entry)

    active_bands = section_grade_bands if section_grade_bands else general_grade_bands
    grades_list = [b['grade'] for b in active_bands]

    def resolve_grade_points(score: float | None) -> tuple[str, int | None]:
        if score is None:
            return ('--', None)
        if not active_bands:
            return ('--', None)

        value = float(score)
        for band in active_bands:
            if band['min_score'] <= value <= band['max_score']:
                return (str(band['grade']), int(band['points']))

        sorted_by_min = sorted(active_bands, key=lambda b: float(b['min_score']), reverse=True)
        for band in sorted_by_min:
            if value >= float(band['min_score']):
                return (str(band['grade']), int(band['points']))

        lowest_band = sorted_by_min[-1]
        return (str(lowest_band['grade']), int(lowest_band['points']))

    return resolve_grade_points, grades_list


@login_required
def subject_analysis_data(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()
    school_any = cast(Any, school)
    class_id = request.GET.get('class_id')
    exam_id = request.GET.get('exam_id')
    term = (request.GET.get('term') or '').strip()
    raw_exam_weights = request.GET.get('exam_weights')
    analysis_scope = (request.GET.get('scope') or 'class').strip().lower()
    if analysis_scope not in {'class', 'subject'}:
        analysis_scope = 'class'

    if not class_id:
        return JsonResponse({'success': False, 'error': 'Please select a class.'}, status=400)

    if not exam_id and not term:
        return JsonResponse({'success': False, 'error': 'Please select an exam or a term.'}, status=400)

    classroom = get_object_or_404(ClassRoom, id=class_id, school=school)
    teacher_scope_subject_ids = None
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        class_pk = cast(Any, classroom).id
        if analysis_scope == 'class':
            allowed_class_ids = _teacher_class_teacher_class_ids(teacher, school)
            if class_pk not in allowed_class_ids:
                return JsonResponse({'success': False, 'error': 'Not allowed for this class in full class analysis.'}, status=403)
        else:
            allowed_class_ids = _teacher_assigned_class_ids(teacher, school)
            if class_pk not in allowed_class_ids:
                return JsonResponse({'success': False, 'error': 'Not allowed for this class in subject analysis.'}, status=403)
            allowed_subject_ids = _teacher_allowed_subject_ids_for_class(teacher, school, class_pk)
            if not allowed_subject_ids:
                return JsonResponse({'success': False, 'error': 'No subject allocation for this class.'}, status=403)
            teacher_scope_subject_ids = allowed_subject_ids
    class_level = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    is_junior = school_any.system_type == 'CBE' and resolved_level == 'Junior'
    is_primary = school_any.system_type == 'CBE' and resolved_level in ('Lower Primary', 'Upper Primary')
    resolve_grade_points, grades_list = _build_grade_resolver_for_class(school, classroom, resolved_level)

    def resolve_grade(score):
        grade, _pts = resolve_grade_points(score)
        return grade

    streams = list(
        Stream.objects.filter(classroom=classroom)
        .values('id', 'name')
        .order_by('name')
    )
    include_no_stream = Student.objects.filter(school=school, classroom=classroom, stream__isnull=True).exists()

    def init_distribution():
        base = {grade: 0 for grade in grades_list}
        dist = {str(s['id']): base.copy() for s in streams}
        if include_no_stream:
            dist['no_stream'] = base.copy()
        return dist

    def compute_exam_distribution(exam, term_filter):
        mark_sheets = MarkSheet.objects.filter(
            school_class=classroom,
            exam=exam,
            status='published',
        )
        if teacher_scope_subject_ids is not None:
            mark_sheets = mark_sheets.filter(subject_id__in=teacher_scope_subject_ids)
        if term_filter:
            mark_sheets = mark_sheets.filter(term=term_filter)

        if not mark_sheets.exists():
            return None

        marks_qs = StudentMark.objects.filter(
            marksheet__in=mark_sheets,
            score__isnull=False,
        ).values('student_id', 'score', 'marksheet__out_of')

        student_pct_map: dict[int, list[float]] = defaultdict(list)
        for row in marks_qs:
            sid = row.get('student_id')
            if sid is None:
                continue
            pct = _normalize_pct(row.get('score'), row.get('marksheet__out_of'))
            student_pct_map[int(sid)].append(pct)

        student_avg_scores = {
            sid: round(sum(vals) / len(vals), 2)
            for sid, vals in student_pct_map.items() if vals
        }
        student_ids = list(student_avg_scores.keys())
        students_map = {
            cast(Any, s).id: s for s in Student.objects.filter(
                id__in=student_ids,
                school=school,
                classroom=classroom
            ).select_related('stream')
        }

        distribution = init_distribution()
        stream_scores = {str(s['id']): [] for s in streams}
        if include_no_stream:
            stream_scores['no_stream'] = []
        
        for sid, score in student_avg_scores.items():
            student = students_map.get(sid)
            if not student:
                continue
            student_any = cast(Any, student)
            stream_key = str(student_any.stream_id) if student_any.stream_id else 'no_stream'
            grade = resolve_grade(score)
            if grade != '--' and stream_key in distribution:
                distribution[stream_key][grade] += 1
            if stream_key in stream_scores:
                stream_scores[stream_key].append(score)

        return {
            'exam_id': exam.id,
            'exam_title': exam.title,
            'term': term_filter or exam.term,
            'distribution': distribution,
            'stream_scores': stream_scores,
            'student_scores': student_avg_scores,
        }

    def compute_metrics(stream_scores):
        """Calculate metrics from stream scores"""
        stream_means = {}
        pass_counts = {}
        total_counts = {}
        all_scores = []
        
        # Define pass mark (assuming 50% is passing - adjust as needed)
        pass_mark = 50.0
        
        for stream_key, scores in stream_scores.items():
            if scores:
                stream_means[stream_key] = round(sum(scores) / len(scores), 1)
                pass_counts[stream_key] = sum(1 for s in scores if s >= pass_mark)
                total_counts[stream_key] = len(scores)
                all_scores.extend(scores)
        
        # Calculate pass rates
        pass_rates = {}
        for stream_key in stream_means.keys():
            if total_counts[stream_key] > 0:
                pass_rates[stream_key] = round((pass_counts[stream_key] / total_counts[stream_key]) * 100)
            else:
                pass_rates[stream_key] = 0
        
        # Calculate class mean
        class_mean = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0
        
        # Find best stream
        best_stream_id = None
        if stream_means:
            # Convert string keys back to int if they're numeric
            numeric_means = {k: v for k, v in stream_means.items() if k != 'no_stream'}
            if numeric_means:
                best_stream_key = max(numeric_means, key=lambda k: numeric_means[k])
                try:
                    best_stream_id = int(best_stream_key)
                except (ValueError, TypeError):
                    best_stream_id = None
        
        # Calculate failure rate
        total_students = len(all_scores)
        failed_students = sum(1 for s in all_scores if s < pass_mark)
        failure_rate = round((failed_students / total_students) * 100) if total_students > 0 else 0
        
        return {
            'stream_means': stream_means,
            'class_mean': class_mean,
            'pass_rate': pass_rates,
            'best_stream_id': best_stream_id,
            'failure_rate': failure_rate,
        }

    if exam_id:
        exam = get_object_or_404(Exam, id=exam_id, school=school)
        term_filter = term or exam.term
        payload = compute_exam_distribution(exam, term_filter)
        if not payload:
            return JsonResponse({'success': True, 'mode': 'single', 'data': None})
        
        metrics = compute_metrics(payload['stream_scores'])
        
        return JsonResponse({
            'success': True,
            'mode': 'single',
            'class_name': classroom.name,
            'streams': streams,
            'include_no_stream': include_no_stream,
            'grades': grades_list,
            'class_level': class_level,
            'is_junior': is_junior,
            'is_primary': is_primary,
            'metrics': metrics,
            'data': payload,
        })

    exams = Exam.objects.filter(school=school, term=term).order_by('-year', 'title')
    exam_ids_for_term = [cast(Any, ex).id for ex in exams]
    exam_weights = _parse_exam_weights_param(raw_exam_weights, exam_ids_for_term)
    results = []
    for exam in exams:
        payload = compute_exam_distribution(exam, term)
        if payload:
            payload['weight'] = float(exam_weights.get(cast(Any, exam).id, 1.0))
            results.append(payload)

    # Calculate weighted metrics from all exams combined (per student)
    student_weighted_scores: dict[int, dict[str, float]] = defaultdict(lambda: {'sum': 0.0, 'weight': 0.0})
    for result in results:
        weight = float(result.get('weight') or 1.0)
        for sid, score in (result.get('student_scores') or {}).items():
            sid_int = int(sid)
            student_weighted_scores[sid_int]['sum'] += float(score) * weight
            student_weighted_scores[sid_int]['weight'] += weight

    students_map = {
        cast(Any, s).id: s for s in Student.objects.filter(
            id__in=list(student_weighted_scores.keys()),
            school=school,
            classroom=classroom
        ).select_related('stream')
    }

    combined_scores = {str(s['id']): [] for s in streams}
    if include_no_stream:
        combined_scores['no_stream'] = []

    for sid, acc in student_weighted_scores.items():
        total_weight = float(acc.get('weight') or 0.0)
        if total_weight <= 0:
            continue
        avg_score = round(float(acc.get('sum') or 0.0) / total_weight, 2)
        student = students_map.get(sid)
        if not student:
            continue
        student_any = cast(Any, student)
        stream_key = str(student_any.stream_id) if student_any.stream_id else 'no_stream'
        if stream_key in combined_scores:
            combined_scores[stream_key].append(avg_score)
    
    metrics = compute_metrics(combined_scores) if results else {}

    return JsonResponse({
        'success': True,
        'mode': 'term',
        'class_name': classroom.name,
        'streams': streams,
        'include_no_stream': include_no_stream,
        'grades': grades_list,
        'class_level': class_level,
        'is_junior': is_junior,
        'is_primary': is_primary,
        'metrics': metrics,
        'data': results,
        'exam_weights': exam_weights,
    })


@login_required
def whole_school_analysis_data(request):
    """Return analysis data for ALL classes in the school"""
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()
    school_any = cast(Any, school)
    exam_id = request.GET.get('exam_id')
    term = (request.GET.get('term') or '').strip()

    if not exam_id and not term:
        return JsonResponse({'success': False, 'error': 'Please select an exam or a term.'}, status=400)

    # Get all classes for this school
    classrooms = ClassRoom.objects.filter(school=school).order_by('name')
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        class_teacher_ids = _teacher_class_teacher_class_ids(teacher, school)
        classrooms = classrooms.filter(id__in=class_teacher_ids)
    
    def process_class_data(classroom, exam, term_filter):
        """Process data for a single class and exam"""
        class_level = classroom.level.name if classroom.level else None
        resolved_level = resolve_cbe_level(school, class_level)
        is_junior = school_any.system_type == 'CBE' and resolved_level == 'Junior'
        is_primary = school_any.system_type == 'CBE' and resolved_level in ('Lower Primary', 'Upper Primary')
        resolve_grade_points, grades_list = _build_grade_resolver_for_class(school, classroom, resolved_level)

        streams = list(
            Stream.objects.filter(classroom=classroom)
            .values('id', 'name')
            .order_by('name')
        )
        
        mark_sheets = MarkSheet.objects.filter(
            school_class=classroom,
            exam=exam,
            status='published',
        )
        if term_filter:
            mark_sheets = mark_sheets.filter(term=term_filter)

        if not mark_sheets.exists():
            return None

        marks_qs = StudentMark.objects.filter(
            marksheet__in=mark_sheets,
            score__isnull=False,
        ).values('student_id').annotate(
            avg_score=Avg('score')
        )

        student_ids = [row['student_id'] for row in marks_qs]
        students_map = {
            cast(Any, s).id: s for s in Student.objects.filter(
                id__in=student_ids,
                school=school,
                classroom=classroom
            ).select_related('stream')
        }

        # Initialize distribution
        distribution = {}
        stream_scores = {}
        
        for s in streams:
            distribution[str(s['id'])] = {grade: 0 for grade in grades_list}
            stream_scores[str(s['id'])] = []
        
        for row in marks_qs:
            student = students_map.get(row['student_id'])
            if not student or not cast(Any, student).stream_id:
                continue
            stream_key = str(cast(Any, student).stream_id)
            score = float(row['avg_score'] or 0)
            grade, _pts = resolve_grade_points(score)
            
            if grade != '--' and stream_key in distribution:
                distribution[stream_key][grade] += 1
            if stream_key in stream_scores:
                stream_scores[stream_key].append(score)

        # Calculate metrics
        stream_means = {}
        pass_rates = {}
        all_scores = []
        pass_mark = 50.0
        
        for stream_key, scores in stream_scores.items():
            if scores:
                stream_means[stream_key] = round(sum(scores) / len(scores), 1)
                pass_count = sum(1 for s in scores if s >= pass_mark)
                pass_rates[stream_key] = round((pass_count / len(scores)) * 100)
                all_scores.extend(scores)
        
        class_mean = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0
        
        best_stream_id = None
        if stream_means:
            numeric_means = {k: v for k, v in stream_means.items()}
            if numeric_means:
                best_stream_key = max(numeric_means, key=lambda k: numeric_means[k])
                try:
                    best_stream_id = int(best_stream_key)
                except (ValueError, TypeError):
                    pass
        
        failed_students = sum(1 for s in all_scores if s < pass_mark)
        failure_rate = round((failed_students / len(all_scores)) * 100) if all_scores else 0
        
        return {
            'exam_id': exam.id,
            'exam_title': exam.title,
            'term': term_filter or exam.term,
            'distribution': distribution,
            'grades': grades_list,
            'class_level': class_level,
            'is_junior': is_junior,
            'is_primary': is_primary,
            'metrics': {
                'stream_means': stream_means,
                'class_mean': class_mean,
                'pass_rate': pass_rates,
                'best_stream_id': best_stream_id,
                'failure_rate': failure_rate,
            }
        }

    classes_data = []
    
    if exam_id:
        # Single exam mode
        exam = get_object_or_404(Exam, id=exam_id, school=school)
        term_filter = term or exam.term
        
        for classroom in classrooms:
            streams = list(
                Stream.objects.filter(classroom=classroom)
                .values('id', 'name')
                .order_by('name')
            )
            
            exam_data = process_class_data(classroom, exam, term_filter)
            if exam_data:
                classes_data.append({
                    'class_name': classroom.name,
                    'streams': streams,
                    'grades': exam_data['grades'],
                    'class_level': exam_data['class_level'],
                    'is_junior': exam_data['is_junior'],
                    'metrics': exam_data['metrics'],
                    'exams': [exam_data]
                })
    else:
        # Term mode - multiple exams
        exams = Exam.objects.filter(school=school, term=term).order_by('-year', 'title')
        
        for classroom in classrooms:
            streams = list(
                Stream.objects.filter(classroom=classroom)
                .values('id', 'name')
                .order_by('name')
            )
            
            class_exams = []
            for exam in exams:
                exam_data = process_class_data(classroom, exam, term)
                if exam_data:
                    class_exams.append(exam_data)
            
            if class_exams:
                # Calculate combined metrics for this class across all exams
                combined_scores = {str(s['id']): [] for s in streams}
                for exam_data in class_exams:
                    # Re-extract scores from the exam data (we need to recalculate)
                    pass  # Using last exam's metrics for now
                
                classes_data.append({
                    'class_name': classroom.name,
                    'streams': streams,
                    'grades': class_exams[-1]['grades'] if class_exams else [],
                    'class_level': class_exams[-1]['class_level'] if class_exams else None,
                    'is_junior': class_exams[-1]['is_junior'] if class_exams else False,
                    'metrics': class_exams[-1]['metrics'] if class_exams else {},
                    'exams': class_exams
                })

    return JsonResponse({
        'success': True,
        'classes': classes_data,
    })


@login_required
def whole_school_subject_stream_analysis(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if not (is_headteacher or is_superuser or is_teacher):
        return HttpResponseForbidden()
    school_any = cast(Any, school)
    class_id = request.GET.get("class_id")
    exam_id = request.GET.get("exam_id")
    term = request.GET.get("term")
    raw_exam_weights = request.GET.get("exam_weights")
    compare_exam_id = request.GET.get("compare_exam_id")
    split_by_stream = request.GET.get("split_by_stream") == "true"
    show_ranking_param = request.GET.get("show_ranking")
    analysis_scope = (request.GET.get('scope') or 'class').strip().lower()
    if analysis_scope not in {'class', 'subject'}:
        analysis_scope = 'class'
    is_cambridge = getattr(school_any, 'school_type', '') == 'CAMBRIDGE'
    if show_ranking_param is None:
        show_ranking = (not is_cambridge) or bool(getattr(school_any, 'cambridge_show_ranking', False))
    else:
        show_ranking = str(show_ranking_param).strip().lower() in {'1', 'true', 'yes', 'on'}

    if not class_id:
        return JsonResponse({"success": False, "error": "Please select a class."}, status=400)

    if not exam_id and not term:
        return JsonResponse({"success": False, "error": "Please select an exam or a term."}, status=400)

    classroom = get_object_or_404(ClassRoom, id=class_id, school=school)
    teacher_scope_subject_ids = None
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        class_pk = cast(Any, classroom).id
        if analysis_scope == 'class':
            allowed_class_ids = _teacher_class_teacher_class_ids(teacher, school)
            if class_pk not in allowed_class_ids:
                return JsonResponse({'success': False, 'error': 'Not allowed for this class in full class analysis.'}, status=403)
        else:
            allowed_class_ids = _teacher_assigned_class_ids(teacher, school)
            if class_pk not in allowed_class_ids:
                return JsonResponse({'success': False, 'error': 'Not allowed for this class in subject analysis.'}, status=403)
            allowed_subject_ids = _teacher_allowed_subject_ids_for_class(teacher, school, class_pk)
            if not allowed_subject_ids:
                return JsonResponse({'success': False, 'error': 'No subject allocation for this class.'}, status=403)
            teacher_scope_subject_ids = allowed_subject_ids
    ensure_cbe_learning_areas(school)
    class_level = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    is_junior = school_any.system_type == 'CBE' and resolved_level == 'Junior'
    is_primary = school_any.system_type == 'CBE' and resolved_level in ('Lower Primary', 'Upper Primary')
    is_senior = school_any.system_type == 'CBE' and resolved_level == 'Senior'

    filters = Q(marksheet__school_class=classroom)
    if exam_id:
        filters &= Q(marksheet__exam_id=exam_id)
    if term:
        filters &= Q(marksheet__exam__term=term)

    results = StudentMark.objects.filter(filters, score__isnull=False).select_related(
        "student", "student__classroom", "student__stream", "marksheet", "marksheet__subject", "marksheet__exam"
    )
    if teacher_scope_subject_ids is not None:
        results = results.filter(marksheet__subject_id__in=teacher_scope_subject_ids)

    exam_ids_for_weights: list[int] = []
    if exam_id:
        try:
            exam_ids_for_weights = [int(exam_id)]
        except Exception:
            exam_ids_for_weights = []
    elif term:
        exam_ids_for_weights = list(
            Exam.objects.filter(school=school, term=term).values_list('id', flat=True)
        )
    exam_weights = _parse_exam_weights_param(raw_exam_weights, exam_ids_for_weights)

    resolve_grade_points, grades_list = _build_grade_resolver_for_class(school, classroom, resolved_level)

    data_by_stream = defaultdict(list)
    subject_grade_distribution = defaultdict(lambda: defaultdict(lambda: {g: 0 for g in grades_list}))
    subject_grade_totals = defaultdict(lambda: {g: 0 for g in grades_list})

    for r in results:
        stream_name = r.student.stream.name if (split_by_stream and r.student.stream) else "All Streams"
        subject_name = r.marksheet.subject.name if r.marksheet.subject else "N/A"

        marksheet_any = cast(Any, r.marksheet)
        exam_pk = int(getattr(marksheet_any, 'exam_id', 0) or 0)
        exam_weight = float(exam_weights.get(exam_pk, 1.0))
        pct_score = _normalize_pct(cast(Any, r).score, getattr(marksheet_any, 'out_of', 100))
        data_by_stream[stream_name].append({
            "subject": subject_name,
            "mark": pct_score,
            "weight": exam_weight,
        })

        grade, _pts = resolve_grade_points(pct_score)
        if grade in grades_list:
            subject_grade_distribution[subject_name][stream_name][grade] += 1
            subject_grade_totals[subject_name][grade] += 1

    subjects = sorted({item["subject"] for s in data_by_stream.values() for item in s})
    subject_shorts = {}
    subject_pathways = {}
    for subject in subjects:
        subject_obj = Subject.objects.filter(school=school, name=subject).select_related('pathway', 'education_level').first()
        subject_shorts[subject] = subject_obj.short_name if subject_obj and subject_obj.short_name else subject[:3].upper()
        subject_pathways[subject] = {
            "code": subject_obj.pathway.code if subject_obj and subject_obj.pathway else None,
            "name": subject_obj.pathway.name if subject_obj and subject_obj.pathway else None,
            "level": subject_obj.education_level.name if subject_obj and subject_obj.education_level else None,
        }

    class_subject_totals = defaultdict(list)
    stream_outputs = []

    for stream_name, records in data_by_stream.items():
        subject_marks: dict[str, list[tuple[float, float]]] = defaultdict(list)

        for rec in records:
            subject_marks[rec["subject"]].append((float(rec["mark"] or 0), float(rec.get("weight") or 1.0)))

        stream_means = {}
        stream_levels = {}
        performance = {}

        all_stream_marks: list[tuple[float, float]] = []

        for subject in subjects:
            marks = subject_marks.get(subject, [])
            if not marks:
                continue

            weighted = _weighted_mean(marks)
            mean_mark = round(float(weighted if weighted is not None else 0.0), 2)
            stream_means[subject] = mean_mark
            stream_levels[subject] = resolve_grade_points(mean_mark)[0]

            class_subject_totals[subject].extend(marks)
            all_stream_marks.extend(marks)

            raw_scores = [m[0] for m in marks]
            performance[subject] = {
                "mean": mean_mark,
                "grade": resolve_grade_points(mean_mark)[0],
                "highest": max(raw_scores) if raw_scores else 0,
                "lowest": min(raw_scores) if raw_scores else 0,
                "entries": len(raw_scores),
                "pass_rate": round(
                    (len([m for m in raw_scores if m >= 50]) / len(raw_scores)) * 100, 1
                )
            }

        class_mean_calc = _weighted_mean(all_stream_marks)
        class_mean = round(float(class_mean_calc), 2) if class_mean_calc is not None else ""

        stream_outputs.append({
            "stream_name": stream_name,
            "means": stream_means,
            "levels": stream_levels,
            "class_mean": class_mean,
            "performance": performance
        })

    class_subject_means = {}
    class_subject_levels = {}

    overall_marks = []

    for subject, marks in class_subject_totals.items():
        weighted = _weighted_mean(marks)
        mean_mark = round(float(weighted if weighted is not None else 0.0), 2)
        class_subject_means[subject] = mean_mark
        class_subject_levels[subject] = resolve_grade_points(mean_mark)[0]
        overall_marks.extend(marks)

    overall_weighted = _weighted_mean(overall_marks)
    overall_class_mean = round(float(overall_weighted), 2) if overall_weighted is not None else ""

    subject_grade_distribution_dict = {
        subject: {stream: dict(counts) for stream, counts in streams.items()}
        for subject, streams in subject_grade_distribution.items()
    }
    subject_grade_totals_dict = {
        subject: dict(counts) for subject, counts in subject_grade_totals.items()
    }

    previous_subject_means = {}
    previous_subject_levels = {}
    previous_stream_mean_points = {}
    previous_class_mean_points = None
    if exam_id:
        current_exam = Exam.objects.filter(id=exam_id, school=school).first()
        previous_exam = None

        if compare_exam_id:
            previous_exam = Exam.objects.filter(id=compare_exam_id, school=school).first()
        elif current_exam:
            previous_exam = Exam.objects.filter(
                school=school,
                term=current_exam.term,
                year=current_exam.year,
                start_date__lt=current_exam.start_date,
            ).order_by('-start_date').first()

        if previous_exam:
            prev_results = StudentMark.objects.filter(
                marksheet__school_class=classroom,
                marksheet__exam=previous_exam,
                score__isnull=False,
            ).select_related("marksheet", "marksheet__subject")

            prev_subject_marks = defaultdict(list)
            for r in prev_results:
                subject_name = r.marksheet.subject.name if r.marksheet.subject else "N/A"
                prev_pct = _normalize_pct(cast(Any, r).score, getattr(cast(Any, r).marksheet, 'out_of', 100))
                prev_subject_marks[subject_name].append(prev_pct)

            for subject in subjects:
                marks = prev_subject_marks.get(subject, [])
                if not marks:
                    continue
                mean_mark = round(sum(marks) / len(marks), 2)
                previous_subject_means[subject] = mean_mark
                previous_subject_levels[subject] = resolve_grade_points(mean_mark)[0]

            prev_student_scores = defaultdict(lambda: {"total": 0, "count": 0, "stream": None})
            for r in prev_results:
                s = r.student
                s_any = cast(Any, s)
                entry = cast(dict[str, Any], prev_student_scores[s_any.id])
                score_val = _normalize_pct(cast(Any, r).score, getattr(cast(Any, r).marksheet, 'out_of', 100))
                entry["total"] = float(entry["total"]) + score_val
                entry["count"] = float(entry["count"]) + 1.0
                entry["stream"] = s_any.stream.name if s_any.stream else "No Stream"

            prev_stream_points = defaultdict(list)
            prev_all_points = []
            for entry in prev_student_scores.values():
                if entry["count"] == 0:
                    continue
                count_val = float(entry["count"] or 0)
                if count_val <= 0:
                    continue
                avg_score = float(entry["total"] or 0) / count_val
                grade, points = resolve_grade_points(avg_score)
                if points is None:
                    pass
                else:
                    prev_stream_points[entry["stream"]].append(points)
                    prev_all_points.append(points)

            for stream_name, points_list in prev_stream_points.items():
                if points_list:
                    previous_stream_mean_points[stream_name] = round(sum(points_list) / len(points_list), 2)

            if prev_all_points:
                previous_class_mean_points = round(sum(prev_all_points) / len(prev_all_points), 2)

    student_scores = defaultdict(lambda: {"total": 0, "count": 0, "stream": None})
    for r in results:
        s = r.student
        s_any = cast(Any, s)
        entry = cast(dict[str, Any], student_scores[s_any.id])
        marksheet_any = cast(Any, r).marksheet
        exam_pk = int(getattr(marksheet_any, 'exam_id', 0) or 0)
        exam_weight = float(exam_weights.get(exam_pk, 1.0))
        score_val = _normalize_pct(cast(Any, r).score, getattr(marksheet_any, 'out_of', 100))
        entry["total"] = float(entry["total"]) + (score_val * exam_weight)
        entry["count"] = float(entry["count"]) + exam_weight
        entry["stream"] = s_any.stream.name if s_any.stream else "No Stream"

    stream_grade_distribution = defaultdict(lambda: {g: 0 for g in grades_list})
    stream_mean_marks = defaultdict(list)
    stream_points = defaultdict(list)
    overall_avg_marks = []
    overall_points = []

    for entry in student_scores.values():
        if float(entry["count"] or 0) <= 0:
            continue
        count_val = float(entry["count"] or 0)
        avg_score = float(entry["total"] or 0) / count_val
        overall_avg_marks.append(avg_score)
        stream_name = entry["stream"]
        stream_mean_marks[stream_name].append(avg_score)

        grade, points = resolve_grade_points(avg_score)
        if grade in grades_list:
            stream_grade_distribution[stream_name][grade] += 1
        if points is not None:
            stream_points[stream_name].append(points)
            overall_points.append(points)

    class_mean_mark = round(sum(overall_avg_marks) / len(overall_avg_marks), 2) if overall_avg_marks else None
    class_mean_points = round(sum(overall_points) / len(overall_points), 2) if overall_points else None

    overall_grade_summary = []
    for stream_name, marks_list in stream_mean_marks.items():
        entries = len(marks_list)
        mean_marks = round(sum(marks_list) / entries, 2) if entries else None
        mm_dev = round(mean_marks - class_mean_mark, 2) if (mean_marks is not None and class_mean_mark is not None) else None

        points_list = stream_points.get(stream_name, [])
        mean_points = round(sum(points_list) / len(points_list), 2) if points_list else None
        mp_dev = round(mean_points - class_mean_points, 2) if (mean_points is not None and class_mean_points is not None) else None

        prev_mp = previous_stream_mean_points.get(stream_name)
        dev = round(mean_points - prev_mp, 2) if (mean_points is not None and prev_mp is not None) else None

        overall_grade_summary.append({
            "stream": stream_name,
            "grade_counts": stream_grade_distribution.get(stream_name, {}),
            "entries": entries,
            "mean_marks": mean_marks,
            "mm_dev": mm_dev,
            "mean_points": None if is_primary else mean_points,
            "mp_dev": None if is_primary else mp_dev,
            "previous_mean_points": None if is_primary else prev_mp,
            "dev": None if is_primary else dev,
            "mean_grade": resolve_grade_points(mean_marks)[0] if mean_marks is not None else "",
        })

    if overall_avg_marks:
        overall_grade_summary.append({
            "stream": "Overall",
            "grade_counts": {g: sum(stream_grade_distribution[s].get(g, 0) for s in stream_grade_distribution) for g in grades_list},
            "entries": len(overall_avg_marks),
            "mean_marks": class_mean_mark,
            "mm_dev": 0,
            "mean_points": None if is_primary else class_mean_points,
            "mp_dev": None if is_primary else 0,
            "previous_mean_points": None if is_primary else previous_class_mean_points,
            "dev": None if is_primary else (round(class_mean_points - previous_class_mean_points, 2) if (class_mean_points is not None and previous_class_mean_points is not None) else None),
            "mean_grade": resolve_grade_points(class_mean_mark)[0] if class_mean_mark is not None else "",
        })

    pathway_analysis = {
        "enabled": False,
        "level": None,
        "reason": "Not a CBE school",
    }

    level_name = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, level_name)
    if school_any.system_type == 'CBE' and resolved_level == 'Senior' and school_allows_pathways(school):
        pathway_analysis = {
            "enabled": True,
            "level": resolved_level,
        }

        if resolved_level == 'Senior':
            pathway_student_points = defaultdict(lambda: defaultdict(float))
            pathway_subject_marks = defaultdict(lambda: defaultdict(list))
            student_total_points = defaultdict(float)
            student_total_marks = defaultdict(float)
            student_subject_counts = defaultdict(int)

            for r in results:
                subject = r.marksheet.subject
                pathway = subject.pathway if subject else None
                if not pathway:
                    continue
                score_pct = _normalize_pct(cast(Any, r).score, getattr(cast(Any, r).marksheet, 'out_of', 100))
                grade, points = resolve_grade_points(score_pct)
                if points is None:
                    points = 0
                r_any = cast(Any, r)
                pathway_student_points[pathway.name][r_any.student_id] += points
                student_total_points[r_any.student_id] += points
                student_total_marks[r_any.student_id] += score_pct
                student_subject_counts[r_any.student_id] += 1
                pathway_subject_marks[pathway.name][subject.name].append(score_pct)

            pathway_stats = []
            best_pathway = None
            best_points = None
            for pathway_name, student_points in pathway_student_points.items():
                points_list = list(student_points.values())
                mean_points = round(sum(points_list) / len(points_list), 2) if points_list else 0
                entries = len(points_list)
                pathway_stats.append({
                    "pathway": pathway_name,
                    "mean_points": mean_points,
                    "entries": entries,
                })
                if best_points is None or mean_points > best_points:
                    best_points = mean_points
                    best_pathway = pathway_name
            subject_contributions = {}
            for pathway_name, subjects_map in pathway_subject_marks.items():
                subject_contributions[pathway_name] = []
                for subject_name, marks_list in subjects_map.items():
                    mean_mark = round(sum(marks_list) / len(marks_list), 2) if marks_list else 0
                    mean_grade, mean_points = resolve_grade_points(mean_mark)
                    subject_contributions[pathway_name].append({
                        "subject": subject_name,
                        "mean_mark": mean_mark,
                        "mean_grade": mean_grade,
                        "mean_points": 0 if mean_points is None else mean_points,
                    })

            student_points = []
            for student_id, total_points in student_total_points.items():
                student = Student.objects.filter(id=student_id).first()
                if not student:
                    continue
                mean_marks = None
                if student_subject_counts[student_id]:
                    mean_marks = round(student_total_marks[student_id] / student_subject_counts[student_id], 2)
                student_points.append({
                    "admission_number": student.admission_number,
                    "name": f"{student.first_name} {student.last_name}",
                    "stream": student.stream.name if student.stream else "",
                    "total_points": round(total_points, 2),
                    "mean_marks": mean_marks,
                })

            pathway_analysis.update({
                "pathway_stats": pathway_stats,
                "best_pathway": best_pathway,
                "student_points": student_points,
                "subject_contributions": subject_contributions,
            })
    elif school_any.system_type == 'CBE':
        pathway_analysis["reason"] = "Pathways are disabled for this school."

        if level_name == 'Junior':
            learning_area_performance = []
            overall_band_counts = {g: 0 for g in grades_list}
            for subject_name in subjects:
                band_counts = subject_grade_totals_dict.get(subject_name, {g: 0 for g in grades_list})
                for grade, count in band_counts.items():
                    overall_band_counts[grade] += count
                learning_area_performance.append({
                    "learning_area": subject_name,
                    "mean_mark": class_subject_means.get(subject_name),
                    "mean_grade": class_subject_levels.get(subject_name),
                    "bands": band_counts,
                })

            pathway_analysis.update({
                "learning_area_performance": learning_area_performance,
                "competency_levels": overall_band_counts,
            })

    top_students_overall = []
    student_rows = []
    for student_id, entry in student_scores.items():
        if float(entry["count"] or 0) <= 0:
            continue
        count_val = float(entry["count"] or 0)
        avg_score = float(entry["total"] or 0) / count_val
        student = Student.objects.filter(id=student_id).first()
        if not student:
            continue
        student_rows.append({
            "admission_number": student.admission_number,
            "name": f"{student.first_name} {student.last_name}",
            "stream": entry["stream"],
            "avg_score": round(avg_score, 2),
            "total_marks": round(float(entry["total"] or 0), 2),
            "grade": resolve_grade_points(avg_score)[0],
        })

    student_rows.sort(key=lambda x: x["avg_score"], reverse=True)
    overall_ranks = {row["admission_number"]: idx + 1 for idx, row in enumerate(student_rows)}

    stream_groups = defaultdict(list)
    for row in student_rows:
        stream_groups[row["stream"]].append(row)

    stream_ranks = {}
    for stream_name, rows in stream_groups.items():
        rows_sorted = sorted(rows, key=lambda x: x["avg_score"], reverse=True)
        for idx, row in enumerate(rows_sorted):
            stream_ranks[(stream_name, row["admission_number"])] = idx + 1

    for row in student_rows[:5]:
        top_students_overall.append({
            "admission_number": row["admission_number"],
            "name": row["name"],
            "stream": row["stream"],
            "stream_rank": stream_ranks.get((row["stream"], row["admission_number"])) if show_ranking else None,
            "overall_rank": overall_ranks.get(row["admission_number"]) if show_ranking else None,
            "mean_marks": row["avg_score"],
            "total_marks": row["total_marks"],
            "mean_points": None if is_primary else resolve_grade_points(row["avg_score"])[1],
            "grade": row["grade"],
        })

    classes_output = [{
        "class_name": classroom.name,
        "subjects": subjects,
        "subject_shorts": subject_shorts,
        "subject_pathways": subject_pathways,
        "class_level": class_level,
        "system_type": school_any.system_type,
        "is_cbe_senior": school_any.system_type == 'CBE' and resolved_level == 'Senior' and school_allows_pathways(school),
        "is_cbe_junior": school_any.system_type == 'CBE' and resolved_level == 'Junior',
        "is_cbe_primary": school_any.system_type == 'CBE' and resolved_level in ('Lower Primary', 'Upper Primary'),
        "streams": stream_outputs,
        "class_subject_means": class_subject_means,
        "class_subject_levels": class_subject_levels,
        "overall_class_mean": overall_class_mean,
        "grades": grades_list,
        "subject_grade_distribution": subject_grade_distribution_dict,
        "subject_grade_totals": subject_grade_totals_dict,
        "previous_subject_means": previous_subject_means,
        "previous_subject_levels": previous_subject_levels,
        "overall_grade_summary": overall_grade_summary,
        "top_students_overall": top_students_overall,
        "pathway_analysis": pathway_analysis,
        "exam_title": cast(Any, results.first()).marksheet.exam.title if results else ""
    }]

    return JsonResponse({
        "mode": "normal",
        "classes": classes_output,
        "exam_weights": exam_weights,
        "show_ranking": show_ranking,
    })


@login_required
def report_cards(request):
    school, denied = _require_school_permission(request, 'academics')
    if denied:
        return denied
    context = {
        'school': school,
        'classes': ClassRoom.objects.filter(school=school).order_by('order', 'name'),
        'exams': Exam.objects.filter(school=school).order_by('-year', 'term', 'start_date', 'title'),
        'terms': ['Term 1', 'Term 2', 'Term 3'],
    }
    return render(request, 'schools/report_cards.html', context)


@login_required
@require_GET
def report_cards_data(request):
    school, denied = _require_school_permission(request, 'academics')
    if denied:
        return denied
    class_id = request.GET.get('class_id')
    exam_id = request.GET.get('exam_id')
    term = (request.GET.get('term') or '').strip()
    raw_exam_weights = request.GET.get('exam_weights')
    show_ranking_param = request.GET.get('show_ranking')
    if not class_id or not exam_id:
        return JsonResponse({'success': False, 'error': 'Please select class and exam.'}, status=400)

    classroom = get_object_or_404(ClassRoom, id=class_id, school=school)
    exam = get_object_or_404(Exam, id=exam_id, school=school)
    term_filter = term or exam.term

    class_level = classroom.level.name if classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    resolve_grade_points, _grades_list = _build_grade_resolver_for_class(school, classroom, resolved_level)
    selected_section = _grading_section_for_classroom(classroom, resolved_level)

    def _level_key(label: str) -> str:
        return re.sub(r'[^A-Z0-9]+', '', (label or '').upper())

    all_scale_bands = GradeScale.objects.filter(school=school).values('min_score', 'max_score', 'grade')
    section_bands = []
    general_bands = []
    for band in all_scale_bands:
        sec, raw_label = _split_section_from_grade(str(band.get('grade') or ''))
        rec = {'min_score': float(band.get('min_score') or 0), 'label': str(raw_label)}
        if sec == selected_section:
            section_bands.append(rec)
        elif sec == 'GENERAL':
            general_bands.append(rec)
    active_threshold_bands = section_bands if section_bands else general_bands
    me2_threshold = next((b['min_score'] for b in active_threshold_bands if _level_key(str(b['label'])) == 'ME2'), None)
    me_threshold = next((b['min_score'] for b in active_threshold_bands if _level_key(str(b['label'])) == 'ME'), None)
    eligibility_threshold = me2_threshold if me2_threshold is not None else me_threshold

    students = list(
        Student.objects.filter(school=school, classroom=classroom)
        .select_related('stream')
        .order_by('admission_number', 'last_name', 'first_name')
    )
    if not students:
        return JsonResponse({'success': False, 'error': 'No students found in selected class.'}, status=404)

    exam_order_rank = {'Term 1': 1, 'Term 2': 2, 'Term 3': 3}
    marksheets = list(
        MarkSheet.objects.filter(
            school_class=classroom,
            exam__school=school,
            exam__term=term_filter,
            subject__isnull=False,
        )
        .select_related('exam', 'subject')
        .order_by('exam__start_date', 'exam__id')
    )
    if not marksheets:
        return JsonResponse({'success': False, 'error': 'No marksheets found for selected class/term.'}, status=404)

    exams_by_id: dict[int, Any] = {}
    for ms in marksheets:
        ex = cast(Any, ms).exam
        exams_by_id[cast(Any, ex).id] = ex
    exams_in_term = sorted(
        exams_by_id.values(),
        key=lambda e: (cast(Any, e).start_date, cast(Any, e).year, exam_order_rank.get(cast(Any, e).term, 9), cast(Any, e).title),
    )
    exam_ids = [cast(Any, e).id for e in exams_in_term]
    exam_index = {eid: idx for idx, eid in enumerate(exam_ids)}
    exam_weights = _parse_exam_weights_param(raw_exam_weights, exam_ids)
    is_cambridge = getattr(school, 'school_type', '') == 'CAMBRIDGE'
    default_cambridge_ranking = bool(getattr(school, 'cambridge_show_ranking', False))
    if show_ranking_param is None:
        show_ranking = (not is_cambridge) or default_cambridge_ranking
    else:
        show_ranking = str(show_ranking_param).strip().lower() in {'1', 'true', 'yes', 'on'}

    pathway_labels = {
        'STEM': 'STEM',
        'SOCIAL_SCIENCES': 'Social Sciences',
        'ARTS_SPORTS_SCIENCE': 'Arts & Sports Science',
    }

    def normalize_category(raw_value: str) -> str:
        value = (raw_value or '').strip()
        if not value:
            return ''
        upper = value.upper()
        if upper in pathway_labels:
            return upper
        folded = re.sub(r'[^A-Z]+', '', upper)
        if 'STEM' in folded:
            return 'STEM'
        if 'SOCIAL' in folded:
            return 'SOCIAL_SCIENCES'
        if 'ART' in folded or 'SPORT' in folded:
            return 'ARTS_SPORTS_SCIENCE'
        return ''

    def fold_key(raw_value: str) -> str:
        return re.sub(r'[^a-z0-9]+', '', (raw_value or '').strip().lower())

    def infer_subject_category(subject_name: str, subject_code: str) -> str:
        text = f"{subject_name or ''} {subject_code or ''}".strip().lower()
        if not text:
            return ''

        stem_tokens = (
            'math', 'mathematics', 'physics', 'chemistry', 'biology', 'science',
            'computer', 'ict', 'technology', 'agriculture', 'agric', 'coding',
        )
        social_tokens = (
            'history', 'geography', 'social', 'religious', 'cre', 'ire', 'hre',
            'business', 'commerce', 'entrepreneur', 'c.r.e', 'i.r.e',
        )
        arts_tokens = (
            'art', 'music', 'sports', 'sport', 'physical education', 'p.e', 'pe',
            'drama', 'theatre',
        )

        if any(token in text for token in stem_tokens):
            return 'STEM'
        if any(token in text for token in social_tokens):
            return 'SOCIAL_SCIENCES'
        if any(token in text for token in arts_tokens):
            return 'ARTS_SPORTS_SCIENCE'
        return ''

    # Fallback map for legacy/duplicate subject records used by old marksheets.
    category_by_name: dict[str, str] = {}
    category_by_code: dict[str, str] = {}
    category_by_folded_name: dict[str, str] = {}
    category_by_folded_code: dict[str, str] = {}
    categorized_subjects = Subject.objects.filter(school=school).exclude(subject_category='')
    for sub in categorized_subjects:
        sub_any = cast(Any, sub)
        normalized = normalize_category(str(getattr(sub_any, 'subject_category', '') or ''))
        if not normalized:
            continue
        name_key = (str(getattr(sub_any, 'name', '') or '').strip().lower())
        code_key = (str(getattr(sub_any, 'code', '') or '').strip().lower())
        short_key = (str(getattr(sub_any, 'short_name', '') or '').strip().lower())
        folded_name = fold_key(name_key)
        folded_code = fold_key(code_key)
        folded_short = fold_key(short_key)
        if name_key and name_key not in category_by_name:
            category_by_name[name_key] = normalized
        if code_key and code_key not in category_by_code:
            category_by_code[code_key] = normalized
        if folded_name and folded_name not in category_by_folded_name:
            category_by_folded_name[folded_name] = normalized
        if folded_code and folded_code not in category_by_folded_code:
            category_by_folded_code[folded_code] = normalized
        if folded_short and folded_short not in category_by_folded_name:
            category_by_folded_name[folded_short] = normalized
        if folded_short and folded_short not in category_by_folded_code:
            category_by_folded_code[folded_short] = normalized

    subject_map: dict[int, dict[str, Any]] = {}
    for ms in marksheets:
        subj = cast(Any, ms).subject
        sid = cast(Any, subj).id
        if sid not in subject_map:
            subject_map[sid] = {
                'id': sid,
                'name': cast(Any, subj).name,
                'code': (cast(Any, subj).short_name or cast(Any, subj).code or cast(Any, subj).name),
                'category': (cast(Any, subj).subject_category or ''),
            }
    subjects = sorted(subject_map.values(), key=lambda s: str(s['code']))
    subject_ids = [cast(int, s['id']) for s in subjects]

    # Teacher initials map per subject
    def _initials_from_name(full_name: str) -> str:
        parts = [p for p in re.split(r'\s+', (full_name or '').strip()) if p]
        if not parts:
            return '-'
        return ''.join([p[0].upper() for p in parts[:3]])

    teacher_initials_by_subject: dict[int, str] = {}
    assignments = (
        TeacherAssignment.objects.filter(classroom=classroom, subject_id__in=subject_ids)
        .select_related('teacher__user', 'subject')
        .order_by('subject_id', 'id')
    )
    for a in assignments:
        a_any = cast(Any, a)
        sid = cast(int, a_any.subject_id)
        if sid in teacher_initials_by_subject:
            continue
        t = a_any.teacher
        full_name = ''
        if t and getattr(t, 'user', None):
            u = cast(Any, t).user
            full_name = (u.get_full_name() or u.username or '').strip()
        teacher_initials_by_subject[sid] = _initials_from_name(full_name)
    for sid in subject_ids:
        teacher_initials_by_subject.setdefault(sid, '-')

    marksheet_by_exam_subject: dict[tuple[int, int], Any] = {}
    marksheet_ids = []
    for ms in marksheets:
        ms_any = cast(Any, ms)
        key = (cast(Any, ms_any.exam).id, cast(Any, ms_any.subject).id)
        marksheet_by_exam_subject[key] = ms
        marksheet_ids.append(ms_any.id)

    student_ids = [cast(Any, s).id for s in students]
    student_marks = list(
        StudentMark.objects.filter(
            marksheet_id__in=marksheet_ids,
            student_id__in=student_ids,
            score__isnull=False,
        ).select_related('marksheet', 'marksheet__exam', 'marksheet__subject')
    )

    comment_marks = list(
        StudentMark.objects.filter(
            marksheet_id__in=marksheet_ids,
            student_id__in=student_ids,
        )
        .exclude(comment_text__isnull=True)
        .exclude(comment_text='')
        .select_related('marksheet', 'marksheet__exam', 'marksheet__subject')
    )

    # (student_id, exam_id, subject_id) -> score/pct/level
    mark_map: dict[tuple[int, int, int], dict[str, Any]] = {}
    comment_map: dict[tuple[int, int, int], str] = {}
    # (student_id, exam_id) -> list[pct]
    exam_pct_by_student: dict[tuple[int, int], list[float]] = defaultdict(list)
    for sm in student_marks:
        sm_any = cast(Any, sm)
        ms = sm_any.marksheet
        ex = cast(Any, ms.exam)
        subj = cast(Any, ms.subject)
        out_of = float(ms.out_of or 0)
        score = float(sm_any.score or 0)
        pct = round(_normalize_pct(score, out_of), 1)
        lvl, _pts = resolve_grade_points(pct)
        key = (sm_any.student_id, ex.id, subj.id)
        mark_map[key] = {
            'score': round(score, 1),
            'out_of': out_of,
            'pct': pct,
            'level': lvl,
        }
        exam_pct_by_student[(sm_any.student_id, ex.id)].append(pct)

    for sm in comment_marks:
        sm_any = cast(Any, sm)
        ms = sm_any.marksheet
        ex = cast(Any, ms.exam)
        subj = cast(Any, ms.subject)
        key = (sm_any.student_id, ex.id, subj.id)
        comment_map[key] = (sm_any.comment_text or '').strip()

    # Exam ranking maps
    class_rank_by_exam: dict[tuple[int, int], int] = {}
    stream_rank_by_exam: dict[tuple[int, int], int] = {}
    stream_size_by_student: dict[int, int] = {}
    class_size = len(students)
    student_stream: dict[int, str] = {}
    student_stream_id: dict[int, Any] = {}
    for s in students:
        s_any = cast(Any, s)
        sid = s_any.id
        stream_name = cast(Any, s_any.stream).name if s_any.stream else 'No Stream'
        student_stream[sid] = stream_name
        student_stream_id[sid] = s_any.stream_id

    for ex in exams_in_term:
        ex_id = cast(Any, ex).id
        scoring = []
        for sid in student_ids:
            vals = exam_pct_by_student.get((sid, ex_id), [])
            avg_pct = round(sum(vals) / len(vals), 2) if vals else -1
            scoring.append((sid, avg_pct))
        scoring.sort(key=lambda t: t[1], reverse=True)

        prev_val = None
        rank = 0
        for idx, (sid, avg) in enumerate(scoring, start=1):
            if avg < 0:
                continue
            if prev_val == avg:
                pass
            else:
                rank = idx
                prev_val = avg
            class_rank_by_exam[(sid, ex_id)] = rank

        group: dict[Any, list[tuple[int, float]]] = defaultdict(list)
        for sid, avg in scoring:
            if avg < 0:
                continue
            group[student_stream_id.get(sid)].append((sid, avg))
        for stream_id, rows in group.items():
            rows.sort(key=lambda t: t[1], reverse=True)
            prev_s = None
            rnk = 0
            size = len(rows)
            for idx, (sid, avg) in enumerate(rows, start=1):
                if prev_s == avg:
                    pass
                else:
                    rnk = idx
                    prev_s = avg
                stream_rank_by_exam[(sid, ex_id)] = rnk
                stream_size_by_student[sid] = size

    # Term overall ranking per student
    term_avg_by_student: dict[int, float] = {}
    for sid in student_ids:
        exam_pairs: list[tuple[float, float]] = []
        for ex_id in exam_ids:
            vals = exam_pct_by_student.get((sid, ex_id), [])
            if not vals:
                continue
            exam_avg = round(sum(vals) / len(vals), 2)
            exam_pairs.append((exam_avg, float(exam_weights.get(ex_id, 1.0))))
        weighted_term_avg = _weighted_mean(exam_pairs)
        term_avg_by_student[sid] = weighted_term_avg if weighted_term_avg is not None else -1

    term_sorted = sorted([(sid, v) for sid, v in term_avg_by_student.items() if v >= 0], key=lambda t: t[1], reverse=True)
    class_term_rank: dict[int, int] = {}
    prev = None
    rnk = 0
    for idx, (sid, val) in enumerate(term_sorted, start=1):
        if prev == val:
            pass
        else:
            rnk = idx
            prev = val
        class_term_rank[sid] = rnk

    stream_term_rank: dict[int, int] = {}
    stream_term_sizes: dict[Any, int] = {}
    grouped_term: dict[Any, list[tuple[int, float]]] = defaultdict(list)
    for sid, val in term_sorted:
        grouped_term[student_stream_id.get(sid)].append((sid, val))
    for stream_id, rows in grouped_term.items():
        rows.sort(key=lambda t: t[1], reverse=True)
        stream_term_sizes[stream_id] = len(rows)
        prev = None
        rnk = 0
        for idx, (sid, val) in enumerate(rows, start=1):
            if prev == val:
                pass
            else:
                rnk = idx
                prev = val
            stream_term_rank[sid] = rnk

    # Term dates for remarks section (source of truth: TermDate records)
    current_term_date = TermDate.objects.filter(
        school=school, year=exam.year, term=term_filter
    ).order_by('start_date').first()
    if not current_term_date:
        # Fallback to most recent configured term date for this term if exact year is missing.
        current_term_date = TermDate.objects.filter(
            school=school, term=term_filter
        ).order_by('-year', 'start_date').first()
    next_term_date = None
    if current_term_date:
        next_term_date = TermDate.objects.filter(
            school=school,
            start_date__gt=current_term_date.end_date
        ).order_by('start_date').first()

    # Trend history (all exams for this class for each student)
    all_class_marksheets = list(
        MarkSheet.objects.filter(school_class=classroom, exam__school=school)
        .select_related('exam', 'subject')
        .order_by('exam__year', 'exam__start_date', 'exam__id')
    )
    all_marksheet_ids = [cast(Any, m).id for m in all_class_marksheets]
    all_student_marks = list(
        StudentMark.objects.filter(
            marksheet_id__in=all_marksheet_ids,
            student_id__in=student_ids,
            score__isnull=False,
        ).select_related('marksheet', 'marksheet__exam', 'marksheet__subject')
    )
    trend_buckets: dict[tuple[int, int], list[float]] = defaultdict(list)  # (student, exam) -> pcts
    exam_meta: dict[int, Any] = {}
    for sm in all_student_marks:
        sm_any = cast(Any, sm)
        ms = sm_any.marksheet
        ex = cast(Any, ms.exam)
        exam_meta[ex.id] = ex
        out_of = float(ms.out_of or 0)
        pct = (float(sm_any.score or 0) / out_of) * 100 if out_of > 0 else 0
        trend_buckets[(sm_any.student_id, ex.id)].append(round(pct, 2))

    cards = []
    school_logo_url = cast(Any, school).logo.url if getattr(cast(Any, school), 'logo', None) else ''
    school_stamp_url = cast(Any, school).stamp.url if getattr(cast(Any, school), 'stamp', None) else ''
    school_signature_url = cast(Any, school).head_signature.url if getattr(cast(Any, school), 'head_signature', None) else ''
    for s in students:
        s_any = cast(Any, s)
        sid = s_any.id
        row_items = []
        exam_totals = [0.0 for _ in exam_ids]
        subject_pct_sum = 0.0
        subject_count = 0
        pathway_points = {key: 0.0 for key in pathway_labels.keys()}
        uncategorized_subjects = 0
        subject_term_pct_index: dict[str, list[float]] = defaultdict(list)

        for subj in subjects:
            subj_id = cast(int, subj['id'])
            subj_category = normalize_category(str(subj.get('category') or ''))
            if subj_category not in pathway_labels:
                subj_name_key = str(subj.get('name') or '').strip().lower()
                subj_code_key = str(subj.get('code') or '').strip().lower()
                subj_folded_name = fold_key(subj_name_key)
                subj_folded_code = fold_key(subj_code_key)
                subj_category = (
                    category_by_name.get(subj_name_key)
                    or category_by_code.get(subj_code_key)
                    or category_by_folded_name.get(subj_folded_name)
                    or category_by_folded_code.get(subj_folded_code)
                    or ''
                )
            if subj_category not in pathway_labels:
                subj_category = infer_subject_category(str(subj.get('name') or ''), str(subj.get('code') or ''))
            item = {
                'subject': subj['name'],
                'teacher': teacher_initials_by_subject.get(subj_id, '-'),
                'exam_cells': [],
                'term_pct': '-',
                'term_level': '-',
                'teacher_comment': '-',
            }
            subj_exam_pcts: list[tuple[float, float]] = []
            for ex in exams_in_term:
                ex_id = cast(Any, ex).id
                rec = mark_map.get((sid, ex_id, subj_id))
                if rec:
                    item['exam_cells'].append({
                        'score': rec['score'],
                        'level': rec['level'],
                    })
                    subj_exam_pcts.append((float(rec['pct']), float(exam_weights.get(ex_id, 1.0))))
                    idx = exam_index[ex_id]
                    exam_totals[idx] += float(rec['score'])
                else:
                    item['exam_cells'].append({'score': '-', 'level': '-'})

            selected_exam_rec = mark_map.get((sid, exam.id, subj_id))
            selected_exam_comment = comment_map.get((sid, exam.id, subj_id))
            if selected_exam_rec or selected_exam_comment:
                item['teacher_comment'] = selected_exam_comment or '-'

            if subj_exam_pcts:
                weighted_term_pct = _weighted_mean(subj_exam_pcts)
                if weighted_term_pct is None:
                    weighted_term_pct = 0.0
                term_pct = round(float(weighted_term_pct), 1)
                lvl, pts = resolve_grade_points(term_pct)
                item['term_pct'] = term_pct
                item['term_level'] = lvl
                subject_pct_sum += term_pct
                subject_count += 1
                subject_name_key = re.sub(r'[^a-z0-9]+', '', str(subj.get('name') or '').lower())
                subject_code_key = re.sub(r'[^a-z0-9]+', '', str(subj.get('code') or '').lower())
                if subject_name_key:
                    subject_term_pct_index[subject_name_key].append(term_pct)
                if subject_code_key:
                    subject_term_pct_index[subject_code_key].append(term_pct)
                if subj_category in pathway_points:
                    pathway_points[subj_category] += float(pts or 0.0)
                else:
                    uncategorized_subjects += 1
            row_items.append(item)

        overall_avg_pct = round(subject_pct_sum / subject_count, 2) if subject_count else 0
        overall_level, _overall_points = resolve_grade_points(overall_avg_pct)
        overall_total = round(subject_pct_sum, 1)

        def _best_subject_pct(candidates: list[str]) -> float | None:
            values: list[float] = []
            for candidate in candidates:
                key = re.sub(r'[^a-z0-9]+', '', candidate.lower())
                values.extend(subject_term_pct_index.get(key, []))
            if not values:
                return None
            return max(values)

        def _meets_pathway_threshold(value: float | None) -> bool:
            if value is None:
                return False
            if eligibility_threshold is None:
                return True
            return float(value) >= float(eligibility_threshold)

        math_pct = _best_subject_pct(['math', 'mathematics'])
        english_pct = _best_subject_pct(['english'])
        kiswahili_pct = _best_subject_pct(['kiswahili', 'swahili'])
        stem_eligible = _meets_pathway_threshold(math_pct)
        social_eligible = _meets_pathway_threshold(english_pct) or _meets_pathway_threshold(kiswahili_pct)
        pathway_eligibility = {
            'STEM': stem_eligible,
            'SOCIAL_SCIENCES': social_eligible,
            'ARTS_SPORTS_SCIENCE': True,
        }

        pathway_scores = [
            {
                'key': key,
                'label': label,
                'points': round(float(pathway_points.get(key, 0.0)), 2),
                'eligible': bool(pathway_eligibility.get(key, True)),
            }
            for key, label in pathway_labels.items()
        ]
        best_pathway = '-'
        eligible_scores = [s for s in pathway_scores if bool(s.get('eligible', True))]
        if eligible_scores:
            best = max(eligible_scores, key=lambda x: float(x['points']))
            if float(best['points']) > 0:
                best_pathway = str(best['label'])

        # Per-exam positions table
        per_exam_positions = []
        for ex in exams_in_term:
            ex_id = cast(Any, ex).id
            s_rank = stream_rank_by_exam.get((sid, ex_id)) if show_ranking else None
            c_rank = class_rank_by_exam.get((sid, ex_id)) if show_ranking else None
            s_size = stream_size_by_student.get(sid, 0) if show_ranking else 0
            per_exam_positions.append({
                'exam': cast(Any, ex).title,
                'stream_pos': (f"{s_rank} of {s_size}" if s_rank else "-"),
                'class_pos': (f"{c_rank} of {class_size}" if c_rank else "-"),
            })

        # Trend points
        trend_points = []
        ordered_exams = sorted(
            exam_meta.items(),
            key=lambda kv: (
                cast(Any, kv[1]).year,
                exam_order_rank.get(cast(Any, kv[1]).term, 9),
                cast(Any, kv[1]).start_date,
                cast(Any, kv[1]).title,
            ),
        )
        trend_exam_labels = [
            f"{cast(Any, ex).year} {cast(Any, ex).term} {cast(Any, ex).title}"
            for _ex_id, ex in ordered_exams
        ]

        for ex_id, ex in sorted(
            exam_meta.items(),
            key=lambda kv: (
                cast(Any, kv[1]).year,
                exam_order_rank.get(cast(Any, kv[1]).term, 9),
                cast(Any, kv[1]).start_date,
                cast(Any, kv[1]).title,
            ),
        ):
            vals = trend_buckets.get((sid, ex_id), [])
            if not vals:
                continue
            trend_points.append({
                'label': f"{cast(Any, ex).year} {cast(Any, ex).term} {cast(Any, ex).title}",
                'value': round(sum(vals) / len(vals), 2),
            })

        cards.append({
            'student': {
                'id': sid,
                'adm': s_any.admission_number,
                'name': f"{s_any.first_name} {s_any.last_name}".strip(),
                'class': classroom.name,
                'stream': (cast(Any, s_any.stream).name if s_any.stream else 'No Stream'),
                'term': term_filter,
                'exam': exam.title,
                'photo_url': (s_any.photo.url if getattr(s_any, 'photo', None) else ''),
            },
            'exams': [{'id': cast(Any, e).id, 'title': cast(Any, e).title} for e in exams_in_term],
            'rows': row_items,
            'exam_totals': [round(v, 1) for v in exam_totals],
            'positions': {
                'per_exam': per_exam_positions,
                'term_stream': (
                    f"{stream_term_rank.get(sid)} of {stream_term_sizes.get(student_stream_id.get(sid), 0)}"
                    if show_ranking and stream_term_rank.get(sid) else "-"
                ),
                'term_class': (f"{class_term_rank.get(sid)} of {class_size}" if show_ranking and class_term_rank.get(sid) else "-"),
            },
            'summary': {
                'overall_total': overall_total,
                'average_pct': overall_avg_pct,
                'overall_level': overall_level,
            },
            'pathway': {
                'priority': best_pathway,
                'scores': pathway_scores,
                'uncategorized_subjects': uncategorized_subjects,
                'eligibility_threshold': eligibility_threshold,
            },
            'trend_exam_labels': trend_exam_labels,
            'trend_points': trend_points,
            'term_dates': {
                'closing_date': (current_term_date.end_date.isoformat() if current_term_date else ''),
                'opening_date': (next_term_date.start_date.isoformat() if next_term_date else ''),
            },
        })

    generated_at = timezone.localtime(timezone.now())
    generated_at_display = generated_at.strftime('%d %b %Y, %I:%M %p %Z')

    return JsonResponse({
        'success': True,
        'school': {
            'name': cast(Any, school).name,
            'motto': getattr(cast(Any, school), 'motto', '') or '',
            'address': getattr(cast(Any, school), 'address', '') or '',
            'phone': getattr(cast(Any, school), 'phone', '') or '',
            'email': getattr(cast(Any, school), 'email', '') or '',
            'school_type': cast(Any, school).get_school_type_display() if getattr(cast(Any, school), 'school_type', None) else '',
            'system_type': cast(Any, school).get_system_type_display() if getattr(cast(Any, school), 'system_type', None) else '',
            'school_category': cast(Any, school).get_school_category_display() if getattr(cast(Any, school), 'school_category', None) else '',
            'logo_url': school_logo_url,
            'stamp_url': school_stamp_url,
            'signature_url': school_signature_url,
        },
        'meta': {
            'class': classroom.name,
            'term': term_filter,
            'exam': exam.title,
            'count': len(cards),
            'generated_at': generated_at_display,
            'timezone': timezone.get_current_timezone_name(),
            'show_ranking': show_ranking,
            'exam_weights': exam_weights,
        },
        'cards': cards,
    })


@login_required
def send_reports(request):
    school, denied = _require_school_permission(request, 'academics')
    if denied:
        return denied
    return render(request, 'schools/send_reports.html', {'school': school})


def _slugify_name_for_username(value: str) -> str:
    return ''.join(ch for ch in (value or '').lower() if ch.isalnum())


def _build_teacher_credentials(school, first_name: str, last_name: str, email_input: str):
    User = get_user_model()
    email_value = (email_input or '').strip().lower()
    base_user = f"{_slugify_name_for_username(first_name)}.{_slugify_name_for_username(last_name)}".strip('.')
    if not base_user:
        base_user = 'teacher'

    school_slug = _slugify_name_for_username(getattr(school, 'name', '') or 'school')
    domain = f"{school_slug}.skulplus.com"

    if email_value:
        exists = User.objects.filter(Q(username=email_value) | Q(email=email_value)).exists()
        if exists:
            return None, None, 'A user with that email already exists.'
        return email_value, email_value, ''

    username_seed = base_user
    counter = 1
    while True:
        candidate_email = f"{username_seed}@{domain}"
        exists = User.objects.filter(Q(username=candidate_email) | Q(email=candidate_email)).exists()
        if not exists:
            return candidate_email, candidate_email, ''
        counter += 1
        username_seed = f"{base_user}{counter}"


def _validate_class_teacher_selection(school, is_class_teacher: bool, class_teacher_for, class_teacher_stream):
    if not is_class_teacher:
        return True, None, None, ''
    if not class_teacher_for:
        return False, None, None, 'Select a class for class teacher assignment.'
    try:
        classroom = ClassRoom.objects.get(id=class_teacher_for, school=school)
    except ClassRoom.DoesNotExist:
        return False, None, None, 'Invalid class selection.'

    stream = None
    if class_teacher_stream:
        try:
            stream = Stream.objects.get(id=class_teacher_stream, classroom=classroom)
        except Stream.DoesNotExist:
            return False, None, None, 'Invalid stream for selected class.'
    return True, classroom, stream, ''


def _apply_class_teacher_assignment(teacher, classroom, stream):
    # Clear previous assignments first.
    StreamClassTeacher.objects.filter(teacher=teacher).delete()
    ClassRoom.objects.filter(class_teacher=teacher).update(class_teacher=None)
    if not classroom:
        return
    if stream:
        stream_class_teacher, _created = StreamClassTeacher.objects.get_or_create(
            classroom=classroom,
            stream=stream,
        )
        stream_class_teacher.teacher = teacher
        stream_class_teacher.save()
    else:
        classroom.class_teacher = teacher
        classroom.save(update_fields=['class_teacher'])


def _ensure_staff_record_for_user(school, user, phone: str = '', is_teacher: bool = False, staff_id=None):
    staff = None
    if staff_id:
        staff = Staff.objects.filter(id=staff_id, school=school).first()
    if not staff and user.email:
        staff = Staff.objects.filter(school=school, email__iexact=user.email).first()

    full_name = user.get_full_name().strip() or user.username
    phone_value = (phone or '').strip()
    role_label = 'Teacher' if is_teacher else 'Staff'

    if not staff:
        staff = Staff.objects.create(
            school=school,
            full_name=full_name,
            role=role_label,
            employee_number='',
            national_id='',
            phone=phone_value,
            email=(user.email or '').strip(),
            kra_pin='N/A',
            nssf_number='N/A',
            nhif_number='',
            basic_salary=Decimal('0'),
            is_teacher=bool(is_teacher),
        )
        return staff

    staff.full_name = full_name
    if phone_value:
        staff.phone = phone_value
    if user.email:
        staff.email = user.email
    if is_teacher:
        staff.is_teacher = True
        if not staff.role:
            staff.role = 'Teacher'
    staff.save()
    return staff


def _split_full_name(value: str):
    parts = [p for p in (value or '').strip().split(' ') if p]
    if not parts:
        return '', ''
    if len(parts) == 1:
        return parts[0], ''
    return parts[0], ' '.join(parts[1:])


def _parent_password_for_student(student: Student) -> str:
    admission = (student.admission_number or '').strip()
    first_name = (student.first_name or '').strip()
    return f'{admission}{first_name}'


def _sync_parent_account_for_student(student: Student):
    phone = (student.parent_phone or '').strip()
    if not phone:
        if student.parent_user_id:
            student.parent_user = None
            student.save(update_fields=['parent_user'])
        return '', '', ''

    User = get_user_model()
    username = phone
    password = _parent_password_for_student(student)
    first_name, last_name = _split_full_name(student.parent_name or '')
    first_name = first_name or 'Parent'

    user = User.objects.filter(username=username).first()
    if user and (hasattr(user, 'headteacher') or hasattr(user, 'teacher')):
        return '', '', 'Parent phone is already used by a staff account.'

    if user is None:
        user = User.objects.create_user(
            username=username,
            email=f'parent.{student.id}@parents.skulplus.local',
            password=password,
            first_name=first_name,
            last_name=last_name,
        )
    else:
        user.first_name = first_name
        user.last_name = last_name
        if not user.email:
            user.email = f'parent.{student.id}@parents.skulplus.local'
        user.set_password(password)
        user.save()

    if student.parent_user_id != user.id:
        student.parent_user = user
        student.save(update_fields=['parent_user'])

    return username, password, ''


@login_required
def new_user(request):
    school = resolve_user_school(request.user)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not school or not has_full_headteacher_access(request.user, school):
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Access denied.'}, status=403)
        return HttpResponseForbidden()
    teachers = Teacher.objects.filter(school=school).select_related('user').order_by('user__first_name', 'user__last_name')
    staff_qs = Staff.objects.filter(school=school).order_by('full_name')
    classes_qs = ClassRoom.objects.filter(school=school)
    classes = classes_qs.order_by('name')

    if request.method == 'POST' and is_ajax:
        try:
            data = json.loads(request.body or '{}')
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid payload.'}, status=400)

        first_name = (data.get('first_name') or '').strip()
        last_name = (data.get('last_name') or '').strip()
        email_input = (data.get('email') or '').strip().lower()
        password = (data.get('password') or '').strip()
        phone = (data.get('phone') or '').strip()
        user_type = (data.get('user_type') or 'teacher').strip().lower()
        teacher_id = data.get('teacher_id')
        staff_id = data.get('staff_id')
        is_class_teacher = bool(data.get('is_class_teacher'))
        class_teacher_for = data.get('class_teacher_for')
        class_teacher_stream = data.get('class_teacher_stream')

        if user_type not in ('teacher', 'staff'):
            return JsonResponse({'success': False, 'error': 'Choose a valid user type.'}, status=400)
        if len(password) < 6:
            return JsonResponse({'success': False, 'error': 'Password must be at least 6 characters.'}, status=400)

        selected_teacher = None
        selected_staff = None
        if user_type == 'teacher':
            if not teacher_id:
                return JsonResponse({'success': False, 'error': 'Select a teacher record.'}, status=400)
            selected_teacher = Teacher.objects.filter(id=teacher_id, school=school).select_related('user').first()
            if not selected_teacher:
                return JsonResponse({'success': False, 'error': 'Teacher record not found.'}, status=404)
            if not first_name:
                first_name = selected_teacher.user.first_name
            if not last_name:
                last_name = selected_teacher.user.last_name
            if not email_input:
                email_input = (selected_teacher.user.email or '').strip().lower()
        else:
            if not staff_id:
                return JsonResponse({'success': False, 'error': 'Select a staff record.'}, status=400)
            selected_staff = Staff.objects.filter(id=staff_id, school=school).first()
            if not selected_staff:
                return JsonResponse({'success': False, 'error': 'Staff record not found.'}, status=404)
            s_first, s_last = _split_full_name(selected_staff.full_name)
            if not first_name:
                first_name = s_first
            if not last_name:
                last_name = s_last
            if not email_input:
                email_input = (selected_staff.email or '').strip().lower()
            if not phone:
                phone = (selected_staff.phone or '').strip()

        if not first_name:
            return JsonResponse({'success': False, 'error': 'First name is required.'}, status=400)

        classroom = None
        stream = None
        if user_type == 'teacher':
            valid, classroom, stream, err = _validate_class_teacher_selection(
                school, is_class_teacher, class_teacher_for, class_teacher_stream
            )
            if not valid:
                return JsonResponse({'success': False, 'error': err}, status=400)
        else:
            is_class_teacher = False

        teacher = None
        try:
            User = get_user_model()
            if user_type == 'teacher':
                teacher = selected_teacher
                user = teacher.user
                user_id = cast(Any, user).id
                if email_input:
                    conflict = User.objects.filter(Q(username=email_input) | Q(email=email_input)).exclude(id=user_id).exists()
                    if conflict:
                        return JsonResponse({'success': False, 'error': 'A user with that email already exists.'}, status=400)
                    user.email = email_input
                    user.username = email_input
                user.first_name = first_name
                user.last_name = last_name
                user.set_password(password)
                user.save()
                teacher.is_class_teacher = is_class_teacher
                teacher.save(update_fields=['is_class_teacher'])
                _apply_class_teacher_assignment(teacher, classroom, stream)
                _ensure_staff_record_for_user(
                    school=school,
                    user=user,
                    phone=phone,
                    is_teacher=True,
                    staff_id=staff_id,
                )
            else:
                existing_user = None
                if email_input:
                    existing_user = User.objects.filter(Q(username=email_input) | Q(email=email_input)).first()

                if existing_user:
                    user = existing_user
                    user_id = cast(Any, user).id
                    conflict = User.objects.filter(Q(username=email_input) | Q(email=email_input)).exclude(id=user_id).exists()
                    if conflict:
                        return JsonResponse({'success': False, 'error': 'A user with that email already exists.'}, status=400)
                    user.username = email_input
                    user.email = email_input
                    user.first_name = first_name
                    user.last_name = last_name
                    user.set_password(password)
                    user.save()
                else:
                    username, email, cred_err = _build_teacher_credentials(school, first_name, last_name, email_input)
                    if cred_err:
                        return JsonResponse({'success': False, 'error': cred_err}, status=400)
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                    )

                _ensure_staff_record_for_user(
                    school=school,
                    user=user,
                    phone=phone,
                    is_teacher=False,
                    staff_id=cast(Any, selected_staff).id if selected_staff else staff_id,
                )

                if selected_staff:
                    selected_staff.full_name = f"{first_name} {last_name}".strip() or selected_staff.full_name
                    if phone:
                        selected_staff.phone = phone
                    if email_input:
                        selected_staff.email = email_input
                    selected_staff.save()
        except Exception as exc:
            return JsonResponse({'success': False, 'error': f'Error creating user: {exc}'}, status=500)

        return JsonResponse({
            'success': True,
            'user': {
                'id': cast(Any, user).id,
                'name': user.get_full_name() or user.username,
                'email': user.email,
                'user_type': user_type,
                'is_class_teacher': bool(teacher.is_class_teacher) if teacher else False,
            }
        })

    return render(request, 'schools/new_user.html', {
        'school': school,
        'classes': classes,
        'teachers': teachers,
        'staff_list': staff_qs,
        'teacher_records': [
            {
                'id': cast(Any, t).id,
                'name': (t.user.get_full_name() or t.user.username),
                'first_name': t.user.first_name or '',
                'last_name': t.user.last_name or '',
                'email': t.user.email or '',
            }
            for t in teachers
        ],
        'staff_records': [
            {
                'id': cast(Any, s).id,
                'name': s.full_name or '',
                'email': s.email or '',
                'phone': s.phone or '',
            }
            for s in staff_qs
        ],
    })


@login_required
def user_updates(request):
    school = resolve_user_school(request.user)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not school or not has_full_headteacher_access(request.user, school):
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Access denied.'}, status=403)
        return HttpResponseForbidden()
    classes = ClassRoom.objects.filter(school=school).order_by('name')
    teachers = list(Teacher.objects.filter(school=school).select_related('user'))
    teachers_by_user_id = {cast(Any, t.user).id: t for t in teachers}
    staff_list = list(Staff.objects.filter(school=school))
    staff_by_email = {(s.email or '').strip().lower(): s for s in staff_list if (s.email or '').strip()}

    if request.method == 'POST' and is_ajax:
        try:
            data = json.loads(request.body or '{}')
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid payload.'}, status=400)

        action = (data.get('action') or 'update').strip().lower()
        user_id = data.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User is required.'}, status=400)

        User = get_user_model()
        user = User.objects.filter(id=user_id).first()
        if not user:
            return JsonResponse({'success': False, 'error': 'User not found.'}, status=404)

        user_email_key = (user.email or '').strip().lower()
        teacher = Teacher.objects.filter(user=user, school=school).first()
        staff_match = staff_by_email.get(user_email_key)
        if not teacher and not staff_match:
            return JsonResponse({'success': False, 'error': 'User is not in this school records.'}, status=403)

        if action == 'delete':
            try:
                user.delete()
                return JsonResponse({'success': True})
            except Exception as exc:
                return JsonResponse({'success': False, 'error': str(exc)}, status=500)

        first_name = (data.get('first_name') or user.first_name).strip()
        last_name = (data.get('last_name') or user.last_name).strip()
        email = (data.get('email') or user.email).strip().lower()
        password = (data.get('password') or '').strip()
        phone = (data.get('phone') or '').strip()
        is_active = bool(data.get('is_active', True))
        role_teacher = bool(data.get('is_teacher'))
        role_staff = bool(data.get('is_staff'))
        access_role = (data.get('access_role') or '').strip().upper()
        is_class_teacher = bool(data.get('is_class_teacher'))
        class_teacher_for = data.get('class_teacher_for')
        class_teacher_stream = data.get('class_teacher_stream')

        if not role_teacher and not role_staff:
            return JsonResponse({'success': False, 'error': 'User must be either Teacher or Staff.'}, status=400)
        if not first_name or not last_name:
            return JsonResponse({'success': False, 'error': 'First and last name are required.'}, status=400)
        if not email:
            return JsonResponse({'success': False, 'error': 'Email is required.'}, status=400)

        email_conflict = User.objects.filter(Q(email=email) | Q(username=email)).exclude(id=cast(Any, user).id).exists()
        if email_conflict:
            return JsonResponse({'success': False, 'error': 'A user with that email already exists.'}, status=400)

        classroom = None
        stream = None
        if role_teacher:
            valid, classroom, stream, err = _validate_class_teacher_selection(
                school, is_class_teacher, class_teacher_for, class_teacher_stream
            )
            if not valid:
                return JsonResponse({'success': False, 'error': err}, status=400)
        else:
            is_class_teacher = False

        if password and len(password) < 6:
            return JsonResponse({'success': False, 'error': 'Password must be at least 6 characters.'}, status=400)

        try:
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.username = email
            user.is_active = is_active
            if password:
                user.set_password(password)
            user.save()

            if role_teacher:
                if not teacher:
                    teacher = Teacher.objects.create(user=user, school=school, is_class_teacher=is_class_teacher)
                teacher.is_class_teacher = is_class_teacher
                teacher.save(update_fields=['is_class_teacher'])
                _apply_class_teacher_assignment(teacher, classroom, stream)
            elif teacher:
                _apply_class_teacher_assignment(teacher, None, None)
                teacher.delete()

            if role_staff:
                _ensure_staff_record_for_user(
                    school=school,
                    user=user,
                    phone=phone,
                    is_teacher=role_teacher,
                    staff_id=(cast(Any, staff_match).id if staff_match else None),
                )
            role_obj = SchoolUserAccess.objects.filter(user=user).first()
            if access_role in dict(SchoolUserAccess.ROLE_CHOICES):
                if role_obj is None:
                    SchoolUserAccess.objects.create(
                        school=school,
                        user=user,
                        role=access_role,
                        is_active=True,
                        granted_by=request.user,
                    )
                else:
                    role_obj.school = school
                    role_obj.role = access_role
                    role_obj.is_active = True
                    role_obj.granted_by = request.user
                    role_obj.save()
            elif role_obj and role_obj.school_id == cast(Any, school).id:
                role_obj.delete()
        except Exception as exc:
            return JsonResponse({'success': False, 'error': str(exc)}, status=500)

        return JsonResponse({
            'success': True,
            'user': {
                'id': cast(Any, user).id,
                'name': user.get_full_name() or user.username,
                'email': user.email,
                'is_active': user.is_active,
                'is_teacher': role_teacher,
                'is_staff': role_staff,
                'access_role': access_role,
            }
        })

    role_emails = set(staff_by_email.keys())
    teacher_user_ids = [cast(Any, t.user).id for t in teachers]
    for t in teachers:
        role_emails.add((t.user.email or '').strip().lower())

    User = get_user_model()
    users = []
    email_values = [e for e in role_emails if e]
    if teacher_user_ids or email_values:
        users = list(
            User.objects.filter(Q(id__in=teacher_user_ids) | Q(email__in=email_values))
            .order_by('first_name', 'last_name', 'email')
            .distinct()
        )

    user_rows = []
    for user in users:
        email_key = (user.email or '').strip().lower()
        teacher = teachers_by_user_id.get(cast(Any, user).id)
        staff = staff_by_email.get(email_key)
        access_role = SchoolUserAccess.objects.filter(user=user, school=school, is_active=True).values_list('role', flat=True).first() or ''

        class_teacher_for = ''
        class_teacher_stream = ''
        class_teacher_for_id = ''
        class_teacher_stream_id = ''
        if teacher:
            classroom = ClassRoom.objects.filter(class_teacher=teacher).first()
            stream_assignment = StreamClassTeacher.objects.filter(teacher=teacher).select_related('classroom', 'stream').first()
            if stream_assignment:
                class_teacher_for = f"{stream_assignment.classroom.name} {stream_assignment.classroom.section}".strip()
                class_teacher_stream = stream_assignment.stream.name
                class_teacher_for_id = cast(Any, stream_assignment.classroom).id
                class_teacher_stream_id = cast(Any, stream_assignment.stream).id
            elif classroom:
                class_teacher_for = f"{classroom.name} {classroom.section}".strip()
                class_teacher_for_id = cast(Any, classroom).id

        user_rows.append({
            'user': user,
            'teacher': teacher,
            'staff': staff,
            'is_teacher': bool(teacher),
            'is_staff': bool(staff),
            'phone': (staff.phone if staff else ''),
            'class_teacher_for': class_teacher_for,
            'class_teacher_stream': class_teacher_stream,
            'class_teacher_for_id': class_teacher_for_id,
            'class_teacher_stream_id': class_teacher_stream_id,
            'access_role': access_role,
        })

    return render(request, 'schools/user_updates.html', {
        'school': school,
        'classes': classes,
        'users_with_meta': user_rows,
        'access_role_choices': SchoolUserAccess.ROLE_CHOICES,
    })

def get_filtered_students(school, class_id, subject_id=None):
    students = Student.objects.filter(school=school, classroom_id=class_id)

    if subject_id:
        allocated_ids = SubjectAllocation.objects.filter(
            subject_id=subject_id,
            student__school=school
        ).values_list('student_id', flat=True)

        students = students.filter(id__in=allocated_ids)

    return students.order_by('first_name', 'last_name')

@login_required
@ensure_csrf_cookie
def enter_marks_page(request):
    # Keep this route for backward compatibility, but use the main enter_marks flow
    # so teacher and headteacher pages remain identical.
    return enter_marks(request)


@login_required
def entered_marks(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponseForbidden()
    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    has_academics_role = user_has_permission(request.user, school, 'academics')
    if not (is_headteacher or is_superuser or is_teacher or has_academics_role):
        return HttpResponseForbidden()

    term = (request.GET.get('term') or '').strip()
    exam_id = (request.GET.get('exam_id') or '').strip()
    class_id = (request.GET.get('class_id') or '').strip()
    stream_id = (request.GET.get('stream_id') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    lock_filter = (request.GET.get('lock') or '').strip()

    base_qs = (
        MarkSheet.objects.filter(exam__school=school)
        .select_related('exam', 'school_class', 'subject')
    )
    teacher_allowed_class_ids = set()
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        teacher_allowed_class_ids = _teacher_class_teacher_class_ids(teacher, school)
        if not teacher_allowed_class_ids:
            return HttpResponseForbidden('Entered marks is available to class teachers only.')
        base_qs = base_qs.filter(school_class_id__in=teacher_allowed_class_ids)

    selected_exam = None
    if exam_id:
        selected_exam = Exam.objects.filter(id=exam_id, school=school).first()
        if selected_exam:
            base_qs = base_qs.filter(exam=selected_exam)
    if term:
        base_qs = base_qs.filter(exam__term=term)
    if class_id:
        base_qs = base_qs.filter(school_class_id=class_id)
    if stream_id:
        base_qs = base_qs.filter(marks__student__stream_id=stream_id)

    stream_count_filter = Q()
    if stream_id:
        stream_count_filter = Q(marks__student__stream_id=stream_id)

    summary_qs = base_qs.distinct()
    summary = {
        'total_marksheets': summary_qs.count(),
        'published': summary_qs.filter(status='published').count(),
        'draft': summary_qs.filter(status='draft').count(),
        'locked': summary_qs.filter(exam__marks_entry_locked=True).count(),
    }

    marksheets_qs = base_qs
    if status_filter in ('draft', 'published'):
        marksheets_qs = marksheets_qs.filter(status=status_filter)
    if lock_filter == 'locked':
        marksheets_qs = marksheets_qs.filter(exam__marks_entry_locked=True)
    elif lock_filter == 'open':
        marksheets_qs = marksheets_qs.filter(exam__marks_entry_locked=False)

    marksheets = list(
        marksheets_qs
        .annotate(
            entered_count=Count('marks', filter=Q(marks__score__isnull=False) & stream_count_filter, distinct=True),
            total_rows=Count('marks', filter=stream_count_filter, distinct=True),
        )
        .order_by('-exam__year', 'exam__term', 'exam__title', 'school_class__name', 'subject__name')
        .distinct()
    )
    draft_row_count = sum(1 for m in marksheets if m.status == 'draft')

    # Draft print payload: raw marks per stream with out_of values, no grading/ranking.
    # Use the current filter scope (term/exam/class/stream), regardless of marksheet status,
    # because verification may be needed even after publish.
    draft_marksheets_qs = (
        base_qs
        .select_related('exam', 'school_class', 'subject')
        .order_by('school_class__name', 'exam__year', 'exam__term', 'exam__title', 'subject__name')
        .distinct()
    )
    if class_id:
        draft_marksheets_qs = draft_marksheets_qs.filter(school_class_id=class_id)

    streams_qs = Stream.objects.filter(classroom__school=school)
    if class_id:
        streams_qs = streams_qs.filter(classroom_id=class_id)
    if stream_id:
        streams_qs = streams_qs.filter(id=stream_id)
    streams_for_print = list(streams_qs.select_related('classroom').order_by('classroom__name', 'name'))

    draft_marksheets = list(draft_marksheets_qs)
    draft_marksheet_ids = [m.id for m in draft_marksheets]
    stream_class_ids = {s.classroom_id for s in streams_for_print}

    students_for_print_qs = Student.objects.filter(school=school, stream__in=streams_for_print).select_related('stream')
    if class_id:
        students_for_print_qs = students_for_print_qs.filter(classroom_id=class_id)
    students_for_print = list(students_for_print_qs.order_by('admission_number', 'last_name', 'first_name'))

    student_ids_for_print = [s.id for s in students_for_print]
    raw_mark_map = {}
    if draft_marksheet_ids and student_ids_for_print:
        raw_marks = StudentMark.objects.filter(
            marksheet_id__in=draft_marksheet_ids,
            student_id__in=student_ids_for_print,
            score__isnull=False,
        ).values('marksheet_id', 'student_id', 'score')
        for rec in raw_marks:
            raw_mark_map[(rec['marksheet_id'], rec['student_id'])] = rec['score']

    grouped_marksheets = defaultdict(list)
    for m in draft_marksheets:
        if m.school_class_id not in stream_class_ids:
            continue
        grouped_marksheets[(m.school_class_id, m.exam_id)].append(m)

    stream_students_map: dict[int, list[Any]] = defaultdict(list)
    for stu in students_for_print:
        if stu.stream_id:
            stream_students_map[stu.stream_id].append(stu)

    draft_print_streams = []
    for stream_obj in streams_for_print:
        stream_students = stream_students_map.get(stream_obj.id, [])
        class_exam_keys = [
            key for key in grouped_marksheets.keys()
            if key[0] == stream_obj.classroom_id
        ]
        class_exam_keys.sort(
            key=lambda k: (
                grouped_marksheets[k][0].exam.year,
                grouped_marksheets[k][0].exam.term,
                grouped_marksheets[k][0].exam.title,
            )
        )

        sections = []
        for key in class_exam_keys:
            sheets = sorted(grouped_marksheets[key], key=lambda x: (x.subject.name, x.id))
            if not sheets:
                continue

            rows = []
            for stu in stream_students:
                marks = []
                for ms in sheets:
                    val = raw_mark_map.get((ms.id, stu.id))
                    marks.append(val if val is not None else '-')
                rows.append({
                    'adm': stu.admission_number,
                    'name': f"{stu.first_name} {stu.last_name}".strip(),
                    'marks': marks,
                })

            first_exam = sheets[0].exam
            sections.append({
                'exam_title': first_exam.title,
                'term': first_exam.term,
                'year': first_exam.year,
                'subjects': [
                    {
                        'name': (ms.subject.short_name or ms.subject.code or ms.subject.name),
                        'out_of': ms.out_of,
                    }
                    for ms in sheets
                ],
                'rows': rows,
            })

        draft_print_streams.append({
            'class_name': stream_obj.classroom.name,
            'stream_name': stream_obj.name,
            'sections': sections,
            'student_count': len(stream_students),
        })

    classes = ClassRoom.objects.filter(school=school)
    if teacher_allowed_class_ids:
        classes = classes.filter(id__in=teacher_allowed_class_ids)
    classes = classes.order_by('order', 'name')
    streams = Stream.objects.filter(classroom__school=school)
    if class_id:
        streams = streams.filter(classroom_id=class_id)
    if teacher_allowed_class_ids:
        streams = streams.filter(classroom_id__in=teacher_allowed_class_ids)
    streams = streams.order_by('classroom__name', 'name')

    return render(request, 'schools/entered_marks.html', {
        'school': school,
        'terms': ['Term 1', 'Term 2', 'Term 3'],
        'exams': Exam.objects.filter(school=school).order_by('-year', 'term', 'title'),
        'classes': classes,
        'streams': streams,
        'selected_term': term,
        'selected_exam_id': exam_id,
        'selected_class_id': class_id,
        'selected_stream_id': stream_id,
        'selected_exam': selected_exam,
        'status_filter': status_filter,
        'lock_filter': lock_filter,
        'marksheets': marksheets,
        'summary': summary,
        'draft_row_count': draft_row_count,
        'draft_print_streams': draft_print_streams,
        'has_draft_print_data': any(s['sections'] for s in draft_print_streams),
    })


@login_required
def set_exams(request):
    school, denied = _require_school_permission(request, 'academics')
    if denied:
        return denied

    exams = Exam.objects.filter(school=school).order_by('-year', 'term', 'start_date', 'title')
    return render(request, 'schools/set_exams.html', {
        'school': school,
        'exams': exams,
    })


@login_required
@require_POST
def toggle_exam_lock(request, exam_id):
    school, denied = _require_school_permission(request, 'academics')
    if denied:
        return denied

    exam = get_object_or_404(Exam, id=exam_id, school=school)
    exam_any = cast(Any, exam)
    exam_any.marks_entry_locked = not bool(exam_any.marks_entry_locked)
    exam_any.save(update_fields=['marks_entry_locked'])
    return JsonResponse({'success': True, 'locked': bool(exam_any.marks_entry_locked)})


@login_required
def load_marks_students(request):
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    stream_id = request.GET.get('stream_id')
    exam_id = request.GET.get('exam_id')
    term = request.GET.get('term')
    out_of = request.GET.get('out_of')
    school = get_user_school(request.user)
    if not school:
        return JsonResponse({'students': [], 'error': 'No school associated with this account.'}, status=403)

    if not all([class_id, subject_id, exam_id, term, out_of]):
        return JsonResponse({'students': [], 'error': 'Missing filters. Please pick term, exam, class, subject, and out of.'}, status=400)

    is_headteacher = hasattr(request.user, 'headteacher')
    is_superuser = request.user.is_superuser
    is_teacher = hasattr(request.user, 'teacher')
    if is_teacher and not (is_headteacher or is_superuser):
        teacher = request.user.teacher
        assignment_qs = TeacherAssignment.objects.filter(
            teacher=teacher,
            classroom__school=school,
            classroom_id=class_id,
            subject_id=subject_id,
        )
        if not assignment_qs.exists():
            return JsonResponse({'students': [], 'error': 'Not allowed for this class/subject allocation.'}, status=403)
        has_all_streams = assignment_qs.filter(stream__isnull=True).exists()
        allowed_stream_ids = set(assignment_qs.exclude(stream__isnull=True).values_list('stream_id', flat=True))
        if stream_id and not has_all_streams:
            try:
                if int(stream_id) not in allowed_stream_ids:
                    return JsonResponse({'students': [], 'error': 'Not allowed for selected stream.'}, status=403)
            except (TypeError, ValueError):
                return JsonResponse({'students': [], 'error': 'Invalid stream selection.'}, status=400)
    else:
        has_all_streams = True
        allowed_stream_ids = set()

    exam_obj = Exam.objects.filter(id=exam_id, school=school).first()
    if not exam_obj:
        return JsonResponse({'students': [], 'error': 'Exam not found for this school.'}, status=404)
    if cast(Any, exam_obj).marks_entry_locked:
        return JsonResponse({'students': [], 'error': 'This exam is locked for marks entry. Ask admin to unlock it in Set Exams.'}, status=423)

    try:
        out_of_value = int(out_of)
    except (TypeError, ValueError):
        return JsonResponse({'students': [], 'error': 'Out Of must be a number.'}, status=400)

    # Load only students in this school/class and allocated the selected subject
    students = get_filtered_students(school, class_id, subject_id)
    
    # Filter by stream if provided
    if stream_id:
        students = students.filter(stream_id=stream_id)
    elif is_teacher and not (is_headteacher or is_superuser) and not has_all_streams and allowed_stream_ids:
        students = students.filter(stream_id__in=allowed_stream_ids)

    marksheet, _ = MarkSheet.objects.get_or_create(
        exam_id=exam_id,
        school_class_id=class_id,
        subject_id=subject_id,
        defaults={'term': term, 'out_of': out_of_value, 'created_by': request.user}
    )

    classroom = ClassRoom.objects.filter(id=class_id, school=school).select_related('level').first()
    if not classroom:
        return JsonResponse({'comments': [], 'suggested': '', 'performance_level': '', 'level': '', 'points': None}, status=400)

    class_level = classroom.level.name if classroom and classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    if resolved_level not in ('Lower Primary', 'Upper Primary') and classroom and classroom.name:
        band = _resolve_primary_band_from_class_name(classroom.name)
        if band:
            resolved_level = band
    is_primary_level = resolved_level in ('Lower Primary', 'Upper Primary')
    is_junior_level = school.system_type == 'CBE' and resolved_level == 'Junior'
    is_cbe_level = is_primary_level or is_junior_level
    is_cambridge_school = getattr(school, 'school_type', '') == 'CAMBRIDGE'
    resolve_grade_points, _grade_list = _build_grade_resolver_for_class(school, classroom, resolved_level)
    subject_obj = Subject.objects.filter(id=subject_id, school=school).first() if subject_id else None

    comment_cache = {}

    data = []
    for s in students:
        mark_obj, _ = StudentMark.objects.get_or_create(marksheet=marksheet, student=s)
        score_value = mark_obj.score
        stored_level = mark_obj.level or ''
        stored_points = mark_obj.points
        comment_manual = getattr(mark_obj, 'comment_manual', False)
        performance_level = ""
        suggested_comments = []
        suggested_comment = ""
        if is_cbe_level and resolved_level and score_value is not None:
            try:
                percentage = (float(score_value) / float(marksheet.out_of or out_of_value)) * 100
            except Exception:
                percentage = None
            performance_level = get_performance_level(resolved_level, percentage)
            if performance_level:
                cache_key = (resolved_level, subject_id, performance_level)
                if cache_key not in comment_cache:
                    comment_cache[cache_key] = get_comment_variants(resolved_level, subject_obj, performance_level, limit=7)
                suggested_comments = comment_cache.get(cache_key, [])
                suggested_comment = get_random_comment(
                    resolved_level,
                    subject_obj,
                    performance_level,
                    student=s,
                    term=term,
                )
        elif is_cambridge_school and score_value is not None:
            try:
                percentage = (float(score_value) / float(marksheet.out_of or out_of_value)) * 100
            except Exception:
                percentage = None
            if percentage is not None:
                level_label, _pts = resolve_grade_points(percentage)
                performance_level = level_label
                suggested_comment = _cambridge_comment_for_level(level_label)
                suggested_comments = [suggested_comment]
        data.append({
            'student_id': cast(Any, s).id,
            'name': f"{s.first_name} {s.last_name}",
            'admission': s.admission_number,
            'score': score_value if score_value is not None else '',
            'comment_text': mark_obj.comment_text or '',
            'level': stored_level,
            'points': stored_points,
            'comment_manual': comment_manual,
            'performance_level': performance_level,
            'suggested_comments': suggested_comments,
            'suggested_comment': suggested_comment,
        })

    marksheet_any = cast(Any, marksheet)
    return JsonResponse({
        'students': data,
        'marksheet_id': marksheet_any.id,
        'status': marksheet_any.status,
        'out_of': marksheet_any.out_of,
        'show_levels': is_cbe_level and resolved_level in ('Lower Primary', 'Upper Primary'),
    })


@login_required
@require_GET
def get_suggested_comments(request):
    school = get_user_school(request.user)
    if not school:
        return JsonResponse({'comments': [], 'suggested': '', 'performance_level': '', 'level': '', 'points': None}, status=403)

    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    student_id = request.GET.get('student_id')
    score = request.GET.get('score')
    out_of = request.GET.get('out_of')
    term = request.GET.get('term')

    if not all([class_id, subject_id, score, out_of, term]):
        return JsonResponse({'comments': [], 'suggested': '', 'performance_level': '', 'level': '', 'points': None}, status=400)

    try:
        score_value = float(score)
        out_of_value = float(out_of)
        percentage = (score_value / out_of_value) * 100 if out_of_value else None
    except Exception:
        return JsonResponse({'comments': [], 'suggested': '', 'performance_level': '', 'level': '', 'points': None}, status=400)

    classroom = ClassRoom.objects.filter(id=class_id, school=school).select_related('level').first()
    class_level = classroom.level.name if classroom and classroom.level else None
    resolved_level = resolve_cbe_level(school, class_level)
    if resolved_level not in ('Lower Primary', 'Upper Primary') and classroom and classroom.name:
        band = _resolve_primary_band_from_class_name(classroom.name)
        if band:
            resolved_level = band
    if resolved_level not in ('Lower Primary', 'Upper Primary', 'Junior'):
        return JsonResponse({'comments': [], 'suggested': '', 'performance_level': '', 'level': '', 'points': None})

    subject_obj = Subject.objects.filter(id=subject_id, school=school).first()
    performance_level = get_performance_level(resolved_level, percentage)
    if not performance_level:
        return JsonResponse({'comments': [], 'suggested': '', 'performance_level': '', 'level': '', 'points': None})

    student = Student.objects.filter(id=student_id, school=school).first() if student_id else None
    comments = get_comment_variants(resolved_level, subject_obj, performance_level, limit=7)
    suggested = get_random_comment(resolved_level, subject_obj, performance_level, student=student, term=term)
    level = performance_level
    points = None
    if resolved_level in ('Lower Primary', 'Upper Primary'):
        _, points = get_primary_level_and_points(percentage, resolved_level)
    elif resolved_level == 'Junior':
        points = get_junior_points(level)

    return JsonResponse({
        'comments': comments,
        'suggested': suggested,
        'performance_level': performance_level,
        'level': level,
        'points': points,
    })


@login_required
@require_POST
def save_marks(request):
    try:
        data = json.loads(request.body)
        marksheet = MarkSheet.objects.get(id=data['marksheet_id'])
        school = get_user_school(request.user)
        if not school:
            return JsonResponse({'success': False, 'error': 'No school found for this user'})
        exam_obj = getattr(marksheet, 'exam', None)
        if exam_obj and cast(Any, exam_obj).school_id != cast(Any, school).id:
            return JsonResponse({'success': False, 'error': 'Invalid marksheet for this school'}, status=403)
        if exam_obj and cast(Any, exam_obj).marks_entry_locked:
            return JsonResponse({'success': False, 'error': 'This exam is locked for marks entry. Ask admin to unlock it in Set Exams.'}, status=423)
        is_headteacher = hasattr(request.user, 'headteacher')
        is_superuser = request.user.is_superuser
        is_teacher = hasattr(request.user, 'teacher')
        if is_teacher and not (is_headteacher or is_superuser):
            teacher = request.user.teacher
            assignment_qs = TeacherAssignment.objects.filter(
                teacher=teacher,
                classroom__school=school,
                classroom_id=cast(Any, marksheet).school_class_id,
                subject_id=cast(Any, marksheet).subject_id,
            )
            if not assignment_qs.exists():
                return JsonResponse({'success': False, 'error': 'Not allowed for this class/subject allocation.'}, status=403)
            has_all_streams = assignment_qs.filter(stream__isnull=True).exists()
            allowed_stream_ids = set(assignment_qs.exclude(stream__isnull=True).values_list('stream_id', flat=True))
            if not has_all_streams and allowed_stream_ids:
                mark_rows = data.get('marks', [])
                student_ids = [m.get('student_id') for m in mark_rows if m.get('student_id') is not None]
                if student_ids:
                    stream_map = {
                        row['id']: row['stream_id']
                        for row in Student.objects.filter(
                            id__in=student_ids,
                            school=school,
                            classroom_id=cast(Any, marksheet).school_class_id,
                        ).values('id', 'stream_id')
                    }
                    for sid in student_ids:
                        if stream_map.get(sid) not in allowed_stream_ids:
                            return JsonResponse({'success': False, 'error': 'One or more students are outside your allocated streams.'}, status=403)

        school_any = cast(Any, school)
        classroom = getattr(marksheet, 'school_class', None)
        class_level = classroom.level.name if classroom and classroom.level else None
        resolved_level = resolve_cbe_level(school, class_level)
        if resolved_level not in ('Lower Primary', 'Upper Primary') and classroom and classroom.name:
            band = _resolve_primary_band_from_class_name(classroom.name)
            if band:
                resolved_level = band
        is_primary = resolved_level in ('Lower Primary', 'Upper Primary')
        is_junior = school_any.system_type == 'CBE' and resolved_level == 'Junior'
        is_cambridge = getattr(school_any, 'school_type', '') == 'CAMBRIDGE'
        resolve_grade_points, _grade_list = _build_grade_resolver_for_class(school, classroom, resolved_level)

        if marksheet.status == 'published' and not (is_headteacher or request.user.is_superuser):
            return JsonResponse({'success': False, 'error': 'Marks already published'})

        # Update out_of if provided
        if 'out_of' in data and data['out_of']:
            try:
                marksheet.out_of = int(data['out_of'])
                marksheet.save()
            except (TypeError, ValueError):
                pass

        for m in data.get('marks', []):
            score_value = m.get('score')
            level = ''
            points = None
            percentage = None
            if score_value is not None:
                try:
                    percentage = (float(score_value) / float(marksheet.out_of or 0)) * 100
                except Exception:
                    percentage = None
            if is_primary and percentage is not None:
                level, points = get_primary_level_and_points(percentage, resolved_level)
            if is_junior and percentage is not None:
                level = get_junior_level(percentage)
                points = get_junior_points(level)
            if is_cambridge and percentage is not None:
                level, points = resolve_grade_points(percentage)

            comment_text = (m.get('comment_text', '') or '').strip()
            comment_manual = bool(m.get('comment_manual'))
            if not comment_manual and percentage is not None and resolved_level in ('Lower Primary', 'Upper Primary', 'Junior'):
                performance_level = get_performance_level(resolved_level, percentage)
                subject_obj = getattr(marksheet, 'subject', None)
                comment_text = get_random_comment(
                    resolved_level,
                    subject_obj,
                    performance_level,
                )
            elif not comment_manual and percentage is not None and is_cambridge:
                comment_text = _cambridge_comment_for_level(level)

            StudentMark.objects.update_or_create(
                marksheet=marksheet,
                student_id=m['student_id'],
                defaults={
                    'score': score_value,
                    'comment_text': comment_text,
                    'level': level,
                    'points': points,
                    'comment_manual': comment_manual,
                }
            )

        # Set marksheet status to draft after editing
        if marksheet.status != 'draft':
            marksheet.status = 'draft'
            marksheet.save(update_fields=['status'])

        return JsonResponse({'success': True})
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


@login_required
@require_POST
def publish_marks(request, marksheet_id):
    try:
        school = get_user_school(request.user)
        if not school:
            return JsonResponse({'success': False, 'error': 'No school found for this user'}, status=403)
        marksheet = MarkSheet.objects.get(id=marksheet_id)
        exam_obj = getattr(marksheet, 'exam', None)
        if exam_obj and cast(Any, exam_obj).school_id != cast(Any, school).id:
            return JsonResponse({'success': False, 'error': 'Invalid marksheet for this school'}, status=403)
        exam_obj = getattr(marksheet, 'exam', None)
        if exam_obj and cast(Any, exam_obj).marks_entry_locked:
            return JsonResponse({'success': False, 'error': 'This exam is locked for marks entry. Unlock it in Set Exams first.'}, status=423)
        marksheet.status = 'published'
        marksheet.save()
        return JsonResponse({'success': True})
    except MarkSheet.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Marksheet not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def unpublish_marks(request, marksheet_id):
    try:
        school = get_user_school(request.user)
        if not school:
            return JsonResponse({'success': False, 'error': 'No school found for this user'}, status=403)
        marksheet = MarkSheet.objects.get(id=marksheet_id)
        exam_obj = getattr(marksheet, 'exam', None)
        if exam_obj and cast(Any, exam_obj).school_id != cast(Any, school).id:
            return JsonResponse({'success': False, 'error': 'Invalid marksheet for this school'}, status=403)
        exam_obj = getattr(marksheet, 'exam', None)
        if exam_obj and cast(Any, exam_obj).marks_entry_locked:
            return JsonResponse({'success': False, 'error': 'This exam is locked for marks entry. Unlock it in Set Exams first.'}, status=423)
        marksheet.status = 'draft'
        marksheet.save()
        return JsonResponse({'success': True})
    except MarkSheet.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Marksheet not found'})







