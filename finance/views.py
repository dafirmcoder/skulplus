from collections import OrderedDict
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
import mimetypes
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import DecimalField, ExpressionWrapper, F, Sum
from django.db import transaction
from django.http import FileResponse, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from schools.models import ClassRoom, Student, TermDate
from schools.access import get_user_school, user_has_permission
from .forms import ExpenditureForm
from .models import Expenditure, FeePayment, FeePaymentAllocation, FeeStructure
from .sms import send_fee_reminder


def _headteacher_school(request):
    school = get_user_school(request.user)
    if school and (request.user.is_superuser or user_has_permission(request.user, school, 'finance')):
        return school
    return None


def _terms_for_structure(item, selected_term=None):
    term_order = [value for value, _ in FeeStructure.TERM_CHOICES]
    if selected_term:
        return [selected_term]
    if item.billing_mode == FeeStructure.BILLING_MODE_ONCE_YEAR and item.due_term:
        return [item.due_term]
    if item.billing_mode == FeeStructure.BILLING_MODE_SELECTED_TERMS:
        selected_terms = item.applied_terms or []
        terms = [t for t in term_order if t in selected_terms]
        extras = [t for t in selected_terms if t not in terms]
        return terms + extras
    return []


def _term_order():
    return [value for value, _ in FeeStructure.TERM_CHOICES]


def _period_index(term_value):
    order = _term_order()
    try:
        return order.index(term_value)
    except ValueError:
        return len(order)


def _is_period_lte(year_a, term_a, year_b, term_b):
    return (int(year_a), _period_index(term_a)) <= (int(year_b), _period_index(term_b))


def _student_due_map_upto(student, upto_term, upto_year):
    due_map = OrderedDict()
    if not student or not student.classroom_id:
        return due_map
    structures = (
        FeeStructure.objects
        .filter(school=student.school, year__lte=upto_year, applicable_classes=student.classroom)
        .distinct()
        .order_by('year', 'vote_head', 'id')
    )
    for item in structures:
        for item_term in _terms_for_structure(item):
            if not _is_period_lte(item.year, item_term, upto_year, upto_term):
                continue
            key = (item.year, item_term, item.vote_head)
            due_map[key] = due_map.get(key, Decimal('0')) + (item.amount or Decimal('0'))
    return due_map


def _student_outstanding_meta(student, upto_term, upto_year):
    if not student:
        return {
            'due_total': Decimal('0'),
            'paid_total': Decimal('0'),
            'balance': Decimal('0'),
            'votehead_rows': [],
            'outstanding_rows': [],
            'target_options': [],
            'target_balance_map': {},
        }

    try:
        upto_year = int(upto_year)
    except (TypeError, ValueError):
        upto_year = date.today().year

    due_map = _student_due_map_upto(student, upto_term, upto_year)
    due_total = sum(due_map.values(), Decimal('0'))

    paid_map = {}
    allocations = (
        FeePaymentAllocation.objects
        .filter(fee_payment__student=student)
        .values('allocation_year', 'allocation_term', 'vote_head')
        .annotate(paid=Sum('amount'))
    )
    for row in allocations:
        if not _is_period_lte(row['allocation_year'], row['allocation_term'], upto_year, upto_term):
            continue
        key = (row['allocation_year'], row['allocation_term'], row['vote_head'])
        paid_map[key] = paid_map.get(key, Decimal('0')) + (row['paid'] or Decimal('0'))

    # Backward compatibility for old payments that had no split rows.
    unallocated_payments = (
        FeePayment.objects
        .filter(student=student)
        .exclude(allocations__isnull=False)
        .values('year', 'term')
        .annotate(total=Sum('amount_paid'))
    )
    for row in unallocated_payments:
        period_due_keys = [k for k in due_map.keys() if k[0] == row['year'] and k[1] == row['term']]
        remaining = row['total'] or Decimal('0')
        for key in sorted(period_due_keys, key=lambda x: x[2]):
            if remaining <= 0:
                break
            key_due = due_map.get(key, Decimal('0'))
            key_paid = paid_map.get(key, Decimal('0'))
            key_balance = key_due - key_paid
            if key_balance <= 0:
                continue
            move = min(remaining, key_balance)
            paid_map[key] = key_paid + move
            remaining -= move

    outstanding_rows = []
    votehead_due = OrderedDict()
    votehead_paid = OrderedDict()
    target_options = []
    target_balance_map = {}

    for key in sorted(due_map.keys(), key=lambda k: (k[0], _period_index(k[1]), k[2])):
        due_amt = due_map[key]
        paid_amt = paid_map.get(key, Decimal('0'))
        bal_amt = due_amt - paid_amt
        year_v, term_v, head_v = key
        votehead_due[head_v] = votehead_due.get(head_v, Decimal('0')) + due_amt
        votehead_paid[head_v] = votehead_paid.get(head_v, Decimal('0')) + paid_amt
        if bal_amt > 0:
            target_key = f'{year_v}||{term_v}||{head_v}'
            label = f'{year_v} {term_v} - {head_v}'
            outstanding_rows.append({
                'key': target_key,
                'year': year_v,
                'term': term_v,
                'vote_head': head_v,
                'due': due_amt,
                'paid': paid_amt,
                'balance': bal_amt,
                'label': label,
            })
            target_options.append({'key': target_key, 'label': label})
            target_balance_map[target_key] = bal_amt

    votehead_rows = []
    for head in votehead_due.keys():
        votehead_rows.append({
            'vote_head': head,
            'due': votehead_due.get(head, Decimal('0')),
            'paid': votehead_paid.get(head, Decimal('0')),
        })

    paid_total = sum((paid_map.get(k, Decimal('0')) for k in due_map.keys()), Decimal('0'))
    balance = sum((row['balance'] for row in outstanding_rows), Decimal('0'))
    return {
        'due_total': due_total,
        'paid_total': paid_total,
        'balance': balance,
        'votehead_rows': votehead_rows,
        'outstanding_rows': outstanding_rows,
        'target_options': target_options,
        'target_balance_map': target_balance_map,
    }


def _financial_period_range(school, period, term, year, anchor_date):
    today = timezone.localdate()
    period_value = (period or 'term').strip().lower()
    resolved_anchor = anchor_date or today
    warning = ''

    if period_value == 'day':
        start_date = resolved_anchor
        end_date = resolved_anchor
        label = f'Daily ({start_date.isoformat()})'
        return start_date, end_date, label, warning

    if period_value == 'week':
        start_date = resolved_anchor - timedelta(days=resolved_anchor.weekday())
        end_date = start_date + timedelta(days=6)
        label = f'Weekly ({start_date.isoformat()} to {end_date.isoformat()})'
        return start_date, end_date, label, warning

    if period_value == 'month':
        start_date = resolved_anchor.replace(day=1)
        if start_date.month == 12:
            next_month_start = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            next_month_start = start_date.replace(month=start_date.month + 1, day=1)
        end_date = next_month_start - timedelta(days=1)
        label = f'Monthly ({start_date.strftime("%b %Y")})'
        return start_date, end_date, label, warning

    if period_value == 'year':
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        label = f'Yearly ({year})'
        return start_date, end_date, label, warning

    # Default / explicit term: use configured school term dates.
    if not TermDate.objects.filter(school=school).exists():
        term = 'Term 1'
    term_date = TermDate.objects.filter(school=school, year=year, term=term).first()
    if term_date:
        start_date = term_date.start_date
        end_date = term_date.end_date
        label = f'{term} {year} ({start_date.isoformat()} to {end_date.isoformat()})'
        return start_date, end_date, label, warning

    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    warning = f'Term dates for {term} {year} are not configured. Using full year range.'
    label = f'{term} {year} (fallback full year)'
    return start_date, end_date, label, warning


@login_required
def fee_structure(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can manage fee structure.')

    term = request.GET.get('term', '').strip()
    year = request.GET.get('year', '').strip()
    class_ids = request.GET.getlist('classrooms')

    structures_qs = (
        FeeStructure.objects
        .filter(school=school)
        .prefetch_related('applicable_classes')
        .order_by('-year', 'vote_head', '-id')
    )

    structures = structures_qs
    if class_ids:
        structures = structures.filter(applicable_classes__id__in=class_ids).distinct()
    if year:
        structures = structures.filter(year=year)

    if term:
        structures = [
            item for item in structures
            if (
                (item.billing_mode == FeeStructure.BILLING_MODE_ONCE_YEAR and item.due_term == term)
                or
                (item.billing_mode == FeeStructure.BILLING_MODE_SELECTED_TERMS and term in (item.applied_terms or []))
            )
        ]

    if not isinstance(structures, list):
        structures = list(structures.distinct())

    classes = ClassRoom.objects.filter(school=school).order_by('name')

    if request.method == 'POST':
        action = request.POST.get('action', 'save').strip().lower()
        structure_id = request.POST.get('structure_id', '').strip()

        if action == 'delete':
            if not structure_id:
                messages.error(request, 'Missing fee structure record to delete.')
                return redirect('fee_structure')

            item = FeeStructure.objects.filter(id=structure_id, school=school).first()
            if not item:
                messages.error(request, 'Fee structure record not found.')
                return redirect('fee_structure')

            item.delete()
            messages.success(request, 'Fee structure record deleted.')
            return redirect('fee_structure')

        vote_head_value = request.POST.get('vote_head', '').strip()
        amount_value = request.POST.get('amount', '').strip()
        year_value = request.POST.get('year', '').strip()
        billing_mode = request.POST.get('billing_mode', FeeStructure.BILLING_MODE_SELECTED_TERMS)
        due_term = request.POST.get('due_term', '').strip()
        applied_terms = request.POST.getlist('applied_terms')
        class_ids = request.POST.getlist('classrooms')

        selected_classes = classes.filter(id__in=class_ids)
        if not selected_classes.exists():
            messages.error(request, 'Select at least one class for the vote head.')
            return redirect('fee_structure')

        if not vote_head_value:
            messages.error(request, 'Vote head is required.')
            return redirect('fee_structure')

        try:
            year_int = int(year_value or date.today().year)
        except ValueError:
            messages.error(request, 'Year must be a number.')
            return redirect('fee_structure')

        amount = amount_value or 0

        item = None
        if structure_id:
            item = FeeStructure.objects.filter(id=structure_id, school=school).first()

        is_update = item is not None
        if item is None:
            item = FeeStructure(school=school)

        item.vote_head = vote_head_value
        item.amount = amount
        item.year = year_int
        item.billing_mode = billing_mode
        item.due_term = due_term
        item.applied_terms = applied_terms

        try:
            item.full_clean()
            item.save()
            item.applicable_classes.set(selected_classes)
            if is_update:
                messages.success(request, 'Fee structure vote head updated successfully.')
            else:
                messages.success(request, 'Fee structure vote head added successfully.')
            return redirect('fee_structure')
        except Exception as exc:
            messages.error(request, f'Unable to save fee structure: {exc}')

    term_order = [value for value, _ in FeeStructure.TERM_CHOICES]
    term_votehead_groups = OrderedDict((t, []) for t in term_order)
    class_groups_map = OrderedDict()

    for item in structures:
        amount = item.amount or Decimal('0')

        terms_for_item = []
        if term:
            terms_for_item = [term]
        elif item.billing_mode == FeeStructure.BILLING_MODE_ONCE_YEAR and item.due_term:
            terms_for_item = [item.due_term]
        elif item.billing_mode == FeeStructure.BILLING_MODE_SELECTED_TERMS:
            selected_terms = item.applied_terms or []
            terms_for_item = [t for t in term_order if t in selected_terms]
            extras = [t for t in selected_terms if t not in terms_for_item]
            terms_for_item.extend(extras)

        class_names = sorted({c.name for c in item.applicable_classes.all()})
        classes_label = ', '.join(class_names) if class_names else '-'

        for row_term in terms_for_item:
            if row_term not in term_votehead_groups:
                term_votehead_groups[row_term] = []
            term_votehead_groups[row_term].append({
                'vote_head': item.vote_head,
                'classes': classes_label,
                'amount': amount,
            })

        for class_name in class_names:
            if class_name not in class_groups_map:
                class_groups_map[class_name] = {
                    'class_name': class_name,
                    'term_groups_map': OrderedDict(),
                    'class_total': Decimal('0'),
                }
            for row_term in terms_for_item:
                term_groups_map = class_groups_map[class_name]['term_groups_map']
                if row_term not in term_groups_map:
                    term_groups_map[row_term] = {
                        'term': row_term,
                        'rows': [],
                        'term_total': Decimal('0'),
                    }
                term_groups_map[row_term]['rows'].append({
                    'vote_head': item.vote_head,
                    'amount': amount,
                })
                term_groups_map[row_term]['term_total'] += amount
                class_groups_map[class_name]['class_total'] += amount

    print_term_groups = []
    for term_key, rows in term_votehead_groups.items():
        if not rows:
            continue
        print_term_groups.append({
            'term': term_key,
            'rows': rows,
            'term_rowspan': len(rows),
        })

    print_class_groups = []
    for class_data in class_groups_map.values():
        term_groups = []
        for term_key in term_order:
            term_data = class_data['term_groups_map'].get(term_key)
            if term_data and term_data['rows']:
                term_groups.append({
                    'term': term_data['term'],
                    'rows': term_data['rows'],
                    'term_total': term_data['term_total'],
                    'term_rowspan': len(term_data['rows']) + 1,
                })
        for term_key, term_data in class_data['term_groups_map'].items():
            if term_key not in term_order and term_data['rows']:
                term_groups.append({
                    'term': term_data['term'],
                    'rows': term_data['rows'],
                    'term_total': term_data['term_total'],
                    'term_rowspan': len(term_data['rows']) + 1,
                })

        if not term_groups:
            continue

        class_rowspan = sum(tg['term_rowspan'] for tg in term_groups) + 1
        render_rows = []
        first_class_row = True
        for term_group in term_groups:
            first_term_row = True
            for detail_row in term_group['rows']:
                render_rows.append({
                    'row_type': 'detail',
                    'show_class': first_class_row,
                    'class_name': class_data['class_name'],
                    'class_rowspan': class_rowspan,
                    'show_term': first_term_row,
                    'term': term_group['term'],
                    'term_rowspan': term_group['term_rowspan'],
                    'vote_head': detail_row['vote_head'],
                    'amount': detail_row['amount'],
                })
                first_class_row = False
                first_term_row = False
            render_rows.append({
                'row_type': 'term_total',
                'vote_head': f"Total {term_group['term']}",
                'amount': term_group['term_total'],
            })
        render_rows.append({
            'row_type': 'class_total',
            'class_name': class_data['class_name'],
            'amount': class_data['class_total'],
        })

        print_class_groups.append({
            'class_name': class_data['class_name'],
            'rows': render_rows,
            'class_total': class_data['class_total'],
        })
    votehead_options = list(structures_qs.values_list('vote_head', flat=True).distinct())
    votehead_records = []
    for item in structures_qs:
        votehead_records.append({
            'id': item.id,
            'vote_head': item.vote_head,
            'year': item.year,
            'billing_mode': item.billing_mode,
            'due_term': item.due_term,
            'applied_terms': item.applied_terms or [],
            'class_ids': [c.id for c in item.applicable_classes.all()],
            'amount': str(item.amount),
        })

    context = {
        'school': school,
        'structures': structures,
        'classes': classes,
        'selected_class_ids': [str(c) for c in class_ids],
        'selected_term': term,
        'selected_year': year,
        'term_choices': FeeStructure.TERM_CHOICES,
        'billing_mode_choices': FeeStructure.BILLING_MODE_CHOICES,
        'default_year': date.today().year,
        'votehead_options': votehead_options,
        'votehead_records': votehead_records,
        'print_single_term': bool(term),
        'print_term_label': term,
        'print_term_groups': print_term_groups,
        'print_class_groups': print_class_groups,
        'print_generated_at': timezone.localtime(timezone.now()).strftime('%d %b %Y, %I:%M %p'),
        'print_prepared_by': (request.user.get_full_name() or request.user.username),
    }
    return render(request, 'finance/fee_structure.html', context)

@login_required
def balance_report(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can view balance reports.')

    term = request.GET.get('term', 'Term 1')
    try:
        year = int(request.GET.get('year', date.today().year))
    except (TypeError, ValueError):
        year = date.today().year
    class_id = request.GET.get('classroom')
    period = (request.GET.get('period') or 'term').strip().lower()
    anchor_date_raw = (request.GET.get('anchor_date') or '').strip()
    anchor_date = None
    if anchor_date_raw:
        try:
            anchor_date = date.fromisoformat(anchor_date_raw)
        except ValueError:
            anchor_date = None
    period_start, period_end, period_label, period_warning = _financial_period_range(
        school=school,
        period=period,
        term=term,
        year=year,
        anchor_date=anchor_date,
    )

    students = Student.objects.select_related('school', 'classroom').filter(school=school)
    classes_qs = ClassRoom.objects.filter(school=school).order_by('name')
    class_label = 'All Classes'
    if class_id:
        class_obj = classes_qs.filter(id=class_id).first()
        if class_obj:
            class_label = class_obj.name

    if class_id:
        students = students.filter(classroom_id=class_id)

    report_data = []
    for student in students:
        meta = _student_outstanding_meta(student, term, year)
        report_data.append({
            'student': student,
            'admission_number': student.admission_number or '-',
            'school': student.school.name,
            'classroom': student.classroom.name if student.classroom else 'N/A',
            'due': meta['due_total'],
            'paid': meta['paid_total'],
            'balance': meta['balance'],
        })

    income_payments = FeePayment.objects.filter(
        student__school=school,
        date_paid__gte=period_start,
        date_paid__lte=period_end,
    )
    income_total = income_payments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    income_by_mode = list(
        income_payments.values('payment_method')
        .annotate(total=Sum('amount_paid'))
        .order_by('-total', 'payment_method')
    )

    income_allocations = FeePaymentAllocation.objects.filter(
        fee_payment__student__school=school,
        fee_payment__date_paid__gte=period_start,
        fee_payment__date_paid__lte=period_end,
    )
    income_by_source = list(
        income_allocations.values('vote_head')
        .annotate(total=Sum('amount'))
        .order_by('-total', 'vote_head')
    )
    unallocated_income = (
        income_payments
        .exclude(allocations__isnull=False)
        .aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    )
    if unallocated_income:
        income_by_source.append({'vote_head': 'Unallocated', 'total': unallocated_income})

    line_total_expr = ExpressionWrapper(
        F('amount') * F('quantity'),
        output_field=DecimalField(max_digits=16, decimal_places=2),
    )
    expenditures_qs = Expenditure.objects.filter(
        school=school,
        date__gte=period_start,
        date__lte=period_end,
    )
    expenditure_total = expenditures_qs.aggregate(total=Sum(line_total_expr))['total'] or Decimal('0')
    expenditure_by_purpose = list(
        expenditures_qs.values('vote_head')
        .annotate(total=Sum(line_total_expr))
        .order_by('-total', 'vote_head')
    )
    expenditure_by_mode = list(
        expenditures_qs.values('payment_method')
        .annotate(total=Sum(line_total_expr))
        .order_by('-total', 'payment_method')
    )
    net_cashflow = income_total - expenditure_total

    if 'export_excel' in request.GET:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Balance Report'

        # Header block (aligned with print layout)
        ws.merge_cells('A1:G1')
        ws['A1'] = school.name
        ws['A1'].font = Font(size=15, bold=True, color='0F3057')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:G2')
        details = ' | '.join([v for v in [school.address, school.phone, school.email] if v])
        ws['A2'] = details or ' '
        ws['A2'].font = Font(size=10, color='4B5563')
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:G3')
        ws['A3'] = f'Balance Report (Cumulative) - Class: {class_label} | Term: {term} | Year: {year}'
        ws['A3'].font = Font(size=11, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:G4')
        ws['A4'] = f'Prepared by: {request.user.get_full_name() or request.user.username}'
        ws['A4'].font = Font(size=10, color='4B5563')
        ws['A4'].alignment = Alignment(horizontal='center')

        header_row = 6
        ws.append([])
        ws.append(['Admission No', 'Student', 'School', 'Class', 'Fees Due', 'Paid', 'Balance'])
        header_fill = PatternFill(fill_type='solid', start_color='F3F4F6', end_color='F3F4F6')
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in 'ABCDEFG':
            cell = ws[f'{col}{header_row}']
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        data_row = header_row + 1
        for row in report_data:
            ws.append([
                row['admission_number'],
                f"{row['student'].first_name} {row['student'].last_name}",
                row['school'], row['classroom'],
                row['due'], row['paid'], row['balance']
            ])
            for col in 'ABCDEFG':
                ws[f'{col}{data_row}'].border = border
            ws[f'E{data_row}'].alignment = Alignment(horizontal='right')
            ws[f'F{data_row}'].alignment = Alignment(horizontal='right')
            ws[f'G{data_row}'].alignment = Alignment(horizontal='right')

            bal = row['balance']
            if bal > 0:
                ws[f'G{data_row}'].font = Font(color='DC2626', bold=True)  # outstanding
            elif bal < 0:
                ws[f'G{data_row}'].font = Font(color='16A34A', bold=True)  # overpaid
            else:
                ws[f'G{data_row}'].font = Font(color='2563EB', bold=True)  # cleared
            data_row += 1

        widths = {'A': 16, 'B': 28, 'C': 24, 'D': 16, 'E': 13, 'F': 13, 'G': 13}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Footer + pagination
        ws.oddFooter.center.text = 'SKUL PLUS SCHOOL MANAGEMENT SYSTEM'
        ws.oddFooter.right.text = 'Page &P of &N'
        ws.evenFooter.center.text = 'SKUL PLUS SCHOOL MANAGEMENT SYSTEM'
        ws.evenFooter.right.text = 'Page &P of &N'

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="balance_report.xlsx"'
        wb.save(response)
        return response

    if 'export_pdf' in request.GET or 'print_pdf' in request.GET:
        response = HttpResponse(content_type='application/pdf')
        disposition = 'inline' if 'print_pdf' in request.GET else 'attachment'
        response['Content-Disposition'] = f'{disposition}; filename="balance_report.pdf"'
        if 'print_pdf' in request.GET:
            response['X-Frame-Options'] = 'SAMEORIGIN'
        p = canvas.Canvas(response, pagesize=landscape(A4))
        page_w, page_h = landscape(A4)
        margin = 32
        row_h = 18
        footer_text = 'SKUL PLUS SCHOOL MANAGEMENT SYSTEM'
        page_no = 1

        def draw_header(y_top):
            # Logo
            if school.logo:
                try:
                    p.drawImage(
                        ImageReader(school.logo.path),
                        margin,
                        y_top - 48,
                        width=44,
                        height=44,
                        preserveAspectRatio=True,
                        mask='auto',
                    )
                except Exception:
                    pass
            p.setFillColor(colors.HexColor('#0f3057'))
            p.setFont('Helvetica-Bold', 14)
            p.drawString(margin + 52, y_top - 10, school.name)
            p.setFillColor(colors.HexColor('#4B5563'))
            p.setFont('Helvetica', 9)
            details = ' | '.join([v for v in [school.address, school.phone, school.email] if v])
            if details:
                p.drawString(margin + 52, y_top - 24, details[:130])
            p.setFillColor(colors.black)
            p.setFont('Helvetica-Bold', 10)
            p.drawString(margin + 52, y_top - 38, f'Balance Report (Cumulative) - Class: {class_label} | Term: {term} | Year: {year}')
            p.setFont('Helvetica', 9)
            p.drawString(margin + 52, y_top - 50, f'Prepared by: {request.user.get_full_name() or request.user.username}')
            p.setStrokeColor(colors.HexColor('#0f3057'))
            p.setLineWidth(1.2)
            p.line(margin, y_top - 58, page_w - margin, y_top - 58)
            return y_top - 70

        def draw_table_header(y_head):
            p.setFillColor(colors.HexColor('#f3f4f6'))
            p.rect(margin, y_head - row_h, page_w - (margin * 2), row_h, fill=1, stroke=1)
            p.setFillColor(colors.black)
            p.setFont('Helvetica-Bold', 8)
            x = [margin + 4, margin + 72, margin + 184, margin + 296, margin + 374, margin + 445, margin + 513]
            labels = ['Adm No', 'Student', 'School', 'Class', 'Fees Due', 'Paid', 'Balance']
            for idx, label in enumerate(labels):
                p.drawString(x[idx], y_head - 12, label)
            return y_head - row_h

        def draw_footer():
            p.setStrokeColor(colors.HexColor('#d1d5db'))
            p.setLineWidth(0.8)
            p.line(margin, 28, page_w - margin, 28)
            p.setFont('Helvetica', 9)
            p.setFillColor(colors.HexColor('#374151'))
            p.drawString(margin, 16, footer_text)
            p.drawRightString(page_w - margin, 16, f'Page {page_no}')

        y = draw_header(page_h - margin)
        y = draw_table_header(y)

        for row in report_data:
            if y < 70:
                draw_footer()
                p.showPage()
                page_no += 1
                y = draw_header(page_h - margin)
                y = draw_table_header(y)

            p.setStrokeColor(colors.HexColor('#e5e7eb'))
            p.rect(margin, y - row_h, page_w - (margin * 2), row_h, fill=0, stroke=1)
            p.setFont('Helvetica', 7.8)
            p.setFillColor(colors.black)
            p.drawString(margin + 4, y - 12, str(row['admission_number'])[:10])
            p.drawString(margin + 72, y - 12, f"{row['student'].first_name} {row['student'].last_name}"[:20])
            p.drawString(margin + 184, y - 12, str(row['school'])[:20])
            p.drawString(margin + 296, y - 12, str(row['classroom'])[:12])
            p.drawRightString(margin + 438, y - 12, str(row['due']))
            p.drawRightString(margin + 506, y - 12, str(row['paid']))

            if row['balance'] > 0:
                p.setFillColor(colors.HexColor('#DC2626'))  # outstanding
            elif row['balance'] < 0:
                p.setFillColor(colors.HexColor('#16A34A'))  # overpaid
            else:
                p.setFillColor(colors.HexColor('#2563EB'))  # cleared
            p.setFont('Helvetica-Bold', 7.8)
            p.drawRightString(page_w - margin - 6, y - 12, str(row['balance']))
            y -= row_h

        draw_footer()
        p.save()
        return response

    context = {
        'report_data': report_data,
        'term': term,
        'year': year,
        'classes': classes_qs,
        'school': school,
        'period': period,
        'anchor_date': anchor_date_raw,
        'period_start': period_start,
        'period_end': period_end,
        'period_label': period_label,
        'period_warning': period_warning,
        'income_total': income_total,
        'income_by_source': income_by_source,
        'income_by_mode': income_by_mode,
        'expenditure_total': expenditure_total,
        'expenditure_by_purpose': expenditure_by_purpose,
        'expenditure_by_mode': expenditure_by_mode,
        'net_cashflow': net_cashflow,
    }
    return render(request, 'finance/balance_report.html', context)


@login_required
def model_reports(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can view model reports.')

    term = request.GET.get('term', 'Term 1')
    try:
        year = int(request.GET.get('year', date.today().year))
    except (TypeError, ValueError):
        year = date.today().year

    period = (request.GET.get('period') or 'term').strip().lower()
    anchor_date_raw = (request.GET.get('anchor_date') or '').strip()
    anchor_date = None
    if anchor_date_raw:
        try:
            anchor_date = date.fromisoformat(anchor_date_raw)
        except ValueError:
            anchor_date = None

    period_start, period_end, period_label, period_warning = _financial_period_range(
        school=school,
        period=period,
        term=term,
        year=year,
        anchor_date=anchor_date,
    )

    income_payments = FeePayment.objects.filter(
        student__school=school,
        date_paid__gte=period_start,
        date_paid__lte=period_end,
    )
    income_total = income_payments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    income_by_mode = list(
        income_payments.values('payment_method')
        .annotate(total=Sum('amount_paid'))
        .order_by('-total', 'payment_method')
    )

    income_allocations = FeePaymentAllocation.objects.filter(
        fee_payment__student__school=school,
        fee_payment__date_paid__gte=period_start,
        fee_payment__date_paid__lte=period_end,
    )
    income_by_source = list(
        income_allocations.values('vote_head')
        .annotate(total=Sum('amount'))
        .order_by('-total', 'vote_head')
    )
    unallocated_income = (
        income_payments
        .exclude(allocations__isnull=False)
        .aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    )
    if unallocated_income:
        income_by_source.append({'vote_head': 'Unallocated', 'total': unallocated_income})

    line_total_expr = ExpressionWrapper(
        F('amount') * F('quantity'),
        output_field=DecimalField(max_digits=16, decimal_places=2),
    )
    expenditures_qs = Expenditure.objects.filter(
        school=school,
        date__gte=period_start,
        date__lte=period_end,
    )
    expenditure_total = expenditures_qs.aggregate(total=Sum(line_total_expr))['total'] or Decimal('0')
    expenditure_by_purpose = list(
        expenditures_qs.values('vote_head')
        .annotate(total=Sum(line_total_expr))
        .order_by('-total', 'vote_head')
    )
    expenditure_by_mode = list(
        expenditures_qs.values('payment_method')
        .annotate(total=Sum(line_total_expr))
        .order_by('-total', 'payment_method')
    )
    net_cashflow = income_total - expenditure_total

    if 'export_pdf' in request.GET or 'print_pdf' in request.GET:
        response = HttpResponse(content_type='application/pdf')
        disposition = 'inline' if 'print_pdf' in request.GET else 'attachment'
        response['Content-Disposition'] = f'{disposition}; filename="model_reports.pdf"'
        if 'print_pdf' in request.GET:
            response['X-Frame-Options'] = 'SAMEORIGIN'

        p = canvas.Canvas(response, pagesize=A4)
        page_w, page_h = A4
        margin = 34
        row_h = 16
        page_no = 1

        def draw_footer():
            p.setStrokeColor(colors.HexColor('#d1d5db'))
            p.setLineWidth(0.8)
            p.line(margin, 28, page_w - margin, 28)
            p.setFont('Helvetica', 8.5)
            p.setFillColor(colors.HexColor('#374151'))
            p.drawString(margin, 16, 'SKUL PLUS SCHOOL MANAGEMENT SYSTEM')
            p.drawRightString(page_w - margin, 16, f'Page {page_no}')

        def draw_header():
            y_top = page_h - margin
            if school.logo:
                try:
                    p.drawImage(
                        ImageReader(school.logo.path),
                        margin,
                        y_top - 42,
                        width=36,
                        height=36,
                        preserveAspectRatio=True,
                        mask='auto',
                    )
                except Exception:
                    pass
            p.setFillColor(colors.HexColor('#0f3057'))
            p.setFont('Helvetica-Bold', 13)
            p.drawString(margin + 44, y_top - 10, f'{school.name} - Income vs Expenditure')
            p.setFillColor(colors.HexColor('#4b5563'))
            p.setFont('Helvetica', 8.5)
            p.drawString(margin + 44, y_top - 24, f'Period: {period_label}')
            p.drawString(margin + 44, y_top - 36, f'Range: {period_start.isoformat()} to {period_end.isoformat()}')
            p.setStrokeColor(colors.HexColor('#0f3057'))
            p.setLineWidth(1)
            p.line(margin, y_top - 48, page_w - margin, y_top - 48)
            return y_top - 60

        def ensure_space(y, needed):
            nonlocal page_no
            if y - needed < 44:
                draw_footer()
                p.showPage()
                page_no += 1
                return draw_header()
            return y

        def draw_ledger_table(y_top, title, left_rows, left_key, right_rows, right_key):
            table_w = page_w - (margin * 2)
            col_part = table_w * 0.38
            col_amt = table_w * 0.12
            x0 = margin
            x1 = x0 + col_part
            x2 = x1 + col_amt
            x3 = x2 + col_part
            x4 = x3 + col_amt

            y = ensure_space(y_top, row_h * 4)
            p.setFont('Helvetica-Bold', 10)
            p.setFillColor(colors.HexColor('#0f3057'))
            p.drawString(margin, y, title)
            y -= 10

            y = ensure_space(y, row_h)
            p.setFillColor(colors.HexColor('#f3f4f6'))
            p.rect(x0, y - row_h + 2, table_w, row_h, fill=1, stroke=0)
            p.setFillColor(colors.black)
            p.setFont('Helvetica-Bold', 8.5)
            p.drawString(x0 + 4, y - 9, 'Income Particulars')
            p.drawRightString(x2 - 4, y - 9, 'Amount')
            p.drawString(x2 + 4, y - 9, 'Expenditure Particulars')
            p.drawRightString(x4 - 4, y - 9, 'Amount')
            y -= row_h

            max_rows = max(len(left_rows), len(right_rows), 1)
            for idx in range(max_rows):
                y = ensure_space(y, row_h)
                left = left_rows[idx] if idx < len(left_rows) else None
                right = right_rows[idx] if idx < len(right_rows) else None

                p.setFont('Helvetica', 8.2)
                p.setFillColor(colors.black)
                if left:
                    l_label = str(left.get(left_key) or '-')[:56]
                    l_total = left.get('total') or Decimal('0')
                    p.drawString(x0 + 4, y - 9, l_label)
                    p.drawRightString(x2 - 4, y - 9, f'{l_total:.2f}')
                if right:
                    r_label = str(right.get(right_key) or '-')[:56]
                    r_total = right.get('total') or Decimal('0')
                    p.drawString(x2 + 4, y - 9, r_label)
                    p.drawRightString(x4 - 4, y - 9, f'{r_total:.2f}')

                p.setStrokeColor(colors.HexColor('#d1d5db'))
                p.setLineWidth(0.6)
                p.line(x0, y - row_h + 2, x4, y - row_h + 2)
                p.line(x1, y + 2, x1, y - row_h + 2)
                p.line(x2, y + 2, x2, y - row_h + 2)
                p.line(x3, y + 2, x3, y - row_h + 2)
                y -= row_h

            # totals row
            y = ensure_space(y, row_h)
            p.setFillColor(colors.HexColor('#f9fafb'))
            p.rect(x0, y - row_h + 2, table_w, row_h, fill=1, stroke=0)
            p.setFillColor(colors.black)
            p.setFont('Helvetica-Bold', 8.5)
            left_total = sum((r.get('total') or Decimal('0') for r in left_rows), Decimal('0'))
            right_total = sum((r.get('total') or Decimal('0') for r in right_rows), Decimal('0'))
            p.drawString(x0 + 4, y - 9, 'Total')
            p.drawRightString(x2 - 4, y - 9, f'{left_total:.2f}')
            p.drawString(x2 + 4, y - 9, 'Total')
            p.drawRightString(x4 - 4, y - 9, f'{right_total:.2f}')
            p.setStrokeColor(colors.HexColor('#9ca3af'))
            p.rect(x0, y - row_h + 2, table_w, row_h, fill=0, stroke=1)
            p.line(x1, y + 2, x1, y - row_h + 2)
            p.line(x2, y + 2, x2, y - row_h + 2)
            p.line(x3, y + 2, x3, y - row_h + 2)
            return y - row_h - 8

        y = draw_header()
        y = ensure_space(y, row_h * 3)
        p.setFont('Helvetica-Bold', 9)
        p.setFillColor(colors.black)
        p.drawString(margin, y, 'Ledger Summary')
        p.drawString(margin + 120, y, f'Income: {income_total:.2f}')
        p.drawString(margin + 280, y, f'Expenditure: {expenditure_total:.2f}')
        p.drawString(margin + 470, y, f'Net: {net_cashflow:.2f}')
        y -= 16

        y = draw_ledger_table(y, 'Ledger A: Source vs Purpose', income_by_source, 'vote_head', expenditure_by_purpose, 'vote_head')
        y = draw_ledger_table(y, 'Ledger B: Payment Mode vs Payment Mode', income_by_mode, 'payment_method', expenditure_by_mode, 'payment_method')
        draw_footer()
        p.save()
        return response

    context = {
        'school': school,
        'term': term,
        'year': year,
        'period': period,
        'anchor_date': anchor_date_raw,
        'period_start': period_start,
        'period_end': period_end,
        'period_label': period_label,
        'period_warning': period_warning,
        'income_total': income_total,
        'income_by_source': income_by_source,
        'income_by_mode': income_by_mode,
        'expenditure_total': expenditure_total,
        'expenditure_by_purpose': expenditure_by_purpose,
        'expenditure_by_mode': expenditure_by_mode,
        'net_cashflow': net_cashflow,
    }
    return render(request, 'finance/model_reports.html', context)


@login_required
def expenditure_report(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can view expenditure.')

    base_expenditures = Expenditure.objects.filter(school=school)
    expenditures = base_expenditures

    if request.method == 'POST':
        form = ExpenditureForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.school = school
            obj.source = Expenditure.SOURCE_MANUAL
            obj.save()
            messages.success(request, 'Expenditure recorded successfully.')
            return redirect('expenditure')
        messages.error(request, 'Please correct expenditure form errors.')
    else:
        form = ExpenditureForm(initial={'date': date.today(), 'quantity': 1})

    filters = {
        'date': (request.GET.get('date') or '').strip(),
        'vote_head': (request.GET.get('vote_head') or '').strip(),
        'source': (request.GET.get('source') or '').strip(),
        'payment_method': (request.GET.get('payment_method') or '').strip(),
    }

    if filters['date']:
        expenditures = expenditures.filter(date=filters['date'])
    if filters['vote_head']:
        expenditures = expenditures.filter(vote_head=filters['vote_head'])
    if filters['source']:
        expenditures = expenditures.filter(source=filters['source'])
    if filters['payment_method']:
        expenditures = expenditures.filter(payment_method=filters['payment_method'])

    expenditures = expenditures.order_by('-date', '-id')
    total_amount = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0')

    if 'export_pdf' in request.GET or 'print_pdf' in request.GET:
        report_rows = list(expenditures)
        response = HttpResponse(content_type='application/pdf')
        disposition = 'inline' if 'print_pdf' in request.GET else 'attachment'
        response['Content-Disposition'] = f'{disposition}; filename="expenditure_report.pdf"'
        if 'print_pdf' in request.GET:
            response['X-Frame-Options'] = 'SAMEORIGIN'

        p = canvas.Canvas(response, pagesize=landscape(A4))
        page_w, page_h = landscape(A4)
        margin = 28
        row_h = 18
        prepared_by = request.user.get_full_name() or request.user.username
        generated_at = timezone.localtime(timezone.now()).strftime('%d %b %Y, %I:%M %p')
        page_no = 1

        def draw_header(y_top):
            if school.logo:
                try:
                    p.drawImage(
                        ImageReader(school.logo.path),
                        margin,
                        y_top - 48,
                        width=44,
                        height=44,
                        preserveAspectRatio=True,
                        mask='auto',
                    )
                except Exception:
                    pass

            p.setFillColor(colors.HexColor('#0f3057'))
            p.setFont('Helvetica-Bold', 14)
            p.drawString(margin + 52, y_top - 12, school.name)

            p.setFillColor(colors.HexColor('#4B5563'))
            p.setFont('Helvetica', 9)
            details = ' | '.join([v for v in [school.address, school.phone, school.email] if v])
            if details:
                p.drawString(margin + 52, y_top - 26, details[:130])

            p.setFillColor(colors.black)
            p.setFont('Helvetica-Bold', 10)
            p.drawString(margin + 52, y_top - 40, 'Expenditure Report')
            p.setFont('Helvetica', 8.5)
            summary = f"Filters: Date={filters['date'] or 'All'}, Votehead={filters['vote_head'] or 'All'}, Source={filters['source'] or 'All'}, Payment={filters['payment_method'] or 'All'}"
            p.drawString(margin + 52, y_top - 52, summary[:150])

            p.setStrokeColor(colors.HexColor('#0f3057'))
            p.setLineWidth(1.1)
            p.line(margin, y_top - 60, page_w - margin, y_top - 60)
            return y_top - 72

        def draw_table_header(y_head):
            col_widths = [58, 150, 95, 86, 64, 44, 68, 66, 98, 56]
            col_x = [margin]
            for width in col_widths[:-1]:
                col_x.append(col_x[-1] + width)

            p.setFillColor(colors.HexColor('#f3f4f6'))
            p.rect(margin, y_head - row_h, sum(col_widths), row_h, fill=1, stroke=1)
            p.setFillColor(colors.black)
            p.setFont('Helvetica-Bold', 7.7)
            headers = ['Date', 'Item', 'Votehead', 'Receipt/Inv', 'Amount', 'Qty', 'Total', 'Source', 'Payment', 'Evidence']
            for idx, label in enumerate(headers):
                p.drawString(col_x[idx] + 4, y_head - 12, label)
            return y_head - row_h, col_x, col_widths

        def draw_footer():
            footer_line_y = 74
            p.setStrokeColor(colors.HexColor('#d1d5db'))
            p.setLineWidth(0.8)
            p.line(margin, footer_line_y, page_w - margin, footer_line_y)

            p.setFont('Helvetica', 8.5)
            p.setFillColor(colors.HexColor('#374151'))
            p.drawString(margin, 62, f'Prepared by: {prepared_by}')
            p.drawString(margin, 50, f'Generated: {generated_at}')
            p.drawString(margin, 38, 'SKUL PLUS SCHOOL MANAGEMENT SYSTEM')
            p.drawRightString(page_w - margin, 38, f'Page {page_no}')

            stamp_x = page_w - margin - 128
            sig_x = page_w - margin - 62
            asset_y = 14
            box_w = 52
            box_h = 52

            p.setFont('Helvetica', 7)
            p.setFillColor(colors.HexColor('#6b7280'))
            p.drawString(stamp_x, 68, 'Stamp')
            p.drawString(sig_x, 68, 'Signature')

            if school.stamp:
                try:
                    p.drawImage(
                        ImageReader(school.stamp.path),
                        stamp_x,
                        asset_y,
                        width=box_w,
                        height=box_h,
                        preserveAspectRatio=True,
                        mask='auto',
                    )
                except Exception:
                    p.rect(stamp_x, asset_y, box_w, box_h, stroke=1, fill=0)
            else:
                p.rect(stamp_x, asset_y, box_w, box_h, stroke=1, fill=0)

            if school.head_signature:
                try:
                    p.drawImage(
                        ImageReader(school.head_signature.path),
                        sig_x,
                        asset_y,
                        width=box_w,
                        height=box_h,
                        preserveAspectRatio=True,
                        mask='auto',
                    )
                except Exception:
                    p.rect(sig_x, asset_y, box_w, box_h, stroke=1, fill=0)
            else:
                p.rect(sig_x, asset_y, box_w, box_h, stroke=1, fill=0)

        y = draw_header(page_h - margin)
        y, col_x, col_widths = draw_table_header(y)
        table_w = sum(col_widths)
        total_value_sum = Decimal('0')

        for row in report_rows:
            if y < 98:
                draw_footer()
                p.showPage()
                page_no += 1
                y = draw_header(page_h - margin)
                y, col_x, col_widths = draw_table_header(y)

            p.setStrokeColor(colors.HexColor('#e5e7eb'))
            p.rect(margin, y - row_h, table_w, row_h, fill=0, stroke=1)
            p.setFont('Helvetica', 7.5)
            p.setFillColor(colors.black)

            total_row = row.total_value or Decimal('0')
            total_value_sum += total_row
            receipt_val = row.receipt_invoice_no or '-'
            evidence_val = 'Yes' if row.evidence_document else 'No'

            p.drawString(col_x[0] + 4, y - 12, str(row.date))
            p.drawString(col_x[1] + 4, y - 12, str(row.item)[:34])
            p.drawString(col_x[2] + 4, y - 12, str(row.vote_head)[:20])
            p.drawString(col_x[3] + 4, y - 12, str(receipt_val)[:16])
            p.drawRightString(col_x[4] + col_widths[4] - 4, y - 12, str(row.amount))
            p.drawRightString(col_x[5] + col_widths[5] - 4, y - 12, str(row.quantity))
            p.drawRightString(col_x[6] + col_widths[6] - 4, y - 12, str(total_row))
            p.drawString(col_x[7] + 4, y - 12, str(row.get_source_display())[:12])
            p.drawString(col_x[8] + 4, y - 12, str(row.payment_method)[:16])
            p.drawString(col_x[9] + 4, y - 12, evidence_val)
            y -= row_h

        if y < 98:
            draw_footer()
            p.showPage()
            page_no += 1
            y = draw_header(page_h - margin)
            y, col_x, col_widths = draw_table_header(y)

        p.setFillColor(colors.HexColor('#f9fafb'))
        p.rect(margin, y - row_h, table_w, row_h, fill=1, stroke=1)
        p.setFillColor(colors.black)
        p.setFont('Helvetica-Bold', 8)
        p.drawString(margin + 4, y - 12, f'Total rows: {len(report_rows)}')
        p.drawRightString(col_x[4] + col_widths[4] - 4, y - 12, str(total_amount))
        p.drawRightString(col_x[6] + col_widths[6] - 4, y - 12, str(total_value_sum))

        draw_footer()
        p.save()
        return response

    total_value = total_amount

    date_choices = (
        base_expenditures.order_by('-date').values_list('date', flat=True).distinct()
    )
    vote_head_choices = (
        base_expenditures.exclude(vote_head='').order_by('vote_head').values_list('vote_head', flat=True).distinct()
    )
    source_choices = [
        choice for choice in Expenditure.SOURCE_CHOICES
        if base_expenditures.filter(source=choice[0]).exists()
    ]
    payment_method_choices = [
        choice for choice in Expenditure.PAYMENT_METHOD_CHOICES
        if base_expenditures.filter(payment_method=choice[0]).exists()
    ]

    context = {
        'form': form,
        'expenditures': expenditures[:100],
        'total_amount': total_amount,
        'total_value': total_value,
        'filters': filters,
        'date_choices': date_choices,
        'vote_head_choices': vote_head_choices,
        'source_choices': source_choices,
        'payment_method_choices': payment_method_choices,
    }
    return render(request, 'finance/expenditure.html', context)


@login_required
def expenditure_evidence(request, expenditure_id):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can access expenditure evidence.')

    row = get_object_or_404(Expenditure, id=expenditure_id, school=school)
    if not row.evidence_document:
        return HttpResponseForbidden('No evidence document attached.')

    file_name = os.path.basename(row.evidence_document.name or 'evidence')
    guessed_type = mimetypes.guess_type(file_name)[0] or 'application/octet-stream'
    mode = (request.GET.get('mode') or 'preview').strip().lower()
    disposition = 'attachment' if mode == 'download' else 'inline'

    file_obj = row.evidence_document.open('rb')
    response = FileResponse(file_obj, content_type=guessed_type)
    response['Content-Disposition'] = f'{disposition}; filename="{file_name}"'
    response['X-Content-Type-Options'] = 'nosniff'
    return response


@login_required
def fee_receipt(request, payment_id):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can generate receipts.')

    payment = get_object_or_404(
        FeePayment.objects.select_related('student', 'student__school', 'student__classroom').prefetch_related('allocations'),
        id=payment_id,
        student__school=school,
    )
    student = payment.student
    receipt_format = request.GET.get('format', '').strip().lower()
    current_meta = _student_outstanding_meta(student, payment.term, payment.year)
    balance_after = current_meta['balance']
    balance_before = balance_after + (payment.amount_paid or Decimal('0'))
    allocations = list(payment.allocations.all())

    if receipt_format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        disposition = 'inline' if request.GET.get('print_pdf') == '1' else 'attachment'
        response['Content-Disposition'] = f'{disposition}; filename="receipt_{payment.id}.pdf"'
        if request.GET.get('print_pdf') == '1':
            response['X-Frame-Options'] = 'SAMEORIGIN'

        p = canvas.Canvas(response, pagesize=A4)
        page_w, page_h = A4
        margin = 36
        y = page_h - margin
        content_w = page_w - (margin * 2)

        # Header
        logo_size = 54
        if student.school.logo:
            try:
                p.drawImage(ImageReader(student.school.logo.path), margin, y - logo_size, width=logo_size, height=logo_size, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        p.setFont('Helvetica-Bold', 16)
        p.setFillColor(colors.HexColor('#0f3057'))
        p.drawString(margin + 64, y - 10, student.school.name or 'School')
        p.setFont('Helvetica', 10)
        p.setFillColor(colors.HexColor('#374151'))
        school_line = ' | '.join([v for v in [student.school.address, student.school.phone, student.school.email] if v])
        if school_line:
            p.drawString(margin + 64, y - 24, school_line[:120])
        p.setFont('Helvetica-Bold', 11)
        p.drawRightString(page_w - margin, y - 8, f'Receipt No: {payment.id}')
        p.setFont('Helvetica', 10)
        p.drawRightString(page_w - margin, y - 22, f'Date: {payment.date_paid}')

        y -= 66
        p.setStrokeColor(colors.HexColor('#0f3057'))
        p.setLineWidth(1.5)
        p.line(margin, y, page_w - margin, y)
        y -= 14

        # Details cards
        box_h = 84
        gap = 10
        box_w = (content_w - gap) / 2
        p.setStrokeColor(colors.HexColor('#d1d5db'))
        p.setLineWidth(0.8)
        p.roundRect(margin, y - box_h, box_w, box_h, 6, stroke=1, fill=0)
        p.roundRect(margin + box_w + gap, y - box_h, box_w, box_h, 6, stroke=1, fill=0)

        p.setFont('Helvetica', 10)
        left_lines = [
            ('Student', f'{student.first_name} {student.last_name}'),
            ('Admission No', f'{student.admission_number or "-"}'),
            ('Class', f'{student.classroom.name if student.classroom else "-"}'),
        ]
        ly = y - 16
        for label, value in left_lines:
            p.setFillColor(colors.HexColor('#6b7280'))
            p.drawString(margin + 10, ly, label)
            p.setFillColor(colors.black)
            p.drawRightString(margin + box_w - 10, ly, str(value))
            ly -= 20

        right_lines = [
            ('Term', payment.term),
            ('Year', str(payment.year)),
            ('Method', payment.payment_method),
            ('Total Paid', str(payment.amount_paid)),
        ]
        ry = y - 16
        for label, value in right_lines:
            p.setFillColor(colors.HexColor('#6b7280'))
            p.drawString(margin + box_w + gap + 10, ry, label)
            p.setFillColor(colors.black)
            p.drawRightString(page_w - margin - 10, ry, str(value))
            ry -= 17

        y -= box_h + 16

        # Allocation table
        table_x = margin
        table_w = content_w
        col_year = table_w * 0.14
        col_term = table_w * 0.18
        col_vote = table_w * 0.43
        col_amt = table_w - (col_year + col_term + col_vote)
        row_h = 20
        p.setFillColor(colors.HexColor('#f3f4f6'))
        p.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=1)
        p.setFillColor(colors.black)
        p.setFont('Helvetica-Bold', 10)
        p.drawString(table_x + 8, y - 14, 'Year')
        p.drawString(table_x + col_year + 8, y - 14, 'Term')
        p.drawString(table_x + col_year + col_term + 8, y - 14, 'Vote Head')
        p.drawRightString(table_x + table_w - 8, y - 14, 'Amount')
        p.line(table_x + col_year, y, table_x + col_year, y - row_h)
        p.line(table_x + col_year + col_term, y, table_x + col_year + col_term, y - row_h)
        p.line(table_x + col_year + col_term + col_vote, y, table_x + col_year + col_term + col_vote, y - row_h)
        y -= row_h

        p.setFont('Helvetica', 10)
        if allocations:
            for allocation in allocations:
                p.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=0)
                p.drawString(table_x + 8, y - 14, str(allocation.allocation_year))
                p.drawString(table_x + col_year + 8, y - 14, str(allocation.allocation_term))
                p.drawString(table_x + col_year + col_term + 8, y - 14, str(allocation.vote_head))
                p.drawRightString(table_x + table_w - 8, y - 14, str(allocation.amount))
                p.line(table_x + col_year, y, table_x + col_year, y - row_h)
                p.line(table_x + col_year + col_term, y, table_x + col_year + col_term, y - row_h)
                p.line(table_x + col_year + col_term + col_vote, y, table_x + col_year + col_term + col_vote, y - row_h)
                y -= row_h
        else:
            p.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=0)
            p.drawString(table_x + 8, y - 14, 'No split rows available.')
            p.line(table_x + col_year, y, table_x + col_year, y - row_h)
            p.line(table_x + col_year + col_term, y, table_x + col_year + col_term, y - row_h)
            p.line(table_x + col_year + col_term + col_vote, y, table_x + col_year + col_term + col_vote, y - row_h)
            y -= row_h

        p.setFillColor(colors.HexColor('#f9fafb'))
        p.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=1)
        p.setFillColor(colors.black)
        p.setFont('Helvetica-Bold', 10)
        p.drawRightString(table_x + col_year + col_term + col_vote - 8, y - 14, 'Total')
        p.drawRightString(table_x + table_w - 8, y - 14, str(payment.amount_paid))
        p.line(table_x + col_year, y, table_x + col_year, y - row_h)
        p.line(table_x + col_year + col_term, y, table_x + col_year + col_term, y - row_h)
        p.line(table_x + col_year + col_term + col_vote, y, table_x + col_year + col_term + col_vote, y - row_h)
        y -= row_h + 16

        # Footer cards
        foot_h = 86
        p.roundRect(margin, y - foot_h, box_w, foot_h, 6, stroke=1, fill=0)
        p.roundRect(margin + box_w + gap, y - foot_h, box_w, foot_h, 6, stroke=1, fill=0)
        p.setFont('Helvetica', 10)
        p.setFillColor(colors.HexColor('#6b7280'))
        p.drawString(margin + 10, y - 20, 'Balance Before Payment')
        p.drawString(margin + 10, y - 44, 'Balance After Payment')
        p.setFillColor(colors.black)
        p.drawRightString(margin + box_w - 10, y - 20, str(balance_before))
        p.drawRightString(margin + box_w - 10, y - 44, str(balance_after))

        p.setFillColor(colors.HexColor('#6b7280'))
        p.drawString(margin + box_w + gap + 10, y - 20, f'Prepared by {request.user.get_full_name() or request.user.username}')
        p.drawString(margin + box_w + gap + 10, y - 36, 'Accounts Office')
        p.drawString(margin + box_w + gap + 10, y - 52, 'Thank you.')
        if student.school.stamp:
            try:
                p.drawImage(
                    ImageReader(student.school.stamp.path),
                    page_w - margin - 76,
                    y - 70,
                    width=64,
                    height=64,
                    preserveAspectRatio=True,
                    mask='auto'
                )
            except Exception:
                pass

        p.showPage()
        p.save()
        return response

    context = {
        'payment': payment,
        'student': student,
        'allocations': allocations,
        'balance_before': balance_before,
        'balance_after': balance_after,
    }
    return render(request, 'finance/fee_receipt.html', context)


@login_required
def add_payment_meta(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can access payment metadata.')

    class_id = request.GET.get('class_id', '').strip()
    term = request.GET.get('term', '').strip() or 'Term 1'
    year = request.GET.get('year', '').strip() or str(date.today().year)
    student_id = request.GET.get('student_id', '').strip()

    students_qs = Student.objects.filter(school=school).order_by('first_name', 'last_name')
    if class_id:
        students_qs = students_qs.filter(classroom_id=class_id)

    students_payload = [
        {
            'id': s.id,
            'name': f'{s.first_name} {s.last_name}',
            'admission_number': s.admission_number,
            'classroom': s.classroom.name if s.classroom else '',
        }
        for s in students_qs
    ]

    meta_payload = {
        'due_total': '0.00',
        'paid_total': '0.00',
        'balance': '0.00',
        'votehead_rows': [],
        'outstanding_rows': [],
        'target_options': [],
    }

    if student_id:
        student = Student.objects.filter(id=student_id, school=school).first()
        if student:
            meta = _student_outstanding_meta(student, term, year)
            meta_payload = {
                'due_total': str(meta['due_total']),
                'paid_total': str(meta['paid_total']),
                'balance': str(meta['balance']),
                'votehead_rows': [
                    {
                        'vote_head': row['vote_head'],
                        'due': str(row['due']),
                        'paid': str(row['paid']),
                        'balance': str((row['due'] or Decimal('0')) - (row['paid'] or Decimal('0'))),
                    }
                    for row in meta['votehead_rows']
                ],
                'outstanding_rows': [
                    {
                        'key': row['key'],
                        'label': row['label'],
                        'year': row['year'],
                        'term': row['term'],
                        'vote_head': row['vote_head'],
                        'due': str(row['due']),
                        'paid': str(row['paid']),
                        'balance': str(row['balance']),
                    }
                    for row in meta['outstanding_rows']
                ],
                'target_options': meta['target_options'],
            }

    return JsonResponse({
        'students': students_payload,
        'meta': meta_payload,
    })


@login_required
def add_payment(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can add payments.')
    classes = ClassRoom.objects.filter(school=school).order_by('name')
    students = Student.objects.filter(school=school).select_related('classroom').order_by('first_name', 'last_name')
    payment_method_choices = FeePayment._meta.get_field('payment_method').choices
    term_choices = FeeStructure.TERM_CHOICES

    selected_class_id = (request.POST.get('classroom') or request.GET.get('classroom') or '').strip()
    selected_student_id = (request.POST.get('student') or '').strip()
    selected_term = (request.POST.get('term') or request.GET.get('term') or 'Term 1').strip()
    selected_year = (request.POST.get('year') or request.GET.get('year') or str(date.today().year)).strip()
    selected_method = (request.POST.get('payment_method') or 'Cash').strip()
    selected_allocation_mode = (request.POST.get('allocation_mode') or 'OLDEST_FIRST').strip()
    if selected_allocation_mode not in ('OLDEST_FIRST', 'MANUAL'):
        selected_allocation_mode = 'OLDEST_FIRST'
    entered_total = (request.POST.get('amount_paid') or '').strip()

    if request.method == 'POST':
        student = Student.objects.filter(id=selected_student_id, school=school).first()
        if not student:
            messages.error(request, 'Select a valid student.')
            return redirect('add_payment')
        if selected_class_id and str(student.classroom_id or '') != selected_class_id:
            messages.error(request, 'Selected student does not belong to selected class.')
            return redirect('add_payment')

        if selected_term not in [t[0] for t in FeeStructure.TERM_CHOICES]:
            messages.error(request, 'Select a valid term.')
            return redirect('add_payment')

        try:
            year_int = int(selected_year)
        except ValueError:
            messages.error(request, 'Year must be a number.')
            return redirect('add_payment')

        try:
            amount_paid = Decimal(entered_total)
        except (InvalidOperation, TypeError):
            messages.error(request, 'Enter a valid total amount paid.')
            return redirect('add_payment')

        if amount_paid <= 0:
            messages.error(request, 'Total amount must be greater than zero.')
            return redirect('add_payment')

        method_values = [choice[0] for choice in payment_method_choices]
        if selected_method not in method_values:
            messages.error(request, 'Invalid payment method selected.')
            return redirect('add_payment')

        payment_meta = _student_outstanding_meta(student, selected_term, year_int)
        outstanding_rows = payment_meta['outstanding_rows']
        target_balance_map = payment_meta['target_balance_map']
        if amount_paid > (payment_meta['balance'] or Decimal('0')):
            messages.error(request, 'Amount paid cannot exceed outstanding balance up to selected term/year.')
            return redirect('add_payment')

        split_rows = []
        if selected_allocation_mode == 'OLDEST_FIRST':
            remaining = amount_paid
            for row in outstanding_rows:
                if remaining <= 0:
                    break
                move = min(remaining, row['balance'])
                if move <= 0:
                    continue
                split_rows.append({
                    'allocation_year': row['year'],
                    'allocation_term': row['term'],
                    'vote_head': row['vote_head'],
                    'amount': move,
                })
                remaining -= move
        else:
            split_targets = request.POST.getlist('split_target')
            split_amounts = request.POST.getlist('split_amount')
            for idx, target_key in enumerate(split_targets):
                target = (target_key or '').strip()
                amount_raw = (split_amounts[idx] if idx < len(split_amounts) else '').strip()
                if not target and not amount_raw:
                    continue
                if not target:
                    messages.error(request, f'Split row {idx + 1}: allocation target is required.')
                    return redirect('add_payment')
                if target not in target_balance_map:
                    messages.error(request, f'Split row {idx + 1}: invalid allocation target.')
                    return redirect('add_payment')
                try:
                    row_amount = Decimal(amount_raw)
                except (InvalidOperation, TypeError):
                    messages.error(request, f'Split row {idx + 1}: amount is invalid.')
                    return redirect('add_payment')
                if row_amount <= 0:
                    messages.error(request, f'Split row {idx + 1}: amount must be greater than zero.')
                    return redirect('add_payment')

                target_year, target_term, target_votehead = target.split('||', 2)
                split_rows.append({
                    'allocation_year': int(target_year),
                    'allocation_term': target_term,
                    'vote_head': target_votehead,
                    'amount': row_amount,
                    'target_key': target,
                })

            if not split_rows:
                messages.error(request, 'Add at least one allocation split row.')
                return redirect('add_payment')

            per_target_alloc = {}
            for row in split_rows:
                key = row['target_key']
                per_target_alloc[key] = per_target_alloc.get(key, Decimal('0')) + row['amount']
            for key, allocated_amount in per_target_alloc.items():
                if allocated_amount > target_balance_map.get(key, Decimal('0')):
                    messages.error(request, f'Allocated amount exceeds outstanding balance for {key}.')
                    return redirect('add_payment')

        if not split_rows:
            messages.error(request, 'No outstanding balances available for allocation.')
            return redirect('add_payment')

        split_total = sum((row['amount'] for row in split_rows), Decimal('0'))
        if split_total != amount_paid:
            messages.error(request, 'Sum of split allocations must match total amount paid.')
            return redirect('add_payment')

        with transaction.atomic():
            payment = FeePayment.objects.create(
                student=student,
                term=selected_term,
                year=year_int,
                amount_paid=amount_paid,
                payment_method=selected_method,
            )
            FeePaymentAllocation.objects.bulk_create([
                FeePaymentAllocation(
                    fee_payment=payment,
                    allocation_term=row['allocation_term'],
                    allocation_year=row['allocation_year'],
                    vote_head=row['vote_head'],
                    amount=row['amount'],
                )
                for row in split_rows
            ])
        messages.success(request, 'Payment saved successfully.')
        return redirect('fee_receipt', payment_id=payment.id)

    history = (
        FeePayment.objects
        .select_related('student', 'student__classroom')
        .prefetch_related('allocations')
        .filter(student__school=school)
        .order_by('-date_paid', '-id')[:40]
    )
    history_rows = []
    for payment in history:
        allocation_text = ', '.join(
            f'{a.allocation_year} {a.allocation_term} {a.vote_head}: {a.amount}'
            for a in payment.allocations.all()
        ) or '-'
        history_rows.append({
            'id': payment.id,
            'student_name': f'{payment.student.first_name} {payment.student.last_name}',
            'class_name': payment.student.classroom.name if payment.student.classroom else '-',
            'term': payment.term,
            'year': payment.year,
            'amount_paid': payment.amount_paid,
            'payment_method': payment.payment_method,
            'date_paid': payment.date_paid,
            'allocations': allocation_text,
        })

    context = {
        'classes': classes,
        'students': students,
        'term_choices': term_choices,
        'payment_method_choices': payment_method_choices,
        'selected_class_id': selected_class_id,
        'selected_student_id': selected_student_id,
        'selected_term': selected_term,
        'selected_year': selected_year,
        'selected_method': selected_method,
        'selected_allocation_mode': selected_allocation_mode,
        'entered_total': entered_total,
        'current_year': date.today().year,
        'history_rows': history_rows,
    }
    return render(request, 'finance/add_payment.html', context)


@login_required
def send_fee_reminders(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can send reminders.')

    term = 'Term 1'
    year = date.today().year
    students = Student.objects.filter(school=school)

    for student in students:
        balance = student.balance(term, year)
        if balance > 0 and student.parent_phone:
            send_fee_reminder(student.parent_phone, student.first_name, balance)

    return HttpResponse('SMS reminders sent!')









