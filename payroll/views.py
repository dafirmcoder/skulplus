from datetime import date
from decimal import Decimal, InvalidOperation
import io

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.db import transaction
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from schools.models import Teacher
from schools.access import get_user_school, has_full_headteacher_access

from .forms import PayrollRecordForm, StaffForm
from .models import PayrollAllowance, PayrollOtherDeduction, PayrollRecord, Staff


def _headteacher_school(request):
    school = get_user_school(request.user)
    if school and (request.user.is_superuser or has_full_headteacher_access(request.user, school)):
        return school
    return None


def _split_name(full_name):
    clean = (full_name or '').strip()
    if not clean:
        return ('', '')
    parts = clean.split()
    if len(parts) == 1:
        return (parts[0], '')
    return (parts[0], ' '.join(parts[1:]))


def _unique_username(base):
    User = get_user_model()
    candidate = base[:150] or 'staffuser'
    counter = 1
    while User.objects.filter(username=candidate).exists():
        suffix = f'{counter}'
        trim = max(1, 150 - len(suffix) - 1)
        candidate = f'{base[:trim]}.{suffix}'
        counter += 1
    return candidate


def _parse_other_deductions(request):
    names = request.POST.getlist('other_deduction_name[]') or request.POST.getlist('other_deduction_name')
    amounts = request.POST.getlist('other_deduction_amount[]') or request.POST.getlist('other_deduction_amount')

    entries = []
    max_rows = max(len(names), len(amounts))

    for idx in range(max_rows):
        name = (names[idx] if idx < len(names) else '').strip()
        amount_raw = (amounts[idx] if idx < len(amounts) else '').strip()

        if not name and not amount_raw:
            continue
        if not name:
            return None, f'Other deduction row {idx + 1} is missing a name.'

        try:
            amount = Decimal(amount_raw or '0')
        except (InvalidOperation, ValueError):
            return None, f'Other deduction amount is invalid on row {idx + 1}.'

        if amount < 0:
            return None, f'Other deduction amount cannot be negative on row {idx + 1}.'

        entries.append({'name': name, 'amount': amount})

    return entries, None


def _parse_allowances(request):
    names = request.POST.getlist('allowance_name[]') or request.POST.getlist('allowance_name')
    amounts = request.POST.getlist('allowance_amount[]') or request.POST.getlist('allowance_amount')

    entries = []
    max_rows = max(len(names), len(amounts))

    for idx in range(max_rows):
        name = (names[idx] if idx < len(names) else '').strip()
        amount_raw = (amounts[idx] if idx < len(amounts) else '').strip()

        if not name and not amount_raw:
            continue
        if not name:
            return None, f'Allowance row {idx + 1} is missing a name.'

        try:
            amount = Decimal(amount_raw or '0')
        except (InvalidOperation, ValueError):
            return None, f'Allowance amount is invalid on row {idx + 1}.'

        if amount < 0:
            return None, f'Allowance amount cannot be negative on row {idx + 1}.'

        entries.append({'name': name, 'amount': amount})

    return entries, None


def _safe_rate(amount, basic_salary):
    if not basic_salary:
        return '0'
    try:
        rate = (Decimal(amount or 0) * Decimal('100')) / Decimal(basic_salary)
    except Exception:
        return '0'
    rate_str = format(rate.quantize(Decimal('0.1')).normalize(), 'f')
    return rate_str if rate_str else '0'


def _payable_basic_salary(basic_salary, days_worked):
    try:
        return (Decimal(basic_salary or 0) * Decimal(days_worked or 0)) / Decimal('30')
    except Exception:
        return Decimal('0')


def _resolve_export_period(records_qs, request):
    month = (request.GET.get('month') or '').strip()
    year_raw = (request.GET.get('year') or '').strip()
    year = int(year_raw) if year_raw.isdigit() else None

    if month and year is not None:
        return month, year

    latest = records_qs.first()
    if latest:
        return latest.month, latest.year
    return '', None


def _build_payroll_period_pdf_bytes(school, month, year, records):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 45
    right = width - 45
    y = height - 42

    logo = _image_reader_from_field(school.logo)
    if logo:
        try:
            p.drawImage(logo, left, y - 44, width=48, height=48, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    center_x = width / 2
    p.setFont('Helvetica-Bold', 14)
    p.drawCentredString(center_x, y, (school.name or '').upper())
    if getattr(school, 'motto', ''):
        p.setFont('Helvetica-Oblique', 10)
        p.drawCentredString(center_x, y - 14, str(school.motto).upper()[:120])
    contact_line = ' | '.join([str(v).upper() for v in [school.phone, school.email] if v])
    if contact_line:
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 28, contact_line[:120])
        text_w = p.stringWidth(contact_line[:120], 'Helvetica', 10)
        p.line(center_x - text_w / 2, y - 30, center_x + text_w / 2, y - 30)
    if getattr(school, 'address', ''):
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 42, str(school.address).upper()[:120])
    p.setFont('Helvetica-Bold', 11)
    p.drawCentredString(center_x, y - 56, f'PAYROLL EVIDENCE - PERIOD: {month} {year}')
    p.setFont('Helvetica', 10)
    p.drawCentredString(center_x, y - 70, f'RECORDS: {len(records)}')
    y -= 84
    p.line(left, y, right, y)
    y -= 14

    p.setFont('Helvetica-Bold', 9)
    p.drawString(left, y, 'Staff')
    p.drawRightString(left + 310, y, 'Payable Basic')
    p.drawRightString(left + 420, y, 'Deductions')
    p.drawRightString(right, y, 'Net')
    y -= 10
    p.line(left, y, right, y)
    y -= 12

    total_net = Decimal('0')
    p.setFont('Helvetica', 9)
    for rec in records:
        if y < 60:
            p.showPage()
            y = height - 50
            p.setFont('Helvetica-Bold', 9)
            p.drawString(left, y, 'Staff')
            p.drawRightString(left + 310, y, 'Payable Basic')
            p.drawRightString(left + 420, y, 'Deductions')
            p.drawRightString(right, y, 'Net')
            y -= 10
            p.line(left, y, right, y)
            y -= 12
            p.setFont('Helvetica', 9)
        net = Decimal(rec.net_salary() or 0)
        total_net += net
        p.drawString(left, y, rec.staff.full_name[:36])
        p.drawRightString(left + 310, y, f'{Decimal(rec.payable_basic_salary() or 0):,.2f}')
        p.drawRightString(left + 420, y, f'{Decimal(rec.total_deductions() or 0):,.2f}')
        p.drawRightString(right, y, f'{net:,.2f}')
        y -= 12

    y -= 4
    p.line(left, y, right, y)
    y -= 14
    p.setFont('Helvetica-Bold', 10)
    p.drawString(left, y, 'TOTAL NET SALARY')
    p.drawRightString(right, y, f'{total_net:,.2f}')

    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer.getvalue()


@login_required
def staff_management(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can manage staff.')

    staff_qs = Staff.objects.filter(school=school).order_by('full_name')
    selected_staff = None
    selected_staff_id = request.GET.get('edit')
    if selected_staff_id:
        selected_staff = staff_qs.filter(id=selected_staff_id).first()

    if request.method == 'POST':
        action = request.POST.get('action', '').strip().lower()

        if action == 'add_staff':
            staff_form = StaffForm(request.POST)
            if staff_form.is_valid():
                with transaction.atomic():
                    staff = staff_form.save(commit=False)
                    staff.school = school
                    staff.save()

                    teacher_username = None
                    if staff.is_teacher:
                        User = get_user_model()
                        first_name, last_name = _split_name(staff.full_name)
                        base_stub = slugify(staff.full_name).replace('-', '.') or 'staff'
                        base_username = f'{base_stub}.{school.id}.{staff.id}'
                        username = _unique_username(base_username)
                        user = User.objects.create(
                            username=username,
                            first_name=first_name,
                            last_name=last_name,
                            email=staff.email or '',
                        )
                        user.set_unusable_password()
                        user.save(update_fields=['password'])
                        Teacher.objects.create(school=school, user=user, is_class_teacher=False)
                        teacher_username = username

                if teacher_username:
                    messages.success(
                        request,
                        f'Staff member added and linked as teacher (username: {teacher_username}). '
                        'Set password from User Management if login access is needed.'
                    )
                else:
                    messages.success(request, 'Staff member added successfully.')
                return redirect('staff_management')
            messages.error(request, 'Please correct staff form errors.')
        elif action == 'update_staff':
            staff_id = request.POST.get('staff_id')
            target_staff = staff_qs.filter(id=staff_id).first()
            if not target_staff:
                messages.error(request, 'Selected staff record was not found.')
                return redirect('staff_management')

            staff_form = StaffForm(request.POST, instance=target_staff)
            if staff_form.is_valid():
                staff_form.save()
                messages.success(request, 'Staff member updated successfully.')
                return redirect('staff_management')
            selected_staff = target_staff
            messages.error(request, 'Please correct staff form errors.')
        else:
            staff_form = StaffForm(instance=selected_staff)
            messages.error(request, 'Unknown action.')
    else:
        staff_form = StaffForm(instance=selected_staff)

    context = {
        'staff_form': staff_form,
        'staff_list': staff_qs,
        'staff_count': staff_qs.count(),
        'selected_staff': selected_staff,
    }
    return render(request, 'payroll/staff_management.html', context)


@login_required
@require_POST
def delete_staff(request, staff_id):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can manage staff.')

    staff = get_object_or_404(Staff, id=staff_id, school=school)
    staff.delete()
    messages.success(request, 'Staff member deleted successfully.')
    return redirect('staff_management')


@login_required
def payroll_overview(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can view payroll.')

    staff = Staff.objects.select_related('school').filter(school=school).order_by('full_name')
    records_base = PayrollRecord.objects.select_related('staff').filter(staff__school=school).order_by('-year', '-id')
    current_month = date.today().strftime('%B')
    current_year_raw = str(date.today().year)
    selected_record = None
    selected_record_id = request.GET.get('record')
    if selected_record_id:
        selected_record = records_base.filter(id=selected_record_id).first()

    filter_month = (request.GET.get('filter_month') or '').strip()
    filter_year_raw = (request.GET.get('filter_year') or '').strip()
    if not filter_month and not filter_year_raw and selected_record:
        filter_month = selected_record.month
        filter_year_raw = str(selected_record.year)
    if not filter_month:
        filter_month = current_month
    if not filter_year_raw:
        filter_year_raw = current_year_raw
    filter_year = int(filter_year_raw) if filter_year_raw.isdigit() else None
    records = records_base
    if filter_month:
        records = records.filter(month=filter_month)
    if filter_year is not None:
        records = records.filter(year=filter_year)
    period_paid = records.exists() and not records.filter(is_paid=False).exists()
    payroll_form = PayrollRecordForm(initial={'year': date.today().year})
    if selected_record_id:
        selected_record = records.filter(id=selected_record_id).first() or selected_record

    if request.method == 'POST':
        action = request.POST.get('action', '').strip().lower()
        if action == 'add_selected_staff':
            month = (request.POST.get('month') or '').strip()
            year_raw = (request.POST.get('year') or '').strip()
            selected_ids = request.POST.getlist('staff_ids[]') or request.POST.getlist('staff_ids')

            if not month:
                messages.error(request, 'Please select a month.')
            elif not year_raw.isdigit():
                messages.error(request, 'Please provide a valid year.')
            elif not selected_ids:
                messages.error(request, 'Please select at least one staff member.')
            else:
                year = int(year_raw)
                if PayrollRecord.objects.filter(staff__school=school, month=month, year=year, is_paid=True).exists():
                    messages.error(request, f'Payroll for {month} {year} is already paid and locked.')
                    return redirect('payroll_overview')
                selected_staff = list(staff.filter(id__in=selected_ids))
                created_count = 0
                skipped_count = 0
                with transaction.atomic():
                    for staff_member in selected_staff:
                        _, created = PayrollRecord.objects.get_or_create(
                            staff=staff_member,
                            month=month,
                            year=year,
                            defaults={
                                'days_worked': 30,
                                'allowances': Decimal('0'),
                                'paye_deduction': Decimal('0'),
                                'nssf_deduction': Decimal('0'),
                                'nhif_deduction': Decimal('0'),
                                'housing_levy_deduction': Decimal('0'),
                                'deductions': Decimal('0'),
                            }
                        )
                        if created:
                            created_count += 1
                        else:
                            skipped_count += 1

                if created_count:
                    messages.success(request, f'Added {created_count} staff member(s) to payroll for {month} {year}.')
                if skipped_count:
                    messages.warning(request, f'{skipped_count} selected staff already had payroll for {month} {year}.')
                return redirect('payroll_overview')

        elif action == 'mark_paid':
            month = (request.POST.get('filter_month') or filter_month).strip()
            year_raw = (request.POST.get('filter_year') or filter_year_raw).strip()
            if not month or not year_raw.isdigit():
                messages.error(request, 'Select a valid month and year to pay.')
                return redirect('payroll_overview')
            year = int(year_raw)
            period_qs = PayrollRecord.objects.filter(staff__school=school, month=month, year=year)
            if not period_qs.exists():
                messages.error(request, f'No payroll records found for {month} {year}.')
                return redirect('payroll_overview')
            updated = period_qs.filter(is_paid=False).update(is_paid=True, paid_at=timezone.now())
            # Auto-post payroll expense as a single row in expenditure.
            try:
                from finance.models import Expenditure
            except Exception:
                Expenditure = None
            if Expenditure is not None:
                existing_expense = Expenditure.objects.filter(
                    school=school,
                    source=Expenditure.SOURCE_PAYROLL,
                    payroll_month=month,
                    payroll_year=year,
                ).first()
                if not existing_expense:
                    payroll_records = list(period_qs.select_related('staff'))
                    total_net = sum((Decimal(rec.net_salary() or 0) for rec in payroll_records), Decimal('0'))
                    payroll_no = f'PAY-{year}-{month[:3].upper()}-{school.id}'
                    evidence_bytes = _build_payroll_period_pdf_bytes(school, month, year, payroll_records)
                    expense = Expenditure(
                        school=school,
                        date=timezone.localdate(),
                        item=f'{month} Salary',
                        amount=total_net,
                        quantity=1,
                        vote_head='Salary',
                        receipt_invoice_no=payroll_no,
                        source=Expenditure.SOURCE_PAYROLL,
                        payroll_month=month,
                        payroll_year=year,
                    )
                    expense.evidence_document.save(
                        f'payroll_{year}_{month.lower()}_{school.id}.pdf',
                        ContentFile(evidence_bytes),
                        save=False,
                    )
                    expense.save()
            if updated:
                messages.success(request, f'Payroll for {month} {year} marked as paid and locked.')
            else:
                messages.info(request, f'Payroll for {month} {year} is already marked paid.')
            return redirect(f"{request.path}?filter_month={month}&filter_year={year}")

        elif action == 'update_payroll':
            record_id = request.POST.get('record_id')
            selected_record = records.filter(id=record_id).first()
            if not selected_record:
                messages.error(request, 'Selected payroll record was not found.')
                return redirect('payroll_overview')
            if selected_record.is_paid:
                messages.error(request, 'This payroll record is paid and locked.')
                return redirect('payroll_overview')

            payroll_form = PayrollRecordForm(request.POST)
            if payroll_form.is_valid():
                allowance_entries, allowance_error = _parse_allowances(request)
                other_entries, other_error = _parse_other_deductions(request)
                if allowance_error:
                    messages.error(request, allowance_error)
                elif other_error:
                    messages.error(request, other_error)
                else:
                    with transaction.atomic():
                        selected_record.month = payroll_form.cleaned_data['month']
                        selected_record.year = payroll_form.cleaned_data['year']
                        payroll_form.apply_statutory_values(selected_record, selected_record.staff)
                        selected_record.save()

                        selected_record.allowances_rows.all().delete()
                        selected_record.other_deductions.all().delete()

                        total_allowances = Decimal('0')
                        for entry in allowance_entries:
                            PayrollAllowance.objects.create(
                                payroll_record=selected_record,
                                name=entry['name'],
                                amount=entry['amount'],
                            )
                            total_allowances += entry['amount']

                        total_other = Decimal('0')
                        for entry in other_entries:
                            PayrollOtherDeduction.objects.create(
                                payroll_record=selected_record,
                                name=entry['name'],
                                amount=entry['amount'],
                            )
                            total_other += entry['amount']

                        selected_record.allowances = total_allowances
                        selected_record.deductions = total_other
                        selected_record.save(update_fields=['allowances', 'deductions'])

                    messages.success(request, 'Payroll record updated successfully.')
                    return redirect('payroll_overview')
            else:
                messages.error(request, 'Please correct payroll form errors.')
        else:
            messages.error(request, 'Unknown action.')
            payroll_form = PayrollRecordForm(initial={'year': date.today().year})
    else:
        payroll_form = PayrollRecordForm(initial={'year': date.today().year})

    if selected_record and request.method != 'POST':
        basic_salary = selected_record.staff.basic_salary or Decimal('0')
        payable_basic = _payable_basic_salary(basic_salary, selected_record.days_worked)
        payroll_form = PayrollRecordForm(initial={
            'month': selected_record.month,
            'year': selected_record.year,
            'days_worked': selected_record.days_worked,
            'paye_rate': _safe_rate(selected_record.paye_deduction, payable_basic),
            'nssf_rate': _safe_rate(selected_record.nssf_deduction, payable_basic),
            'nhif_rate': _safe_rate(selected_record.nhif_deduction, payable_basic),
            'housing_levy_rate': _safe_rate(selected_record.housing_levy_deduction, payable_basic),
        })

    selected_allowances = []
    selected_deductions = []
    if selected_record:
        selected_allowances_qs = list(selected_record.allowances_rows.all())
        selected_deductions_qs = list(selected_record.other_deductions.all())
        selected_allowances = [{'name': row.name, 'amount': str(row.amount)} for row in selected_allowances_qs]
        selected_deductions = [{'name': row.name, 'amount': str(row.amount)} for row in selected_deductions_qs]

    context = {
        'payroll_form': payroll_form,
        'staff_list': staff,
        'record_list': records[:30],
        'staff_count': staff.count(),
        'record_count': records.count(),
        'period_paid': period_paid,
        'selected_record': selected_record,
        'staff_salary_map': {str(member.id): str(member.basic_salary) for member in staff},
        'month_options': PayrollRecordForm.MONTH_CHOICES,
        'current_year': date.today().year,
        'filter_month': filter_month,
        'filter_year': filter_year_raw,
        'selected_allowances': selected_allowances,
        'selected_deductions': selected_deductions,
    }
    return render(request, 'payroll/overview.html', context)


@login_required
def generate_payslip(request, record_id):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can generate payslips.')

    record = get_object_or_404(PayrollRecord, id=record_id, staff__school=school)
    staff = record.staff
    allowance_rows = list(record.allowances_rows.all())
    allowance_total = sum((row.amount for row in allowance_rows), Decimal('0')) if allowance_rows else (record.allowances or Decimal('0'))
    other_rows = list(record.other_deductions.all())
    other_total = sum((row.amount for row in other_rows), Decimal('0')) if other_rows else (record.deductions or Decimal('0'))
    payable_basic = record.payable_basic_salary()
    payable_percent = (Decimal(record.days_worked or 0) / Decimal('30')) * Decimal('100')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payslip_{staff.full_name}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    left = 45
    right = width - 45
    y = height - 42

    def money(value):
        try:
            amount = Decimal(value or 0)
        except Exception:
            amount = Decimal('0')
        return f"{amount:,.2f}"

    # Header with school details and logo
    logo = _image_reader_from_field(school.logo)
    if logo:
        try:
            p.drawImage(logo, left, y - 46, width=52, height=52, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    center_x = width / 2
    p.setFont('Helvetica-Bold', 16)
    p.drawCentredString(center_x, y, (school.name or '').upper())
    if getattr(school, 'motto', ''):
        p.setFont('Helvetica-Oblique', 10)
        p.drawCentredString(center_x, y - 14, str(school.motto).upper()[:120])
    contact_line = " | ".join([str(v).upper() for v in [school.phone, school.email] if v])
    if contact_line:
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 28, contact_line[:120])
        text_w = p.stringWidth(contact_line[:120], 'Helvetica', 10)
        p.line(center_x - text_w / 2, y - 30, center_x + text_w / 2, y - 30)
    if getattr(school, 'address', ''):
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 42, str(school.address).upper()[:120])

    p.setFont('Helvetica-Bold', 15)
    p.drawRightString(right, y, 'PAYSLIP')
    p.setFont('Helvetica', 10)
    p.drawRightString(right, y - 14, f'{record.month} {record.year}')

    y -= 74
    p.line(left, y, right, y)
    y -= 16

    # Staff details
    p.setFont('Helvetica', 10)
    p.drawString(left, y, f'Staff Name: {staff.full_name}')
    p.drawString(left + 280, y, f'Role: {staff.role}')
    y -= 14
    p.drawString(left, y, f'Employee Number: {staff.employee_number or "-"}')
    p.drawString(left + 280, y, f'KRA PIN: {staff.kra_pin or "-"}')
    y -= 14
    p.drawString(left, y, f'Days Worked: {record.days_worked}/30')
    p.drawString(left + 280, y, f'Payable: {payable_percent:,.2f}%')
    y -= 22

    # Tabulated section (no grid lines)
    p.setFont('Helvetica-Bold', 11)
    p.drawString(left, y, 'Earnings')
    p.drawString(left + 295, y, 'Deductions')
    y -= 12
    p.line(left, y, right, y)
    y -= 12

    earnings = [
        ('Basic Salary (Full)', staff.basic_salary),
        ('Payable Basic Salary', payable_basic),
    ]
    for row in allowance_rows:
        earnings.append((row.name, row.amount))
    earnings.append(('Total Allowances', allowance_total))

    deductions = [
        ('PAYE', record.paye_deduction),
        ('NSSF', record.nssf_deduction),
        ('NHIF/SHIF', record.nhif_deduction),
        ('Housing Levy', record.housing_levy_deduction),
    ]
    for row in other_rows:
        deductions.append((row.name, row.amount))
    deductions.append(('Total Deductions', record.total_deductions()))

    max_rows = max(len(earnings), len(deductions))
    p.setFont('Helvetica', 10)
    for idx in range(max_rows):
        left_name, left_value = earnings[idx] if idx < len(earnings) else ('', '')
        right_name, right_value = deductions[idx] if idx < len(deductions) else ('', '')

        if left_name:
            p.drawString(left, y, str(left_name))
            p.drawRightString(left + 250, y, money(left_value))
        if right_name:
            p.drawString(left + 295, y, str(right_name))
            p.drawRightString(right, y, money(right_value))

        y -= 14
        if y < 120:
            p.showPage()
            y = height - 60
            p.setFont('Helvetica', 10)

    y -= 6
    p.line(left, y, right, y)
    y -= 20
    p.setFont('Helvetica-Bold', 12)
    p.drawString(left, y, 'Net Salary')
    p.drawRightString(right, y, money(record.net_salary()))

    # Footer: stamp and signature
    y -= 60
    stamp_x = left
    sign_x = left + 300
    p.setFont('Helvetica-Bold', 10)
    p.drawString(stamp_x, y, 'School Stamp')
    p.drawString(sign_x, y, 'Authorized Signature')

    y -= 48
    stamp_drawn = False
    stamp_img = _image_reader_from_field(school.stamp)
    if stamp_img:
        try:
            p.drawImage(stamp_img, stamp_x, y - 8, width=95, height=52, preserveAspectRatio=True, mask='auto')
            stamp_drawn = True
        except Exception:
            stamp_drawn = False
    if not stamp_drawn:
        p.line(stamp_x, y + 4, stamp_x + 120, y + 4)

    sign_drawn = False
    sign_img = _image_reader_from_field(school.head_signature)
    if sign_img:
        try:
            p.drawImage(sign_img, sign_x, y - 8, width=130, height=52, preserveAspectRatio=True, mask='auto')
            sign_drawn = True
        except Exception:
            sign_drawn = False
    if not sign_drawn:
        p.line(sign_x, y + 4, sign_x + 150, y + 4)

    p.showPage()
    p.save()
    return response


@login_required
def generate_p9(request, record_id):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can generate P9 forms.')

    december_record = get_object_or_404(PayrollRecord, id=record_id, staff__school=school)
    if (december_record.month or '').lower() != 'december':
        return HttpResponseForbidden('P9 can only be generated from a December payroll row.')

    staff = december_record.staff
    year = december_record.year
    annual_records = (
        PayrollRecord.objects
        .filter(staff=staff, year=year)
        .order_by('id')
        .prefetch_related('allowances_rows', 'other_deductions')
    )
    month_order = [value for value, _ in PayrollRecordForm.MONTH_CHOICES]
    month_summary = {
        month: {
            'payable_basic': Decimal('0'),
            'allowances': Decimal('0'),
            'taxable': Decimal('0'),
            'paye': Decimal('0'),
            'net': Decimal('0'),
        }
        for month in month_order
    }

    for rec in annual_records:
        month_key = rec.month
        if month_key not in month_summary:
            continue
        payable_basic = Decimal(rec.payable_basic_salary() or 0)
        allowances = Decimal(rec.total_allowances() or 0)
        taxable = payable_basic + allowances
        paye = Decimal(rec.paye_deduction or 0)
        net = Decimal(rec.net_salary() or 0)
        month_summary[month_key]['payable_basic'] += payable_basic
        month_summary[month_key]['allowances'] += allowances
        month_summary[month_key]['taxable'] += taxable
        month_summary[month_key]['paye'] += paye
        month_summary[month_key]['net'] += net

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="p9_{staff.full_name}_{year}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    left = 45
    right = width - 45
    y = height - 42

    # Header
    logo = _image_reader_from_field(school.logo)
    if logo:
        try:
            p.drawImage(logo, left, y - 46, width=52, height=52, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    center_x = width / 2
    p.setFont('Helvetica-Bold', 16)
    p.drawCentredString(center_x, y, (school.name or '').upper())
    if getattr(school, 'motto', ''):
        p.setFont('Helvetica-Oblique', 10)
        p.drawCentredString(center_x, y - 14, str(school.motto).upper()[:120])
    contact_line = " | ".join([str(v).upper() for v in [school.phone, school.email] if v])
    if contact_line:
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 28, contact_line[:120])
        text_w = p.stringWidth(contact_line[:120], 'Helvetica', 10)
        p.line(center_x - text_w / 2, y - 30, center_x + text_w / 2, y - 30)
    if getattr(school, 'address', ''):
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 42, str(school.address).upper()[:120])
    p.setFont('Helvetica-Bold', 13)
    p.drawRightString(right, y, f'P9 FORM - {year}')
    p.setFont('Helvetica', 10)
    p.drawString(left, y - 16, f'Staff: {staff.full_name}')
    p.drawString(left, y - 30, f'KRA PIN: {staff.kra_pin or "-"}')
    y -= 62
    p.line(left, y, right, y)
    y -= 14

    # Table headings
    p.setFont('Helvetica-Bold', 9)
    p.drawString(left, y, 'Month')
    p.drawString(left + 95, y, 'Payable Basic')
    p.drawString(left + 200, y, 'Allowances')
    p.drawString(left + 295, y, 'Taxable Pay')
    p.drawString(left + 390, y, 'PAYE')
    p.drawString(left + 470, y, 'Net Pay')
    y -= 10
    p.line(left, y, right, y)
    y -= 12

    total_payable_basic = Decimal('0')
    total_allowances = Decimal('0')
    total_taxable = Decimal('0')
    total_paye = Decimal('0')
    total_net = Decimal('0')

    p.setFont('Helvetica', 9)
    for month_name in month_order:
        if y < 70:
            p.showPage()
            y = height - 50
            p.setFont('Helvetica-Bold', 9)
            p.drawString(left, y, 'Month')
            p.drawString(left + 95, y, 'Payable Basic')
            p.drawString(left + 200, y, 'Allowances')
            p.drawString(left + 295, y, 'Taxable Pay')
            p.drawString(left + 390, y, 'PAYE')
            p.drawString(left + 470, y, 'Net Pay')
            y -= 10
            p.line(left, y, right, y)
            y -= 12
            p.setFont('Helvetica', 9)

        month_values = month_summary.get(month_name, {})
        payable_basic = Decimal(month_values.get('payable_basic') or 0)
        allowances = Decimal(month_values.get('allowances') or 0)
        taxable = Decimal(month_values.get('taxable') or 0)
        paye = Decimal(month_values.get('paye') or 0)
        net = Decimal(month_values.get('net') or 0)

        p.drawString(left, y, month_name)
        p.drawRightString(left + 185, y, f'{payable_basic:,.2f}')
        p.drawRightString(left + 280, y, f'{allowances:,.2f}')
        p.drawRightString(left + 375, y, f'{taxable:,.2f}')
        p.drawRightString(left + 455, y, f'{paye:,.2f}')
        p.drawRightString(right, y, f'{net:,.2f}')
        y -= 12

        total_payable_basic += payable_basic
        total_allowances += allowances
        total_taxable += taxable
        total_paye += paye
        total_net += net

    y -= 4
    p.line(left, y, right, y)
    y -= 14
    p.setFont('Helvetica-Bold', 9)
    p.drawString(left, y, 'TOTAL')
    p.drawRightString(left + 185, y, f'{total_payable_basic:,.2f}')
    p.drawRightString(left + 280, y, f'{total_allowances:,.2f}')
    p.drawRightString(left + 375, y, f'{total_taxable:,.2f}')
    p.drawRightString(left + 455, y, f'{total_paye:,.2f}')
    p.drawRightString(right, y, f'{total_net:,.2f}')
    y -= 14
    p.drawString(left, y, 'TOTAL PAYE')
    p.drawRightString(left + 455, y, f'{total_paye:,.2f}')

    p.showPage()
    p.save()
    return response


@login_required
@require_POST
def delete_payroll_record(request, record_id):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can manage payroll.')

    record = get_object_or_404(PayrollRecord, id=record_id, staff__school=school)
    if record.is_paid:
        messages.error(request, 'This payroll record is paid and cannot be deleted.')
        return redirect('payroll_overview')
    record.delete()
    messages.success(request, 'Payroll record deleted successfully.')
    return redirect('payroll_overview')


@login_required
def export_payroll_excel(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can export payroll.')

    records_base = PayrollRecord.objects.select_related('staff').filter(staff__school=school).order_by('-year', '-id')
    export_month, export_year = _resolve_export_period(records_base, request)
    records = records_base
    if export_month and export_year is not None:
        records = records.filter(month=export_month, year=export_year)
    records = records.prefetch_related('allowances_rows', 'other_deductions')

    allowance_names = sorted({
        n for n in PayrollAllowance.objects.filter(payroll_record__in=records).values_list('name', flat=True) if n
    })
    deduction_names = sorted({
        n for n in PayrollOtherDeduction.objects.filter(payroll_record__in=records).values_list('name', flat=True) if n
    })

    wb = Workbook()
    ws = wb.active
    ws.title = 'Payroll'

    period_label = f'{export_month} {export_year}' if export_month and export_year is not None else 'All Periods'
    ws.append([f'{(school.name or "").upper()} PAYROLL REPORT'])
    ws.append([f'Payroll Month: {period_label}'])
    ws.append([])

    headers = [
        'Staff',
        'Role',
        'Days Worked',
        'Basic Salary (Full)',
        'Payable Basic Salary',
    ]
    headers.extend([f'Allowance: {name}' for name in allowance_names])
    headers.extend([
        'Total Allowances',
        'PAYE',
        'NSSF',
        'NHIF/SHIF',
        'Housing Levy',
    ])
    headers.extend([f'Deduction: {name}' for name in deduction_names])
    headers.extend([
        'Total Deductions',
        'Net Salary',
    ])
    ws.append(headers)
    header_row = 4
    data_start_row = 5

    # Column index groups (1-based for openpyxl)
    allowance_start_col = 6  # after Staff, Role, Days, Basic, Payable
    allowance_end_col = allowance_start_col + len(allowance_names)  # includes total allowances
    deduction_start_col = allowance_end_col + 1  # PAYE starts
    deduction_end_col = deduction_start_col + 4 + len(deduction_names)  # includes total deductions

    green_font = Font(color='16A34A')
    red_font = Font(color='DC2626')

    # Color allowance/deduction headers
    for col_idx in range(allowance_start_col, allowance_end_col + 1):
        ws.cell(row=header_row, column=col_idx).font = green_font
    for col_idx in range(deduction_start_col, deduction_end_col + 1):
        ws.cell(row=header_row, column=col_idx).font = red_font

    total_full_basic = Decimal('0')
    total_payable_basic = Decimal('0')
    total_allowances = Decimal('0')
    total_paye = Decimal('0')
    total_nssf = Decimal('0')
    total_nhif = Decimal('0')
    total_housing = Decimal('0')
    total_deductions = Decimal('0')
    total_net = Decimal('0')
    totals_allowance_by_name = {name: Decimal('0') for name in allowance_names}
    totals_deduction_by_name = {name: Decimal('0') for name in deduction_names}

    row_count = 0
    for row_offset, rec in enumerate(records):
        allowance_map = {row.name: row.amount for row in rec.allowances_rows.all()}
        deduction_map = {row.name: row.amount for row in rec.other_deductions.all()}

        row = [
            rec.staff.full_name,
            rec.staff.role,
            rec.days_worked,
            float(rec.staff.basic_salary or 0),
            float(rec.payable_basic_salary() or 0),
        ]
        row.extend([float(allowance_map.get(name, 0) or 0) for name in allowance_names])
        row.extend([
            float(rec.total_allowances() or 0),
            float(rec.paye_deduction or 0),
            float(rec.nssf_deduction or 0),
            float(rec.nhif_deduction or 0),
            float(rec.housing_levy_deduction or 0),
        ])
        row.extend([float(deduction_map.get(name, 0) or 0) for name in deduction_names])
        row.extend([
            float(rec.total_deductions() or 0),
            float(rec.net_salary() or 0),
        ])

        ws.append([
            *row
        ])
        current_row = data_start_row + row_offset
        for col_idx in range(allowance_start_col, allowance_end_col + 1):
            ws.cell(row=current_row, column=col_idx).font = green_font
        for col_idx in range(deduction_start_col, deduction_end_col + 1):
            ws.cell(row=current_row, column=col_idx).font = red_font

        row_count += 1
        total_full_basic += Decimal(rec.staff.basic_salary or 0)
        total_payable_basic += Decimal(rec.payable_basic_salary() or 0)
        total_allowances += Decimal(rec.total_allowances() or 0)
        total_paye += Decimal(rec.paye_deduction or 0)
        total_nssf += Decimal(rec.nssf_deduction or 0)
        total_nhif += Decimal(rec.nhif_deduction or 0)
        total_housing += Decimal(rec.housing_levy_deduction or 0)
        total_deductions += Decimal(rec.total_deductions() or 0)
        total_net += Decimal(rec.net_salary() or 0)
        for name in allowance_names:
            totals_allowance_by_name[name] += Decimal(allowance_map.get(name, 0) or 0)
        for name in deduction_names:
            totals_deduction_by_name[name] += Decimal(deduction_map.get(name, 0) or 0)

    if row_count:
        total_row = [
            'TOTAL',
            '',
            '',
            float(total_full_basic),
            float(total_payable_basic),
        ]
        total_row.extend([float(totals_allowance_by_name[name]) for name in allowance_names])
        total_row.extend([
            float(total_allowances),
            float(total_paye),
            float(total_nssf),
            float(total_nhif),
            float(total_housing),
        ])
        total_row.extend([float(totals_deduction_by_name[name]) for name in deduction_names])
        total_row.extend([
            float(total_deductions),
            float(total_net),
        ])
        ws.append(total_row)
        total_excel_row = data_start_row + row_count
        ws.cell(row=total_excel_row, column=1).font = Font(bold=True)
        for col_idx in range(allowance_start_col, allowance_end_col + 1):
            ws.cell(row=total_excel_row, column=col_idx).font = Font(color='16A34A', bold=True)
        for col_idx in range(deduction_start_col, deduction_end_col + 1):
            ws.cell(row=total_excel_row, column=col_idx).font = Font(color='DC2626', bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="payroll_{school.name}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_payroll_pdf(request):
    school = _headteacher_school(request)
    if not school:
        return HttpResponseForbidden('Only headteachers can export payroll.')

    records_base = PayrollRecord.objects.select_related('staff').filter(staff__school=school).order_by('-year', '-id')
    export_month, export_year = _resolve_export_period(records_base, request)
    records = records_base
    if export_month and export_year is not None:
        records = records.filter(month=export_month, year=export_year)
    records = records.prefetch_related('allowances_rows', 'other_deductions')

    allowance_names = sorted({
        n for n in PayrollAllowance.objects.filter(payroll_record__in=records).values_list('name', flat=True) if n
    })
    deduction_names = sorted({
        n for n in PayrollOtherDeduction.objects.filter(payroll_record__in=records).values_list('name', flat=True) if n
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payroll_{school.name}.pdf"'

    p = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = A4
    width, height = height, width
    left = 40
    y = height - 42
    period_label = f'{export_month} {export_year}' if export_month and export_year is not None else 'All Periods'

    logo = _image_reader_from_field(school.logo)
    if logo:
        try:
            p.drawImage(logo, left, y - 36, width=42, height=42, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    center_x = width / 2
    p.setFont('Helvetica-Bold', 14)
    p.drawCentredString(center_x, y, (school.name or '').upper())
    if getattr(school, 'motto', ''):
        p.setFont('Helvetica-Oblique', 10)
        p.drawCentredString(center_x, y - 14, str(school.motto).upper()[:120])
    contact_line = " | ".join([str(v).upper() for v in [school.phone, school.email] if v])
    if contact_line:
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 28, contact_line[:120])
        text_w = p.stringWidth(contact_line[:120], 'Helvetica', 10)
        p.line(center_x - text_w / 2, y - 30, center_x + text_w / 2, y - 30)
    if getattr(school, 'address', ''):
        p.setFont('Helvetica', 10)
        p.drawCentredString(center_x, y - 42, str(school.address).upper()[:120])
    p.setFont('Helvetica', 10)
    p.drawCentredString(center_x, y - 56, f'PAYROLL REPORT - {period_label}')
    p.drawCentredString(center_x, y - 70, f'TOTAL RECORDS: {records.count()}')
    y -= 82

    headers = [
        'Staff', 'Role', 'Days', 'Basic', 'Payable',
    ]
    headers.extend([n[:10] for n in allowance_names])
    headers.extend(['AllowTot', 'PAYE', 'NSSF', 'NHIF', 'House'])
    headers.extend([n[:10] for n in deduction_names])
    headers.extend(['DedTot', 'Net'])

    table_width = width - (left * 2)
    col_count = max(1, len(headers))
    col_width = table_width / col_count

    def draw_header_row(curr_y):
        p.setFont('Helvetica-Bold', 7)
        x = left
        for head in headers:
            p.drawString(x + 1, curr_y, head)
            x += col_width
        p.line(left, curr_y - 2, width - left, curr_y - 2)
        return curr_y - 12

    y = draw_header_row(y)
    p.setFont('Helvetica', 7)
    allowance_start_idx = 5
    allowance_end_idx = allowance_start_idx + len(allowance_names) + 1  # includes total allowances
    deduction_start_idx = allowance_end_idx  # starts at PAYE
    deduction_end_idx = deduction_start_idx + 4 + len(deduction_names) + 1  # statutory + dynamic + total deductions

    total_full_basic = Decimal('0')
    total_payable_basic = Decimal('0')
    total_allowances = Decimal('0')
    total_paye = Decimal('0')
    total_nssf = Decimal('0')
    total_nhif = Decimal('0')
    total_housing = Decimal('0')
    total_deductions = Decimal('0')
    total_net = Decimal('0')
    totals_allowance_by_name = {name: Decimal('0') for name in allowance_names}
    totals_deduction_by_name = {name: Decimal('0') for name in deduction_names}
    row_count = 0

    for rec in records:
        if y < 45:
            p.showPage()
            y = height - 45
            y = draw_header_row(y)
            p.setFont('Helvetica', 7)

        allowance_map = {row.name: row.amount for row in rec.allowances_rows.all()}
        deduction_map = {row.name: row.amount for row in rec.other_deductions.all()}

        values = [
            rec.staff.full_name[:16],
            rec.staff.role[:12],
            str(rec.days_worked),
            f'{Decimal(rec.staff.basic_salary or 0):,.0f}',
            f'{Decimal(rec.payable_basic_salary() or 0):,.0f}',
        ]
        values.extend([f'{Decimal(allowance_map.get(name, 0) or 0):,.0f}' for name in allowance_names])
        values.extend([
            f'{Decimal(rec.total_allowances() or 0):,.0f}',
            f'{Decimal(rec.paye_deduction or 0):,.0f}',
            f'{Decimal(rec.nssf_deduction or 0):,.0f}',
            f'{Decimal(rec.nhif_deduction or 0):,.0f}',
            f'{Decimal(rec.housing_levy_deduction or 0):,.0f}',
        ])
        values.extend([f'{Decimal(deduction_map.get(name, 0) or 0):,.0f}' for name in deduction_names])
        values.extend([
            f'{Decimal(rec.total_deductions() or 0):,.0f}',
            f'{Decimal(rec.net_salary() or 0):,.0f}',
        ])

        x = left
        for idx, value in enumerate(values):
            if allowance_start_idx <= idx < allowance_end_idx:
                p.setFillColorRGB(0.086, 0.639, 0.290)  # green
            elif deduction_start_idx <= idx < deduction_end_idx:
                p.setFillColorRGB(0.863, 0.149, 0.149)  # red
            else:
                p.setFillColorRGB(0, 0, 0)
            p.drawString(x + 1, y, str(value))
            x += col_width
        p.setFillColorRGB(0, 0, 0)
        y -= 11
        row_count += 1

        total_full_basic += Decimal(rec.staff.basic_salary or 0)
        total_payable_basic += Decimal(rec.payable_basic_salary() or 0)
        total_allowances += Decimal(rec.total_allowances() or 0)
        total_paye += Decimal(rec.paye_deduction or 0)
        total_nssf += Decimal(rec.nssf_deduction or 0)
        total_nhif += Decimal(rec.nhif_deduction or 0)
        total_housing += Decimal(rec.housing_levy_deduction or 0)
        total_deductions += Decimal(rec.total_deductions() or 0)
        total_net += Decimal(rec.net_salary() or 0)
        for name in allowance_names:
            totals_allowance_by_name[name] += Decimal(allowance_map.get(name, 0) or 0)
        for name in deduction_names:
            totals_deduction_by_name[name] += Decimal(deduction_map.get(name, 0) or 0)

    if row_count:
        if y < 45:
            p.showPage()
            y = height - 45
            y = draw_header_row(y)
        p.setFont('Helvetica-Bold', 7)
        total_values = [
            'TOTAL',
            '',
            '',
            f'{total_full_basic:,.0f}',
            f'{total_payable_basic:,.0f}',
        ]
        total_values.extend([f'{totals_allowance_by_name[name]:,.0f}' for name in allowance_names])
        total_values.extend([
            f'{total_allowances:,.0f}',
            f'{total_paye:,.0f}',
            f'{total_nssf:,.0f}',
            f'{total_nhif:,.0f}',
            f'{total_housing:,.0f}',
        ])
        total_values.extend([f'{totals_deduction_by_name[name]:,.0f}' for name in deduction_names])
        total_values.extend([
            f'{total_deductions:,.0f}',
            f'{total_net:,.0f}',
        ])
        x = left
        for idx, value in enumerate(total_values):
            if allowance_start_idx <= idx < allowance_end_idx:
                p.setFillColorRGB(0.086, 0.639, 0.290)
            elif deduction_start_idx <= idx < deduction_end_idx:
                p.setFillColorRGB(0.863, 0.149, 0.149)
            else:
                p.setFillColorRGB(0, 0, 0)
            p.drawString(x + 1, y, str(value))
            x += col_width
        p.setFillColorRGB(0, 0, 0)

    p.showPage()
    p.save()
    return response
def _image_reader_from_field(field):
    if not field:
        return None
    try:
        if hasattr(field, 'path'):
            return ImageReader(field.path)
    except Exception:
        pass
    try:
        if hasattr(field, 'url'):
            return ImageReader(field.url)
    except Exception:
        return None
